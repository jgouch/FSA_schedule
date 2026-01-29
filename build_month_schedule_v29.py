#!/usr/bin/env python3
"""
build_month_schedule_v29.py

End-to-end month builder:
- Reads TimeOff.xlsx (monthly request sheet + Sales Ranking tab)
- Reads prior month schedule JSON for cross-month constraints
- Generates schedule for target month (primaries then BU)
- Exports:
    - schedule JSON (machine truth)
    - Excel calendar workbook (Calendar, Summary, Weekly Hours, Pay Period Hours, Violations)

LOCKED CALENDAR GEOMETRY:
- Grid weeks: Dynamic (5 or 6) based on month shape.
- Each day block: 3 columns x 8 rows
- Role mapping (R = week_start_row, cols = start/mid/end):
    PN label: (start, R+1), PN name: (mid, R+1)
    AN label: (start, R+2), AN name: (mid, R+2)
    W  label: (start, R+3), W  name: (mid, R+3)
    BU1 name: (mid, R+4)
    BU2 name: (mid, R+5)
    BU3 name: (mid, R+6)
    BU4 name (push-week Saturday only): (mid, R+7)

Logic Changes (v29):
- REFACTOR: Shared 'calculate_daily_hours' helper for both Export and Audit.
- FIXED: Audit wording ("Avoid Warning", "Limit 40/80").
- CLEANUP: Deleted unused MAX_WEEKS constant.
- CLARITY: Renamed B2B violation to "Back-to-Back (Same Role)".
"""

from __future__ import annotations

import argparse
import calendar
import json
import random
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


# ----------------------------
# Constants / settings
# ----------------------------

SUNDAY_FILL = "FFF2CC"
NONMONTH_FILL = "F2F2F2"
WHITE_FILL_HEX = "FFFFFF"
OVERTIME_FILL = "FFC7CE" # Red fill for >40h weeks

FONT_MAIN = Font(name="Calibri", size=11)
FONT_DAY_NUM = Font(name="Calibri", size=11, bold=True)
FONT_HEADER = Font(name="Calibri", size=24, bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

GRID_WEEK_START_ROW = 3
WEEK_HEIGHT = 8
DAY_WIDTH = 3
SUN_START_COL = 2  # B

ROLE_PRIMARY = ["PN", "AN", "W"]
ROLE_BUS = ["BU1", "BU2", "BU3", "BU4"]

OBSERVED_HOLIDAYS_BY_YEAR = {
    2026: {
        date(2026, 1, 1),   # New Year's Day
        date(2026, 5, 25),  # Memorial Day
        date(2026, 7, 3),   # Independence Day observed
        date(2026, 9, 7),   # Labor Day
        date(2026, 11, 26), # Thanksgiving
        date(2026, 12, 25), # Christmas
    }
}

DEFAULT_ROSTER = ["Greg", "Mark", "Dawn", "Will", "CJ", "Kyle"]
SENIORITY_ORDER = ["Mark", "Dawn", "Will", "Kyle", "CJ", "Greg"]


# ----------------------------
# Helpers
# ----------------------------

def weekday_sun0(d: date) -> int:
    return (d.weekday() + 1) % 7

def is_sunday(d: date) -> bool:
    return weekday_sun0(d) == 0

def is_saturday(d: date) -> bool:
    return weekday_sun0(d) == 6

def is_weekday(d: date) -> bool:
    return d.weekday() < 5

def month_days(year: int, month: int) -> List[date]:
    n = calendar.monthrange(year, month)[1]
    return [date(year, month, d) for d in range(1, n+1)]

def quarter_tag_for_month(year: int, month: int) -> str:
    q = (month - 1) // 3 + 1
    return f"{year}Q{q}"

def norm_name(x) -> str:
    return re.sub(r"\s+", " ", str(x).strip())

def normalize_quarter_key(s: str) -> str:
    """Normalize '2026 Q1' -> '2026Q1'."""
    return re.sub(r"\s+", "", str(s).strip().upper())

def parse_month_arg(s: str) -> Tuple[int, int]:
    m = re.match(r"^\s*(\d{4})-(\d{2})\s*$", s)
    if not m:
        raise ValueError("Month must be YYYY-MM, e.g. 2026-02")
    return int(m.group(1)), int(m.group(2))

def compute_observed_holidays_us(year: int) -> Set[date]:
    out: Set[date] = set()
    out.add(date(year, 1, 1))
    d = date(year, 5, 31)
    while d.weekday() != 0: d -= timedelta(days=1)
    out.add(d)
    july4 = date(year, 7, 4)
    if july4.weekday() == 5: out.add(date(year, 7, 3))
    elif july4.weekday() == 6: out.add(date(year, 7, 5))
    else: out.add(july4)
    d = date(year, 9, 1)
    while d.weekday() != 0: d += timedelta(days=1)
    out.add(d)
    d = date(year, 11, 1)
    while d.weekday() != 3: d += timedelta(days=1)
    out.add(d + timedelta(days=21))
    out.add(date(year, 12, 25))
    return out

def observed_holidays_for_year(year: int) -> Set[date]:
    return OBSERVED_HOLIDAYS_BY_YEAR.get(year, compute_observed_holidays_us(year))

def week_start_sun(d: date) -> date:
    return d - timedelta(days=weekday_sun0(d))

def generate_payperiods(anchor_start: date, start_date: date, end_date: date) -> List[Tuple[date, date]]:
    a = anchor_start
    while a > start_date:
        a -= timedelta(days=14)
    periods: List[Tuple[date, date]] = []
    cur = a
    while cur <= end_date:
        periods.append((cur, cur + timedelta(days=13)))
        cur += timedelta(days=14)
    return periods

def parse_excel_date(value) -> date:
    if isinstance(value, datetime): return value.date()
    if isinstance(value, date): return value
    if isinstance(value, (int, float)):
        try: return from_excel(value).date()
        except Exception as e: 
            raise ValueError(f"Failed to parse numeric Excel date: {value}. Error: {e}")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try: return datetime.strptime(s, fmt).date()
        except: continue
    raise ValueError(f"Unrecognized date format: {value!r}")

def _rotate_left(lst: List[str], k: int) -> List[str]:
    if not lst: return lst
    k = k % len(lst)
    return lst[k:] + lst[:k]

# Global logic checks
def week_contains_push(ws: date, push_days: Set[date]) -> bool:
    for i in range(7):
        if (ws + timedelta(days=i)) in push_days: return True
    return False

def week_contains_holiday(ws: date, hols: Set[date]) -> bool:
    for i in range(7):
        if (ws + timedelta(days=i)) in hols: return True
    return False

# Global hours calculator (Single Source of Truth)
def calculate_daily_hours(dd: date, name: str, schedule: Dict, prev_sched: Dict, year: int, month: int, hols: Set[date]) -> int:
    """Calculate hours for a person on a specific date, handling month boundaries."""
    if dd.year == year and dd.month == month:
        roles = schedule.get(dd.isoformat(), {})
        dh = 0
        for r, n in roles.items():
            if n == name: dh = max(dh, hours_for_role(dd, r, hols))
        return dh
    elif dd < date(year, month, 1):
        roles = prev_sched.get(dd.isoformat(), {})
        dh = 0
        for r, n in roles.items():
            if n == name: 
                # Use global observed holidays logic for previous year dates if needed
                prev_hols = observed_holidays_for_year(dd.year)
                dh = max(dh, hours_for_role(dd, r, prev_hols))
        return dh
    return 0


# ----------------------------
# Inputs
# ----------------------------

@dataclass
class TimeOffRule:
    name: str
    d: date
    hard: bool = True
    avoid_roles: Set[str] = field(default_factory=set)

@dataclass
class Constraints:
    hard_off: Dict[date, Set[str]]
    avoid_roles: Dict[Tuple[date, str], Set[str]]

def compile_constraints(timeoff: List[TimeOffRule]) -> Constraints:
    hard_off: Dict[date, Set[str]] = {}
    avoid_roles: Dict[Tuple[date, str], Set[str]] = {}
    for r in timeoff:
        if r.hard:
            hard_off.setdefault(r.d, set()).add(r.name)
        if r.avoid_roles:
            for role in r.avoid_roles:
                avoid_roles.setdefault((r.d, role.upper()), set()).add(r.name)
    return Constraints(hard_off=hard_off, avoid_roles=avoid_roles)

def load_timeoff_from_xlsx(xlsx_path: str, sheet_name: str) -> List[TimeOffRule]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing timeoff sheet '{sheet_name}'")
    ws = wb[sheet_name]
    
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v: headers[str(v).strip().lower()] = c

    c_name = headers.get("name")
    if not c_name: raise ValueError("TimeOff sheet missing 'Name' column")
    c_date = headers.get("date")
    c_start = headers.get("start")
    c_end = headers.get("end")
    c_hard = headers.get("hard")
    c_avoid = headers.get("avoidroles")

    def parse_hard(v) -> bool:
        if v is None or str(v).strip() == "": return True
        return str(v).strip().lower() in ("true", "1", "yes", "y")

    rules = []
    for r in range(2, ws.max_row + 1):
        nv = ws.cell(r, c_name).value
        if not nv: continue
        name = norm_name(nv)
        
        dates = []
        dv = ws.cell(r, c_date).value if c_date else None
        sv = ws.cell(r, c_start).value if c_start else None
        ev = ws.cell(r, c_end).value if c_end else None

        if dv:
            dates = [parse_excel_date(dv)]
        elif sv and ev:
            sd = parse_excel_date(sv)
            ed = parse_excel_date(ev)
            cur = sd
            while cur <= ed:
                dates.append(cur)
                cur += timedelta(days=1)
        else:
            continue

        hard = parse_hard(ws.cell(r, c_hard).value) if c_hard else True
        av = set()
        if c_avoid:
            val = ws.cell(r, c_avoid).value
            if val:
                av = {p.strip().upper() for p in str(val).split(",") if p.strip()}

        for d in dates:
            rules.append(TimeOffRule(name, d, hard, av))
    return rules

def load_roster_from_json(path: str) -> List[str]:
    data = json.loads(Path(path).read_text("utf-8"))
    if not isinstance(data, list) or not data:
        raise ValueError("Roster JSON must be a non-empty list of names")
    return [norm_name(x) for x in data if str(x).strip()]

def load_payperiods_from_json(path: str) -> List[Tuple[date, date]]:
    data = json.loads(Path(path).read_text("utf-8"))
    out = []
    for obj in data:
        s = datetime.strptime(str(obj["start"]).strip(), "%Y-%m-%d").date()
        e = datetime.strptime(str(obj["end"]).strip(), "%Y-%m-%d").date()
        if e < s:
            raise ValueError(f"Pay period end before start: {s}..{e}")
        out.append((s, e))
    out.sort(key=lambda t: t[0])
    return out

def load_sales_ranking_from_timeoff_xlsx(xlsx_path, sheet_name, year, month, roster_names):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing sheet '{sheet_name}'")
    ws = wb[sheet_name]
    target = normalize_quarter_key(quarter_tag_for_month(year, month))
    
    rows = {}
    for r in range(2, ws.max_row + 1):
        p = ws.cell(r, 1).value
        if not p: continue
        p = normalize_quarter_key(str(p)) 
        ranks = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and str(v).strip():
                ranks.append(norm_name(v))
        if ranks: rows[p] = ranks

    def key(x):
        m = re.match(r"^\s*(\d{4})\s*Q([1-4])\s*$", x)
        return (int(m.group(1)), int(m.group(2))) if m else (0,0)

    order = rows.get(target)
    if not order:
        candidates = sorted(rows.keys(), key=key)
        if candidates: order = rows[candidates[-1]]
    
    roster_set = set(roster_names)
    if not order:
        return [n for n in SENIORITY_ORDER if n in roster_set] + [n for n in roster_names if n not in SENIORITY_ORDER]
    
    order = [x for x in order if x in roster_set]
    missing = roster_set - set(order)
    order.extend(sorted(list(missing)))
    return order

def load_schedule_json(path: str) -> Dict[str, Dict[str, str]]:
    data = json.loads(Path(path).read_text("utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Schedule JSON must be a dictionary")
    iso_pat = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    for k, v in data.items():
        if not iso_pat.match(k):
             raise ValueError(f"Invalid date key in schedule JSON: {k} (Must be YYYY-MM-DD)")
        if not isinstance(v, dict):
             raise ValueError(f"Schedule JSON invalid at key {k}: value must be a dictionary")
    return data

def role_of(prev, d, name):
    roles = prev.get(d.isoformat(), {})
    for r, n in roles.items():
        if n == name: return r
    return None

def worked_prev(prev, d, name):
    return name in prev.get(d.isoformat(), {}).values()


# ----------------------------
# Core Logic
# ----------------------------

def roles_required_for_day(d: date, hols: Set[date]) -> List[str]:
    if d in hols: return []
    if is_sunday(d): return ["PN"]
    req = ["PN", "AN"]
    if is_weekday(d): req.append("W")
    return req

def hours_for_role(d: date, role: str, hols: Set[date]) -> int:
    if d in hols: return 0
    if is_sunday(d) and role == "PN": return 5
    return 8

def push_days_for_month(year: int, month: int) -> Set[date]:
    """Return set of dates that are 'Push Days' (Last occurrence of Mon-Sat)."""
    days = set()
    last_dom = calendar.monthrange(year, month)[1]
    last = date(year, month, last_dom)
    for wd in range(6): 
        cur = last
        while cur.weekday() != wd:
            cur -= timedelta(days=1)
        if cur.month == month:
            days.add(cur)
    return days

@dataclass
class BuildConfig:
    year: int
    month: int
    roster: List[str]
    sales_order: List[str]
    max_consecutive_days: int = 5
    lookback_days: int = 14
    rng_seed: int = 11

def compute_primary_targets(days, roster, sales_order, hols):
    opp = {"PN": 0, "AN": 0, "W": 0}
    for d in days:
        req = roles_required_for_day(d, hols)
        for r in opp:
            if r in req: opp[r] += 1
            
    n = len(roster)
    targets = {r: {p: 0 for p in roster} for r in opp}
    
    roster_set = set(roster)
    valid_order = [x for x in sales_order if x in roster_set]
    missing = [x for x in roster if x not in valid_order]
    full_priority = valid_order + missing
    
    def distribute(role, total):
        base = total // n
        rem = total % n
        for p in roster: targets[role][p] = base
        
        offset = 0
        rem_pn = opp["PN"] % n
        rem_an = opp["AN"] % n
        
        if role == "AN": offset = rem_pn
        if role == "W": offset = (rem_pn + rem_an) % n
        
        rotated_order = _rotate_left(full_priority, offset)
        for i in range(rem):
            targets[role][rotated_order[i]] += 1

    for r in opp: distribute(r, opp[r])
    return targets


def build_schedule(config: BuildConfig,
                   prev_sched: Dict[str, Dict[str, str]],
                   cons: Constraints,
                   payperiods: Optional[List[Tuple[date, date]]] = None) -> Dict[str, Dict[str, str]]:
    
    rng = random.Random(config.rng_seed)
    year, month = config.year, config.month
    days = month_days(year, month)
    
    hols = observed_holidays_for_year(year)
    push_days = push_days_for_month(year, month)
    targets = compute_primary_targets(days, config.roster, config.sales_order, hols)

    schedule: Dict[str, Dict[str, str]] = {}
    role_counts = {r: {n: 0 for n in config.roster} for r in ["PN", "AN", "W"]}
    total_hours = {n: 0 for n in config.roster}
    pn_weekday_counts = {n: {i: 0 for i in range(7)} for n in config.roster}
    monday_pn_used = {n: 0 for n in config.roster}

    def reset_state():
        nonlocal schedule, role_counts, total_hours, pn_weekday_counts, monday_pn_used
        schedule = {d.isoformat(): {} for d in days}
        role_counts = {r: {n: 0 for n in config.roster} for r in ["PN", "AN", "W"]}
        total_hours = {n: 0 for n in config.roster}
        pn_weekday_counts = {n: {i: 0 for i in range(7)} for n in config.roster}
        monday_pn_used = {n: 0 for n in config.roster}

    def last_role(name, d):
        for k in range(1, config.lookback_days + 1):
            p = d - timedelta(days=k)
            if p.year == year and p.month == month:
                roles = schedule.get(p.isoformat(), {})
                for r, n in roles.items():
                    if n == name: return r
            else:
                rr = role_of(prev_sched, p, name)
                if rr: return rr
        return None

    def consecutive_streak(name, d):
        streak = 0
        for k in range(1, config.lookback_days + 1):
            p = d - timedelta(days=k)
            worked = False
            if p.year == year and p.month == month:
                worked = name in schedule.get(p.isoformat(), {}).values()
            else:
                worked = worked_prev(prev_sched, p, name)
            if worked: streak += 1
            else: break
        return streak

    def week_hours(name, d):
        ws = week_start_sun(d)
        tot = 0
        first_day = date(year, month, 1)
        for i in range(7):
            dd = ws + timedelta(days=i)
            tot += calculate_daily_hours(dd, name, schedule, prev_sched, year, month, hols)
        return tot

    def find_pp(d):
        if not payperiods: return None
        for s, e in payperiods:
            if s <= d <= e: return (s, e)
        return None

    def pp_hours(name, d):
        pp = find_pp(d)
        if not pp: return 0
        s, e = pp
        tot = 0
        cur = s
        while cur <= e:
            tot += calculate_daily_hours(cur, name, schedule, prev_sched, year, month, hols)
            cur += timedelta(days=1)
        return tot

    def can_assign_primary(d, role, name, strict_mode):
        if d in hols: return False
        if name in cons.hard_off.get(d, set()): return False
        if name in cons.avoid_roles.get((d, role), set()): return False
        if name in schedule[d.isoformat()].values(): return False
        
        if last_role(name, d) == role: return False
        
        if consecutive_streak(name, d) >= config.max_consecutive_days: return False
        
        ws = week_start_sun(d)
        is_holiday_week = week_contains_holiday(ws, hols)
        is_push_week = week_contains_push(ws, push_days)
        
        if not (is_holiday_week or is_push_week):
            wh = week_hours(name, d)
            if wh + hours_for_role(d, role, hols) > 40: return False
        
        if not is_holiday_week:
            days_worked = 0
            for i in range(7):
                dd = ws + timedelta(days=i)
                worked = False
                if dd.year == year and dd.month == month:
                    if name in schedule.get(dd.isoformat(), {}).values(): worked = True
                elif dd < date(year, month, 1):
                    if worked_prev(prev_sched, dd, name): worked = True
                if worked: days_worked += 1
            if days_worked >= 5: return False

        if not is_push_week:
             curr_pp_h = pp_hours(name, d)
             if curr_pp_h + hours_for_role(d, role, hols) > 80: return False

        if strict_mode:
            if role_counts[role][name] >= targets[role][name]: return False
            if role == "PN" and weekday_sun0(d) == 1 and monday_pn_used[name] >= 1: return False
            
        return True

    def score_primary(d, role, name):
        s = 0.0
        t = targets[role][name]
        c = role_counts[role][name]
        s += (t - c) * 50.0 
        if role == "PN" and weekday_sun0(d) == 1 and monday_pn_used[name] >= 1: s -= 500.0
        s -= total_hours[name] * 2.0
        s -= pp_hours(name, d) * 1.0
        if role == "PN": s -= pn_weekday_counts[name][weekday_sun0(d)] * 10.0
        return s + rng.random()

    def can_bu(d, role, p):
        if d in hols: return False
        if p in cons.hard_off.get(d, set()): return False
        if p in schedule[d.isoformat()].values(): return False
        if consecutive_streak(p, d) >= config.max_consecutive_days: return False
        
        if last_role(p, d) == role: return False # Hard B2B for BU

        ws = week_start_sun(d)
        is_push = week_contains_push(ws, push_days)
        is_hol = week_contains_holiday(ws, hols)
        
        if not (is_push or is_hol):
             wh = week_hours(p, d)
             if wh + 8 > 40: return False
             days_w = 0
             for i in range(7):
                 dd = ws + timedelta(days=i)
                 if dd.year == year and dd.month == month:
                     if p in schedule.get(dd.isoformat(), {}).values(): days_w += 1
                 elif dd < date(year, month, 1):
                     if worked_prev(prev_sched, dd, p): days_w += 1
             if days_w >= 5: return False
        
        if not is_push:
             if pp_hours(p, d) + 8 > 80: return False

        return True

    def run_solver(strict_mode: bool) -> bool:
        sundays = [d for d in days if is_sunday(d) and d not in hols]
        
        prev_pn = None
        pd = sorted(prev_sched.keys())
        for x in reversed(pd):
            xd = datetime.strptime(x, "%Y-%m-%d").date()
            if is_sunday(xd):
                r = prev_sched[x].get("PN")
                if r in config.roster:
                    prev_pn = r
                    break
        
        start_idx = 0
        if prev_pn:
            start_idx = (config.roster.index(prev_pn) + 1) % len(config.roster)
            
        for i, sd in enumerate(sundays):
            assigned_sunday = False
            for offset in range(len(config.roster)):
                who = config.roster[(start_idx + i + offset) % len(config.roster)]
                
                if who in cons.hard_off.get(sd, set()): continue
                
                if strict_mode:
                    if role_counts["PN"][who] >= targets["PN"][who]: continue
                
                if not can_assign_primary(sd, "PN", who, strict_mode=False): continue

                schedule[sd.isoformat()]["PN"] = who
                role_counts["PN"][who] += 1
                total_hours[who] += 5
                pn_weekday_counts[who][0] += 1
                assigned_sunday = True
                
                sat = sd - timedelta(days=1)
                if sat.month == month and sat not in hols:
                    if can_assign_primary(sat, "AN", who, strict_mode=False):
                        schedule[sat.isoformat()]["AN"] = who
                        role_counts["AN"][who] += 1
                        total_hours[who] += 8
                
                break 
            
            if not assigned_sunday and strict_mode: return False

        primary_slots = []
        for d in days:
            req = roles_required_for_day(d, hols)
            for r in ["PN", "AN", "W"]:
                if r in req and r not in schedule[d.isoformat()]:
                    primary_slots.append((d, r))
        
        def count_candidates(slot):
            d, r = slot
            return sum(1 for n in config.roster if can_assign_primary(d, r, n, strict_mode))

        def backtrack(slots_left):
            if not slots_left: return True
            
            best_slot = min(slots_left, key=count_candidates)
            d, role = best_slot
            remaining_slots = [s for s in slots_left if s != best_slot]
            
            cands = [n for n in config.roster if can_assign_primary(d, role, n, strict_mode)]
            rng.shuffle(cands)
            cands.sort(key=lambda n: score_primary(d, role, n), reverse=True)
            
            for p in cands:
                schedule[d.isoformat()][role] = p
                role_counts[role][p] += 1
                hrs = hours_for_role(d, role, hols)
                total_hours[p] += hrs
                if role == "PN":
                    wd = weekday_sun0(d)
                    pn_weekday_counts[p][wd] += 1
                    if wd == 1: monday_pn_used[p] += 1
                
                if backtrack(remaining_slots): return True
                
                del schedule[d.isoformat()][role]
                role_counts[role][p] -= 1
                total_hours[p] -= hrs
                if role == "PN":
                    wd = weekday_sun0(d)
                    pn_weekday_counts[p][wd] -= 1
                    if wd == 1: monday_pn_used[p] -= 1
            
            return False

        return backtrack(primary_slots)

    print("Attempting STRICT solve (Pass 1)...")
    reset_state()
    if not run_solver(strict_mode=True):
        print(">> Strict solve failed. Switching to RELAXED solve (Pass 2)...")
        reset_state()
        if not run_solver(strict_mode=False):
            raise RuntimeError("❌ Solver failed even in relaxed mode. Constraints are too tight.")
        else:
            print(">> Relaxed solve successful.")
    else:
        print(">> Strict solve successful.")

    # --- BU FILL ---
    bu_queue = []
    for d in days:
        if d in hols: continue
        ideal = []
        if d in push_days:
            ideal = ["BU1", "BU2", "BU3", "BU4"] if is_saturday(d) else ["BU1", "BU2", "BU3"]
        elif weekday_sun0(d) == 1: ideal = ["BU1", "BU2", "BU3"] # Mon
        elif is_weekday(d): ideal = ["BU1", "BU2"] # Tue-Fri
        
        if not ideal: continue

        assigned = set(schedule[d.isoformat()].values())
        hard = cons.hard_off.get(d, set())
        pool = sorted([p for p in config.roster if p not in assigned and p not in hard])
        
        temp_pool = pool[:]
        for rname in ideal:
            if not temp_pool: break
            
            valid_cands = [p for p in temp_pool if can_bu(d, rname, p)]
            if not valid_cands: continue

            bu_queue.append((d, rname))
            person_to_reserve = valid_cands[0]
            temp_pool.remove(person_to_reserve)

    def bu_score(d, role, p):
        s = rng.random()
        s -= total_hours[p] * 0.5
        s -= pp_hours(p, d) * 1.0
        if p in cons.avoid_roles.get((d, role), set()): s -= 1000.0
        return s

    def solve_bu(idx):
        if idx >= len(bu_queue): return True
        d, role = bu_queue[idx]
        cands = [p for p in config.roster if can_bu(d, role, p)]
        rng.shuffle(cands)
        cands.sort(key=lambda p: bu_score(d, role, p), reverse=True)
        
        for p in cands:
            schedule[d.isoformat()][role] = p
            total_hours[p] += 8
            if solve_bu(idx + 1): return True
            del schedule[d.isoformat()][role]
            total_hours[p] -= 8
        return False

    if not solve_bu(0):
        print("Warning: Could not fill all desired BU slots.")

    def compact_bu_roles(sched):
        for d_iso, roles in sched.items():
            if not roles: continue
            bus = []
            for r in ["BU1", "BU2", "BU3", "BU4"]:
                if r in roles:
                    if roles[r]: bus.append(roles[r])
                    del roles[r]
            for i, p in enumerate(bus):
                roles[f"BU{i+1}"] = p
    compact_bu_roles(schedule)

    print("\n--- Schedule Fairness Audit ---")
    print(f"{'Name':<10} {'PN (Tgt)':<10} {'AN':<5} {'W':<5} {'TotHrs':<8}")
    for p in config.roster:
        pn = role_counts["PN"][p]
        an = role_counts["AN"][p]
        w = role_counts["W"][p]
        tgt = targets["PN"][p]
        print(f"{p:<10} {pn} ({tgt})      {an:<5} {w:<5} {total_hours[p]:<8}")

    return schedule


# ----------------------------
# Violations & Export
# ----------------------------

def write_violations_tab(wb, schedule, year, month, roster, targets, cons, prev_sched, payperiods, push_days, hols, week_starts):
    ws_vio = wb.create_sheet("Violations")
    ws_vio["A1"] = "Rule Violations / Audits"
    ws_vio["A1"].font = Font(bold=True, size=14)
    ws_vio.append(["Type", "Person", "Details"])
    for cell in ws_vio[2]: cell.font = Font(bold=True)

    # 1. Target Audit
    real_counts = {r: {p: 0 for p in roster} for r in ["PN", "AN", "W"]}
    for d_iso, rmap in schedule.items():
        for r in ["PN", "AN", "W"]:
            p = rmap.get(r)
            if p: real_counts[r][p] += 1
    for p in roster:
        for r in ["PN", "AN", "W"]:
            tgt = targets[r][p]
            act = real_counts[r][p]
            if act != tgt:
                ws_vio.append(["Target Mismatch", p, f"{r} Count {act} != Target {tgt}"])

    # 2. Daily Constraints (Unified Loop)
    sorted_dates = sorted(schedule.keys())
    
    for d_iso in sorted_dates:
        d = date.fromisoformat(d_iso)
        rmap = schedule[d_iso]
        
        seen_today = {} # p -> role
        hard_off_names = cons.hard_off.get(d, set())
        
        for r, p in rmap.items():
            # A. Duplicate Assignment
            if p in seen_today:
                first_r = seen_today[p]
                ws_vio.append(["Duplicate Assignment", p, f"Assigned {first_r} AND {r} on {d}"])
            else:
                seen_today[p] = r
            
            # B. Hard Off
            if p in hard_off_names:
                ws_vio.append(["Hard Off Violation", p, f"Worked {r} on {d} (Requested Off)"])
            
            # C. Avoid Role
            avoid_list = cons.avoid_roles.get((d, r), set())
            if p in avoid_list:
                rtype = "Avoid Violation (Hard)" if r in ROLE_PRIMARY else "Avoid Warning (Soft)"
                ws_vio.append([rtype, p, f"Assigned {r} on {d}"])
                
            # D. Back-to-Back (Same Role)
            prev_d = d - timedelta(days=1)
            prev_roles = schedule.get(prev_d.isoformat(), {})
            if prev_d.month != month:
                prev_roles = prev_sched.get(prev_d.isoformat(), {})
            
            if prev_roles.get(r) == p:
                ws_vio.append(["Back-to-Back (Same Role)", p, f"Worked {r} on {prev_d} and {d}"])

    # 3. Hours Audit
    for ws_date in week_starts:
        is_relaxed = week_contains_holiday(ws_date, hols) or week_contains_push(ws_date, push_days)
        if is_relaxed: continue
        for p in roster:
            tot = 0
            for i in range(7):
                dd = ws_date + timedelta(days=i)
                tot += calculate_daily_hours(dd, p, schedule, prev_sched, year, month, hols)
            if tot > 40:
                ws_vio.append(["Weekly Cap Violation", p, f"{tot}h in week of {ws_date} (Limit 40)"])

    if payperiods:
        for s, e in payperiods:
            # Relax if PP contains ANY push week (start)
            pp_relaxed = False
            cur_wk = week_start_sun(s)
            while cur_wk <= e:
                if week_contains_push(cur_wk, push_days):
                    pp_relaxed = True
                    break
                cur_wk += timedelta(days=7)
            
            if pp_relaxed: continue
            
            for p in roster:
                h = 0
                cur = s
                while cur <= e:
                    h += calculate_daily_hours(cur, p, schedule, prev_sched, year, month, hols)
                    cur += timedelta(days=1)
                if h > 80:
                    ws_vio.append(["PayPeriod Cap Violation", p, f"{h}h in PP {s}:{e} (Limit 80)"])


def export_to_excel(schedule, year, month, out_path, roster, hols, prev_sched, payperiods, sales_order, cons):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = calendar.month_name[month]
    ws.sheet_view.showGridLines = False

    for col in range(2, 23):
        ws.column_dimensions[get_column_letter(col)].width = 10
    
    first = date(year, month, 1)
    last = date(year, month, calendar.monthrange(year, month)[1])
    offset = weekday_sun0(first)
    weeks = (offset + last.day + 6) // 7
    grid_weeks = 6 if weeks > 5 else 5
    
    for r in range(3, GRID_WEEK_START_ROW + WEEK_HEIGHT * grid_weeks):
        ws.row_dimensions[r].height = 18

    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=22)
    c = ws.cell(1, 2, f"{calendar.month_name[month]} {year}")
    c.font = FONT_HEADER
    c.alignment = ALIGN_CENTER

    # Setup Fills
    white_fill = PatternFill("solid", fgColor=WHITE_FILL_HEX)
    sunday_fill = PatternFill("solid", fgColor=SUNDAY_FILL)
    nonmonth_fill = PatternFill("solid", fgColor=NONMONTH_FILL)

    for wk in range(grid_weeks):
        R = GRID_WEEK_START_ROW + WEEK_HEIGHT * wk
        for dow in range(7):
            col_start = SUN_START_COL + DAY_WIDTH * dow
            day_num = (wk * 7 + dow) - offset + 1
            
            in_month = (1 <= day_num <= last.day)
            
            fill = None
            if in_month:
                d = date(year, month, day_num)
                if is_sunday(d):
                    fill = sunday_fill
                else:
                    fill = white_fill # Explicit White
            else:
                fill = nonmonth_fill

            for rr in range(R, R + WEEK_HEIGHT):
                for cc in range(col_start, col_start + DAY_WIDTH):
                    cell = ws.cell(rr, cc)
                    cell.font = FONT_MAIN
                    cell.alignment = ALIGN_CENTER
                    cell.border = THIN_BORDER
                    if fill: cell.fill = fill

            if in_month:
                d = date(year, month, day_num)
                c = ws.cell(R, col_start + 2, day_num)
                c.font = FONT_DAY_NUM
                c.alignment = Alignment(horizontal="right", vertical="top")
                
                ws.cell(R+1, col_start, "PN-")
                ws.cell(R+2, col_start, "AN-")
                if is_weekday(d): ws.cell(R+3, col_start, "W-")
                
                roles = schedule.get(d.isoformat(), {})
                def write(off, r): ws.cell(R+off, col_start+1, roles.get(r, ""))
                
                write(1, "PN")
                if not is_sunday(d): write(2, "AN")
                if is_weekday(d): write(3, "W")
                write(4, "BU1"); write(5, "BU2"); write(6, "BU3"); write(7, "BU4")
                
                if d in hols:
                    ws.merge_cells(start_row=R, start_column=col_start, end_row=R+7, end_column=col_start+2)
                    c = ws.cell(R, col_start, "HOLIDAY")
                    c.alignment = ALIGN_CENTER
                    c.font = Font(bold=True, size=14)

    # Weekly Hours Sheet
    ws_weekly = wb.create_sheet("Weekly Hours")
    ws_weekly.sheet_view.showGridLines = False
    ws_weekly["A1"] = "Weekly Hours (Sun-Sat)"
    ws_weekly["A1"].font = Font(bold=True, size=14)
    
    headers = ["Week Start"] + roster
    ws_weekly.append(headers)
    
    week_starts = sorted({week_start_sun(d) for d in month_days(year, month)})
    ot_fill = PatternFill("solid", fgColor=OVERTIME_FILL)
    
    for ws_date in week_starts:
        row_vals = [ws_date.isoformat()]
        numeric_vals = []
        for p in roster:
            tot = 0
            for i in range(7):
                dd = ws_date + timedelta(days=i)
                tot += calculate_daily_hours(dd, p, schedule, prev_sched, year, month, hols)
            numeric_vals.append(tot)
        
        ws_weekly.append(row_vals + numeric_vals)
        
        curr_row = ws_weekly.max_row
        for idx, val in enumerate(numeric_vals):
            if val > 40:
                ws_weekly.cell(row=curr_row, column=2+idx).fill = ot_fill
    
    # Summary Tab
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Summary"
    ws2["A1"].font = Font(bold=True, size=14)
    
    headers = ["Name"] + ROLE_PRIMARY + ROLE_BUS + ["Total", "Hours"]
    ws2.append(headers)
    
    for p in roster:
        row = [p]
        tot_ass = 0
        tot_hrs = 0
        counts = {r: 0 for r in ROLE_PRIMARY + ROLE_BUS}
        for d_iso, rmap in schedule.items():
            for r, name in rmap.items():
                if name == p:
                    counts[r] += 1
                    tot_hrs += hours_for_role(date.fromisoformat(d_iso), r, hols)
        for r in ROLE_PRIMARY + ROLE_BUS:
            row.append(counts[r])
            tot_ass += counts[r]
        row.append(tot_ass)
        row.append(tot_hrs)
        ws2.append(row)

    if payperiods:
        ws3 = wb.create_sheet("Pay Period Hours")
        ws3.append(["Pay Period", "Start", "End"] + roster)
        for i, (s, e) in enumerate(payperiods, 1):
            row = [f"PP{i}", s.isoformat(), e.isoformat()]
            for p in roster:
                h = 0
                cur = s
                while cur <= e:
                    h += calculate_daily_hours(cur, p, schedule, prev_sched, year, month, hols)
                    cur += timedelta(days=1)
                row.append(h)
            ws3.append(row)

    # VIOLATIONS TAB (Refactored)
    # Compute computed push days once for audit
    push_days = push_days_for_month(year, month)
    targets = compute_primary_targets(month_days(year, month), roster, sales_order, hols)
    
    write_violations_tab(wb, schedule, year, month, roster, targets, cons, prev_sched, payperiods, push_days, hols, week_starts)

    wb.save(out_path)


# ----------------------------
# Main
# ----------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", required=True)
    ap.add_argument("--timeoff-xlsx", required=True)
    ap.add_argument("--timeoff-sheet", required=True)
    ap.add_argument("--prev-schedule", required=True)
    ap.add_argument("--sales-sheet", default="Sales Ranking")
    ap.add_argument("--roster-json")
    ap.add_argument("--payperiods-json")
    ap.add_argument("--payperiod-anchor-start", default="2026-01-25")
    ap.add_argument("--out-json")
    ap.add_argument("--out-xlsx")
    ap.add_argument("--seed", type=int, default=11)
    args = ap.parse_args()

    year, month = parse_month_arg(args.month)
    roster = load_roster_from_json(args.roster_json) if args.roster_json else DEFAULT_ROSTER[:]
    timeoff = load_timeoff_from_xlsx(args.timeoff_xlsx, args.timeoff_sheet)
    cons = compile_constraints(timeoff)
    sales = load_sales_ranking_from_timeoff_xlsx(args.timeoff_xlsx, args.sales_sheet, year, month, roster)
    prev = load_schedule_json(args.prev_schedule)
    cfg = BuildConfig(year, month, roster, sales, rng_seed=args.seed)

    if args.payperiods_json:
        pps = load_payperiods_from_json(args.payperiods_json)
    else:
        ms = date(year, month, 1)
        me = date(year, month, calendar.monthrange(year, month)[1])
        anc = datetime.strptime(args.payperiod_anchor_start, "%Y-%m-%d").date()
        raw = generate_payperiods(anc, ms - timedelta(days=28), me + timedelta(days=28))
        pps = [p for p in raw if p[0] <= me and p[1] >= ms]

    sched = build_schedule(cfg, prev, cons, pps)

    oj = args.out_json or str(Path(args.prev_schedule).with_name(f"{year}_{month:02d}_schedule.json"))
    ox = args.out_xlsx or str(Path(args.prev_schedule).with_name(f"{calendar.month_name[month]}_{year}_Schedule.xlsx"))

    Path(oj).write_text(json.dumps(sched, indent=2), "utf-8")
    print(f"✅ JSON: {oj}")
    
    # PASS CONS object for correct audit
    export_to_excel(sched, year, month, ox, roster, observed_holidays_for_year(year), prev, pps, sales, cons)
    print(f"✅ Excel: {ox}")

if __name__ == "__main__":
    main()
