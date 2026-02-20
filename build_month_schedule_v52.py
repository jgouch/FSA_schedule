#!/usr/bin/env python3
"""
build_month_schedule_v52.py

End-to-end month builder:
- Reads TimeOff.xlsx (monthly request sheet + Sales Ranking tab)
- Reads prior month schedule JSON for cross-month constraints
- Generates schedule for target month (primaries then BU)
- Exports:
    - schedule JSON (machine truth)
    - Excel calendar workbook (Calendar, Summary, Weekly Hours, Pay Period Hours, Violations)

LOCKED CALENDAR GEOMETRY:
- Grid weeks: Dynamic (5 or 6) based on month shape.
- Each day block: 3 columns x 8 rows
- Role mapping (R = week_start_row, cols = start/mid/end):
    PN label: (start, R+1), PN name: (mid, R+1)
    AN label: (start, R+2), AN name: (mid, R+2)
    W  label: (start, R+3), W  name: (mid, R+3)
    BU1 name: (mid, R+4)
    BU2 name: (mid, R+5)
    BU3 name: (mid, R+6)
    BU4 name (push-week Saturday only): (mid, R+7)

Logic Changes (v34):
- CLEANUP: MAX_WEEKS constant physically deleted.
- OPTIMIZATION: Added caching for holiday lookup (speeds up hours calc).
- UNIFICATION: 5-day cap logic now uses global 'calculate_daily_hours' helper.
- FIX (v44): Make sunday_month_dup_used reflect true month-duplicate usage; strict-mode fails if Sunday dupes occur.
- v45: add auto-seed retry loop with staffing validator (no external bash loop needed).
- FIX (v46): Enforce hard minimum weekday staffing (>=4 unique people Mon–Fri except holidays) and make auto-seed loop resilient to per-seed failures.
- FIX (v47): add optional weekday target staffing optimization (aim for 5 unique people Mon–Fri when feasible)
- FIX (v48): target-weekday metrics only count weekdays where target is feasible; early-exit truly stops seed sweep when perfect is found.
- FIX (v49): cap auto-seed search to 400 attempts by default (seed_end default=400), while still allowing override via CLI.
- FIX (v50): add seed selection tie-breakers (BU deficits + hour cap violation counts) when optimize-weekday-staff is enabled.
- FIX (v51): default seed search capped at 50 attempts; add unconditional empty seed-range guard in auto-seed.
- FIX (v52): add Weekday Target Repair Pass (BU-only micro-swaps within week) to improve eligible weekdays reaching target staffing.
"""

from __future__ import annotations

import argparse
import calendar
import json
import random
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


# ----------------------------
# Constants / settings
# ----------------------------

SUNDAY_FILL = "FFF2CC"
NONMONTH_FILL = "F2F2F2"
WHITE_FILL_HEX = "FFFFFF"
OVERTIME_FILL = "FFC7CE" # Red fill for >40h weeks

FONT_MAIN = Font(name="Calibri", size=11)
FONT_DAY_NUM = Font(name="Calibri", size=11, bold=True)
FONT_HEADER = Font(name="Calibri", size=24, bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN_BORDER = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

GRID_WEEK_START_ROW = 3
WEEK_HEIGHT = 8
DAY_WIDTH = 3
SUN_START_COL = 2  # B

ROLE_PRIMARY = ["PN", "AN", "W"]
ROLE_BUS = ["BU1", "BU2", "BU3", "BU4"]

OBSERVED_HOLIDAYS_BY_YEAR = {
    2026: {
        date(2026, 1, 1),   # New Year's Day
        date(2026, 5, 25),  # Memorial Day
        date(2026, 7, 3),   # Independence Day observed
        date(2026, 9, 7),   # Labor Day
        date(2026, 11, 26), # Thanksgiving
        date(2026, 12, 25), # Christmas
    }
}

DEFAULT_ROSTER = ["Greg", "Mark", "Dawn", "Will", "CJ", "Kyle"]
SENIORITY_ORDER = ["Mark", "Dawn", "Will", "Kyle", "CJ", "Greg"]


# ----------------------------
# Helpers
# ----------------------------

def weekday_sun0(d: date) -> int:
    return (d.weekday() + 1) % 7

def is_sunday(d: date) -> bool:
    return weekday_sun0(d) == 0

def is_saturday(d: date) -> bool:
    return weekday_sun0(d) == 6

def is_weekday(d: date) -> bool:
    return d.weekday() < 5

def month_days(year: int, month: int) -> List[date]:
    n = calendar.monthrange(year, month)[1]
    return [date(year, month, d) for d in range(1, n+1)]

def quarter_tag_for_month(year: int, month: int) -> str:
    q = (month - 1) // 3 + 1
    return f"{year}Q{q}"

def norm_name(x) -> str:
    return re.sub(r"\s+", " ", str(x).strip())

def _normalize_header_key(value: str) -> str:
    """Normalize column headers so minor spacing/punctuation differences still match."""
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())

def normalize_quarter_key(s: str) -> str:
    """Normalize '2026 Q1' -> '2026Q1'."""
    return re.sub(r"\s+", "", str(s).strip().upper())

def parse_month_arg(s: str) -> Tuple[int, int]:
    m = re.match(r"^\s*(\d{4})-(\d{2})\s*$", s)
    if not m:
        raise ValueError("Month must be YYYY-MM, e.g. 2026-02")
    year = int(m.group(1))
    month = int(m.group(2))
    if not 1 <= month <= 12:
        raise ValueError(f"Month out of range in '{s}'. Expected 01..12")
    return year, month

def compute_observed_holidays_us(year: int) -> Set[date]:
    out: Set[date] = set()
    out.add(date(year, 1, 1))
    d = date(year, 5, 31)
    while d.weekday() != 0: d -= timedelta(days=1)
    out.add(d)
    july4 = date(year, 7, 4)
    if july4.weekday() == 5: out.add(date(year, 7, 3))
    elif july4.weekday() == 6: out.add(date(year, 7, 5))
    else: out.add(july4)
    d = date(year, 9, 1)
    while d.weekday() != 0: d += timedelta(days=1)
    out.add(d)
    d = date(year, 11, 1)
    while d.weekday() != 3: d += timedelta(days=1)
    out.add(d + timedelta(days=21))

    # Christmas observed rule (Fri if Saturday holiday, Mon if Sunday holiday)
    xmas = date(year, 12, 25)
    if xmas.weekday() == 5:
        out.add(date(year, 12, 24))
    elif xmas.weekday() == 6:
        out.add(date(year, 12, 26))
    else:
        out.add(xmas)
    return out

_hols_cache = {}
def observed_holidays_for_year(year: int) -> Set[date]:
    if year in _hols_cache:
        return _hols_cache[year]
    
    hols = OBSERVED_HOLIDAYS_BY_YEAR.get(year, compute_observed_holidays_us(year))
    _hols_cache[year] = hols
    return hols

def week_start_sun(d: date) -> date:
    return d - timedelta(days=weekday_sun0(d))

def generate_payperiods(anchor_start: date, start_date: date, end_date: date) -> List[Tuple[date, date]]:
    a = anchor_start
    while a > start_date:
        a -= timedelta(days=14)
    periods: List[Tuple[date, date]] = []
    cur = a
    while cur <= end_date:
        periods.append((cur, cur + timedelta(days=13)))
        cur += timedelta(days=14)
    return periods

def parse_excel_date(value) -> date:
    if isinstance(value, datetime): return value.date()
    if isinstance(value, date): return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try: return from_excel(value).date()
        except Exception as e: 
            raise ValueError(f"Failed to parse numeric Excel date: {value}. Error: {e}")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date format: {value!r}")

def _rotate_left(lst: List[str], k: int) -> List[str]:
    if not lst: return lst
    k = k % len(lst)
    return lst[k:] + lst[:k]

# Global logic checks
def week_contains_push(ws: date, push_days: Set[date]) -> bool:
    for i in range(7):
        if (ws + timedelta(days=i)) in push_days: return True
    return False

def week_contains_holiday(ws: date, hols: Set[date]) -> bool:
    for i in range(7):
        if (ws + timedelta(days=i)) in hols: return True
    return False

# Global hours calculator (Single Source of Truth)
def calculate_daily_hours(dd: date, name: str, schedule: Dict, prev_sched: Dict, year: int, month: int) -> int:
    """Calculate hours for a person on a specific date, handling month boundaries."""
    
    # Dynamic holiday lookup based on the specific date's year
    hols = observed_holidays_for_year(dd.year)
    
    if dd.year == year and dd.month == month:
        roles = schedule.get(dd.isoformat(), {})
        dh = 0
        for r, n in roles.items():
            if n == name: dh = max(dh, hours_for_role(dd, r, hols))
        return dh
    elif dd < date(year, month, 1):
        roles = prev_sched.get(dd.isoformat(), {})
        dh = 0
        for r, n in roles.items():
            if n == name: 
                dh = max(dh, hours_for_role(dd, r, hols))
        return dh
    return 0


# ----------------------------
# Inputs
# ----------------------------

@dataclass
class TimeOffRule:
    name: str
    d: date
    hard: bool = True
    avoid_roles: Set[str] = field(default_factory=set)

@dataclass
class Constraints:
    hard_off: Dict[date, Set[str]]
    avoid_roles: Dict[Tuple[date, str], Set[str]]

def compile_constraints(timeoff: List[TimeOffRule]) -> Constraints:
    hard_off: Dict[date, Set[str]] = {}
    avoid_roles: Dict[Tuple[date, str], Set[str]] = {}
    for r in timeoff:
        if r.hard:
            hard_off.setdefault(r.d, set()).add(r.name)
        if r.avoid_roles:
            for role in r.avoid_roles:
                avoid_roles.setdefault((r.d, role.upper()), set()).add(r.name)
    return Constraints(hard_off=hard_off, avoid_roles=avoid_roles)

def load_timeoff_from_xlsx(xlsx_path: str, sheet_name: str) -> List[TimeOffRule]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Missing timeoff sheet '{sheet_name}'")
        ws = wb[sheet_name]

        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            if v:
                headers[_normalize_header_key(v)] = c

        c_name = headers.get("name")
        if not c_name:
            raise ValueError("TimeOff sheet missing 'Name' column")
        c_date = headers.get("date")
        c_start = headers.get("start")
        c_end = headers.get("end")
        c_hard = headers.get("hard")
        c_avoid = headers.get("avoidroles")

        def parse_hard(v) -> bool:
            if v is None or str(v).strip() == "": return True
            return str(v).strip().lower() in ("true", "1", "yes", "y")

        rules = []
        for r in range(2, ws.max_row + 1):
            nv = ws.cell(r, c_name).value
            if not nv: continue
            name = norm_name(nv)

            dates = []
            dv = ws.cell(r, c_date).value if c_date else None
            sv = ws.cell(r, c_start).value if c_start else None
            ev = ws.cell(r, c_end).value if c_end else None

            if dv:
                dates = [parse_excel_date(dv)]
            elif sv and ev:
                sd = parse_excel_date(sv)
                ed = parse_excel_date(ev)
                if ed < sd:
                    raise ValueError(f"Invalid time-off range for {name}: start {sd} is after end {ed}")
                cur = sd
                while cur <= ed:
                    dates.append(cur)
                    cur += timedelta(days=1)
            else:
                continue

            hard = parse_hard(ws.cell(r, c_hard).value) if c_hard else True
            av = set()
            if c_avoid:
                val = ws.cell(r, c_avoid).value
                if val:
                    av = {p.strip().upper() for p in str(val).split(",") if p.strip()}
                    invalid = av - set(ROLE_PRIMARY + ROLE_BUS)
                    if invalid:
                        raise ValueError(
                            f"Unknown role(s) in AvoidRoles for {name} on row {r}: {sorted(invalid)}"
                        )

            for d in dates:
                rules.append(TimeOffRule(name, d, hard, av))
        return rules
    finally:
        wb.close()

def load_roster_from_json(path: str) -> List[str]:
    data = json.loads(Path(path).read_text("utf-8"))
    if not isinstance(data, list) or not data:
        raise ValueError("Roster JSON must be a non-empty list of names")

    roster = [norm_name(x) for x in data if str(x).strip()]
    if not roster:
        raise ValueError("Roster JSON must contain at least one non-blank name")
    return roster

def load_payperiods_from_json(path: str) -> List[Tuple[date, date]]:
    data = json.loads(Path(path).read_text("utf-8"))
    if not isinstance(data, list):
        raise ValueError("Pay periods JSON must be a list of objects with 'start' and 'end'")
    out = []
    for obj in data:
        if not isinstance(obj, dict):
            raise ValueError("Each pay period entry must be an object with 'start' and 'end'")
        if "start" not in obj or "end" not in obj:
            raise ValueError(f"Pay period entry missing required keys: {obj}")
        s = datetime.strptime(str(obj["start"]).strip(), "%Y-%m-%d").date()
        e = datetime.strptime(str(obj["end"]).strip(), "%Y-%m-%d").date()
        if e < s:
            raise ValueError(f"Pay period end before start: {s}..{e}")
        out.append((s, e))
    out.sort(key=lambda t: t[0])
    return out

def load_sales_ranking_from_timeoff_xlsx(xlsx_path, sheet_name, year, month, roster_names):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Missing sheet '{sheet_name}'")
        ws = wb[sheet_name]
        target = normalize_quarter_key(quarter_tag_for_month(year, month))

        rows = {}
        for r in range(2, ws.max_row + 1):
            p = ws.cell(r, 1).value
            if not p: continue
            p = normalize_quarter_key(str(p))
            ranks = []
            for c in range(2, ws.max_column + 1):
                v = ws.cell(r, c).value
                if v and str(v).strip():
                    ranks.append(norm_name(v))
            if ranks: rows[p] = ranks

        def key(x):
            # Support both raw values ("2026 Q1") and normalized values ("2026Q1").
            m = re.match(r"^(\d{4})Q([1-4])$", normalize_quarter_key(x))
            return (int(m.group(1)), int(m.group(2))) if m else (0,0)

        order = rows.get(target)
        if not order:
            candidates = sorted(rows.keys(), key=key)
            if candidates: order = rows[candidates[-1]]

        roster_set = set(roster_names)
        if not order:
            return [n for n in SENIORITY_ORDER if n in roster_set] + [n for n in roster_names if n not in SENIORITY_ORDER]

        order = [x for x in order if x in roster_set]
        missing = roster_set - set(order)
        order.extend(sorted(list(missing)))
        return order
    finally:
        wb.close()

def load_schedule_json(path: str) -> Dict[str, Dict[str, str]]:
    data = json.loads(Path(path).read_text("utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Schedule JSON must be a dictionary")
    iso_pat = re.compile(r"^\d{4}-\d{2}-\d{2}$")
    for k, v in data.items():
        if not iso_pat.match(k):
             raise ValueError(f"Invalid date key in schedule JSON: {k} (Must be YYYY-MM-DD)")
        if not isinstance(v, dict):
             raise ValueError(f"Schedule JSON invalid at key {k}: value must be a dictionary")
    return data

def role_of(prev, d, name):
    roles = prev.get(d.isoformat(), {})
    for r, n in roles.items():
        if n == name: return r
    return None

def worked_prev(prev, d, name):
    return name in prev.get(d.isoformat(), {}).values()


# ----------------------------
# Core Logic
# ----------------------------

def roles_required_for_day(d: date, hols: Set[date]) -> List[str]:
    if d in hols: return []
    if is_sunday(d): return ["PN"]
    req = ["PN", "AN"]
    if is_weekday(d): req.append("W")
    return req

def hours_for_role(d: date, role: str, hols: Set[date]) -> int:
    if d in hols: return 0
    if is_sunday(d) and role == "PN": return 5
    return 8

def push_days_for_month(year: int, month: int) -> Set[date]:
    """Return set of dates that are 'Push Days' (Last occurrence of Mon-Sat)."""
    days = set()
    last_dom = calendar.monthrange(year, month)[1]
    last = date(year, month, last_dom)
    for wd in range(6): 
        cur = last
        while cur.weekday() != wd:
            cur -= timedelta(days=1)
        if cur.month == month:
            days.add(cur)
    return days

@dataclass
class BuildConfig:
    year: int
    month: int
    roster: List[str]
    sales_order: List[str]
    max_consecutive_days: int = 5
    lookback_days: int = 14
    rng_seed: int = 11
    enable_weekday_target_repair: bool = False
    weekday_target_repair_verbose: bool = False
    target_weekday_staff: int = 5

def compute_primary_targets(days, roster, sales_order, hols):
    if not roster:
        raise ValueError("Roster cannot be empty when computing primary targets")

    opp = {"PN": 0, "AN": 0, "W": 0}
    for d in days:
        req = roles_required_for_day(d, hols)
        for r in opp:
            if r in req: opp[r] += 1
            
    n = len(roster)
    targets = {r: {p: 0 for p in roster} for r in opp}
    
    roster_set = set(roster)
    valid_order = [x for x in sales_order if x in roster_set]
    missing = [x for x in roster if x not in valid_order]
    full_priority = valid_order + missing
    
    def distribute(role, total):
        base = total // n
        rem = total % n
        for p in roster: targets[role][p] = base
        
        offset = 0
        rem_pn = opp["PN"] % n
        rem_an = opp["AN"] % n
        
        if role == "AN": offset = rem_pn
        if role == "W": offset = (rem_pn + rem_an) % n
        
        rotated_order = _rotate_left(full_priority, offset)
        for i in range(rem):
            targets[role][rotated_order[i]] += 1

    for r in opp: distribute(r, opp[r])
    return targets


def build_schedule(config: BuildConfig,
                   prev_sched: Dict[str, Dict[str, str]],
                   cons: Constraints,
                   payperiods: Optional[List[Tuple[date, date]]] = None) -> Dict[str, Dict[str, str]]:
    
    rng = random.Random(config.rng_seed)
    year, month = config.year, config.month
    days = month_days(year, month)
    
    hols = observed_holidays_for_year(year)
    push_days = push_days_for_month(year, month)
    targets = compute_primary_targets(days, config.roster, config.sales_order, hols)

    schedule: Dict[str, Dict[str, str]] = {}
    role_counts = {r: {n: 0 for n in config.roster} for r in ["PN", "AN", "W"]}
    total_hours = {n: 0 for n in config.roster}
    pn_weekday_counts = {n: {i: 0 for i in range(7)} for n in config.roster}
    monday_pn_used = {n: 0 for n in config.roster}

    def reset_state():
        nonlocal schedule, role_counts, total_hours, pn_weekday_counts, monday_pn_used
        schedule = {d.isoformat(): {} for d in days}
        role_counts = {r: {n: 0 for n in config.roster} for r in ["PN", "AN", "W"]}
        total_hours = {n: 0 for n in config.roster}
        pn_weekday_counts = {n: {i: 0 for i in range(7)} for n in config.roster}
        monday_pn_used = {n: 0 for n in config.roster}

    def consecutive_streak(name, d):
        streak = 0
        for k in range(1, config.lookback_days + 1):
            p = d - timedelta(days=k)
            worked = False
            if p.year == year and p.month == month:
                worked = name in schedule.get(p.isoformat(), {}).values()
            else:
                worked = worked_prev(prev_sched, p, name)
            if worked: streak += 1
            else: break
        return streak
    def consecutive_span(name, d):

        """Return (before, after) consecutive worked-day counts around date d, given current partial schedule.

        Counts contiguous days immediately before d and immediately after d where the person already has ANY role.

        Uses prev_sched for days < month start, and schedule for days within the target month.

        """

        before = 0

        for k in range(1, config.lookback_days + 1):

            p = d - timedelta(days=k)

            if p.year == year and p.month == month:

                worked = name in schedule.get(p.isoformat(), {}).values()

            else:

                worked = worked_prev(prev_sched, p, name)

            if worked:

                before += 1

            else:

                break

        after = 0

        for k in range(1, config.lookback_days + 1):

            n = d + timedelta(days=k)

            if not (n.year == year and n.month == month):

                break

            worked = name in schedule.get(n.isoformat(), {}).values()

            if worked:

                after += 1

            else:

                break

        return before, after


    def week_hours(name, d):
        ws = week_start_sun(d)
        tot = 0
        for i in range(7):
            dd = ws + timedelta(days=i)
            tot += calculate_daily_hours(dd, name, schedule, prev_sched, year, month)
        return tot

    def find_pp(d):
        if not payperiods: return None
        for s, e in payperiods:
            if s <= d <= e: return (s, e)
        return None

    def pp_hours(name, d):
        pp = find_pp(d)
        if not pp: return 0
        s, e = pp
        tot = 0
        cur = s
        while cur <= e:
            tot += calculate_daily_hours(cur, name, schedule, prev_sched, year, month)
            cur += timedelta(days=1)
        return tot

    def can_assign_primary(d, role, name, strict_mode):
        if d in hols: return False
        if name in cons.hard_off.get(d, set()): return False
        if name in cons.avoid_roles.get((d, role), set()): return False
        if name in schedule[d.isoformat()].values(): return False
        
        # No consecutive Saturdays (any role) — relaxed during push/holiday weeks
        ws_sat = week_start_sun(d)
        if is_saturday(d) and (not (week_contains_push(ws_sat, push_days) or week_contains_holiday(ws_sat, hols))) and calculate_daily_hours(d - timedelta(days=7), name, schedule, prev_sched, year, month) != 0: return False
        # B2B (Same Role) should mean previous CALENDAR day, not previous WORKED role
        prev_d = d - timedelta(days=1)
        prev_roles = schedule.get(prev_d.isoformat(), {})
        if prev_d.month != month:
            prev_roles = prev_sched.get(prev_d.isoformat(), {})
        if prev_roles.get(role) == name: return False
        if (d + timedelta(days=1)).month == month and schedule.get((d + timedelta(days=1)).isoformat(), {}).get(role) == name: return False

        # Sat AN / Sun PN pairing guard (both directions) to avoid unpaired weekend handoff.
        # Default: Sat AN must match Sun PN.
        # Only exception: the Sunday PN person is hard-off on that Saturday.
        if role == 'PN' and is_sunday(d):
            sat = d - timedelta(days=1)
            sat_hols = observed_holidays_for_year(sat.year)
            if sat in sat_hols or sat in hols:
                pass
            elif name in cons.hard_off.get(sat, set()):
                pass
            else:
                if sat.month == month:
                    sat_an = schedule.get(sat.isoformat(), {}).get('AN')
                    if sat_an and sat_an != name:
                        return False
                else:
                    if prev_sched.get(sat.isoformat(), {}).get('AN') != name:
                        return False


        if role == 'AN' and is_saturday(d):
            sun = d + timedelta(days=1)
            sun_hols = observed_holidays_for_year(sun.year)
            if sun in hols or sun in sun_hols:
                pass
            elif sun.month == month:
                sun_pn = schedule.get(sun.isoformat(), {}).get('PN')
                # Exception: if Sunday PN person is hard-off today (Saturday), allow mismatch.
                # Sanity example: Feb 21 hard-off for Greg allows Sun 22 PN=Greg without forcing Sat 21 AN=Greg.
                if sun_pn and (sun_pn not in cons.hard_off.get(d, set())) and sun_pn != name:
                    return False
            else:
                # Cross-month: Sunday is in next month; pairing is validated when that month is built.
                pass


        # Push week relaxation: consecutive-days cap not enforced during push week
        
        ws = week_start_sun(d)
        is_holiday_week = week_contains_holiday(ws, hols)
        is_push_week = week_contains_push(ws, push_days)
        # Boundary safeguard: if this is the LAST day of the month and it's Saturday AN,
        # cross-month pairing forces next-day Sunday PN. Prevent exceeding max consecutive days.
        before, after = consecutive_span(name, d)
        if role == 'AN' and is_saturday(d) and d.day == calendar.monthrange(year, month)[1] and (before + 2) > config.max_consecutive_days: return False
        if (not is_push_week) and (before + 1 + after) > config.max_consecutive_days: return False
        
        if not (is_holiday_week or is_push_week):
            wh = week_hours(name, d)
            if wh + hours_for_role(d, role, hols) > 40: return False
        
        if not (is_holiday_week or is_push_week):
            days_worked = 0
            for i in range(7):
                dd = ws + timedelta(days=i)
                # Unify logic: Count worked days via hours > 0
                if calculate_daily_hours(dd, name, schedule, prev_sched, year, month) > 0:
                    days_worked += 1
            if days_worked >= 5: return False

        if not is_push_week:
             curr_pp_h = pp_hours(name, d)
             if curr_pp_h + hours_for_role(d, role, hols) > 80: return False

        # PN fairness stays enforced even in RELAXED mode (prevents PN delta > 1)
        if role == 'PN' and role_counts[role][name] >= targets[role][name]: return False
        if strict_mode:
            if role_counts[role][name] >= targets[role][name]: return False
            if role == "PN" and weekday_sun0(d) == 1 and monday_pn_used[name] >= 1: return False
            
        return True

    def score_primary(d, role, name):
        s = 0.0
        t = targets[role][name]
        c = role_counts[role][name]
        s += (t - c) * 50.0 
        if role == "PN" and weekday_sun0(d) == 1 and monday_pn_used[name] >= 1: s -= 500.0
        s -= total_hours[name] * 2.0
        s -= pp_hours(name, d) * 1.0
        if role == "PN": s -= pn_weekday_counts[name][weekday_sun0(d)] * 10.0
        return s + rng.random()

    def can_bu(d, role, p):
        if d in hols: return False
        if p in cons.hard_off.get(d, set()): return False
        if p in schedule[d.isoformat()].values(): return False
        if role == 'BU1' and enforce_bu1_cap and bu1_counts[p] >= bu1_targets[p]: return False
        # No consecutive Saturdays (any role) — relaxed during push/holiday weeks
        ws_sat = week_start_sun(d)
        if is_saturday(d) and (not (week_contains_push(ws_sat, push_days) or week_contains_holiday(ws_sat, hols))) and calculate_daily_hours(d - timedelta(days=7), p, schedule, prev_sched, year, month) != 0: return False
        # Push week relaxation: consecutive-days cap not enforced during push week

        ws = week_start_sun(d)
        is_push = week_contains_push(ws, push_days)
        before, after = consecutive_span(p, d)
        if (not is_push) and (before + 1 + after) > config.max_consecutive_days: return False
        is_hol = week_contains_holiday(ws, hols)
        
        if not (is_push or is_hol):
             wh = week_hours(p, d)
             if wh + 8 > 40: return False
             days_w = 0
             for i in range(7):
                 dd = ws + timedelta(days=i)
                 if calculate_daily_hours(dd, p, schedule, prev_sched, year, month) > 0:
                     days_w += 1
             if days_w >= 5: return False
        
        if not is_push:
             if pp_hours(p, d) + 8 > 80: return False

        return True

    def run_solver(strict_mode: bool) -> bool:
        sundays = [d for d in days if is_sunday(d) and d not in hols]
        month_sunday_count = {p: 0 for p in config.roster}
        sunday_month_dup_used = False

        # 6-week rolling Sunday PN rotation across months, driven by actual prior schedules
        # This continues from the last actual Sunday PN in prev_sched, so manual edits are respected.
        from collections import deque
        recent_sundays = deque(maxlen=len(config.roster))
        prev_pn = None
        # Collect up to N most recent Sunday PN assignees from prev month schedule
        for x in sorted(prev_sched.keys(), reverse=True):
            xd = datetime.strptime(x, '%Y-%m-%d').date()
            if not is_sunday(xd):
                continue
            r = prev_sched.get(x, {}).get('PN')
            if r in config.roster:
                if prev_pn is None:
                    prev_pn = r
                recent_sundays.appendleft(r)
                if len(recent_sundays) >= len(config.roster):
                    break

        start_idx = 0
        if prev_pn:
            start_idx = (config.roster.index(prev_pn) + 1) % len(config.roster)
        rot_idx = start_idx

        for sd in sundays:
            assigned_sunday = False
            recent_set = set(recent_sundays)
            # Three-pass selection:
            #   A) strict rotation + no month duplicate
            #   B) relaxed rotation + no month duplicate
            #   C) relaxed rotation + allow month duplicate (last resort)
            for allow_recent, allow_month_dup in ((False, False), (True, False), (True, True)):
                for offset in range(len(config.roster)):
                    who = config.roster[(rot_idx + offset) % len(config.roster)]
                    if (not allow_recent) and (who in recent_set):
                        continue
                    if (not allow_month_dup) and month_sunday_count[who] >= 1:
                        continue
                    if who in cons.hard_off.get(sd, set()):
                        continue
                    if strict_mode and role_counts['PN'][who] >= targets['PN'][who]:
                        continue
                    if not can_assign_primary(sd, 'PN', who, strict_mode=strict_mode):
                        continue

                    # Pairing guard: require Sat AN to match this Sun PN (including cross-month boundary),
                    # except when Saturday is a holiday/observed holiday or Sunday PN is hard-off on Saturday.
                    sat = sd - timedelta(days=1)
                    sat_hols = observed_holidays_for_year(sat.year)
                    if sat in sat_hols or sat in hols:
                        pass
                    elif who in cons.hard_off.get(sat, set()):
                        pass
                    elif sat.month == month:
                        if not can_assign_primary(sat, 'AN', who, strict_mode=strict_mode):
                            continue
                    else:
                        prev_roles = prev_sched.get(sat.isoformat(), {})
                        prev_an = prev_roles.get('AN')
                        if (prev_an is None) or (str(prev_an).strip() == '') or (prev_an != who):
                            continue

                    # Assign Sunday PN
                    schedule[sd.isoformat()]['PN'] = who
                    role_counts['PN'][who] += 1
                    total_hours[who] += 5
                    pn_weekday_counts[who][0] += 1
                    already_had_sunday = (month_sunday_count[who] >= 1)
                    month_sunday_count[who] += 1
                    if allow_month_dup and already_had_sunday:
                        sunday_month_dup_used = True
                    assigned_sunday = True
                    # Advance rotation based on actual assignment
                    rot_idx = (config.roster.index(who) + 1) % len(config.roster)
                    recent_sundays.append(who)

                    # If Saturday is in-month and open, assign Sat AN to match Sunday PN
                    sat = sd - timedelta(days=1)
                    sat_hols = observed_holidays_for_year(sat.year)
                    if sat.month == month and sat not in hols and sat not in sat_hols:
                        if can_assign_primary(sat, 'AN', who, strict_mode=strict_mode):
                            schedule[sat.isoformat()]['AN'] = who
                            role_counts['AN'][who] += 1
                            total_hours[who] += 8
                    break
                if assigned_sunday:
                    break
            if not assigned_sunday and strict_mode:
                return False

        has_sunday_dupes = any(cnt > 1 for cnt in month_sunday_count.values())
        if strict_mode and has_sunday_dupes:
            return False


        primary_slots = []
        for d in days:
            req = roles_required_for_day(d, hols)
            for r in ["PN", "AN", "W"]:
                if r in req and r not in schedule[d.isoformat()]:
                    primary_slots.append((d, r))
        
        def count_candidates(slot):
            d, r = slot
            return sum(1 for n in config.roster if can_assign_primary(d, r, n, strict_mode))

        def backtrack(slots_left):
            if not slots_left: return True
            
            best_slot = min(slots_left, key=count_candidates)
            d, role = best_slot
            remaining_slots = [s for s in slots_left if s != best_slot]
            
            cands = [n for n in config.roster if can_assign_primary(d, role, n, strict_mode)]
            rng.shuffle(cands)
            cands.sort(key=lambda n: score_primary(d, role, n), reverse=True)
            
            for p in cands:
                schedule[d.isoformat()][role] = p
                role_counts[role][p] += 1
                hrs = hours_for_role(d, role, hols)
                total_hours[p] += hrs
                if role == "PN":
                    wd = weekday_sun0(d)
                    pn_weekday_counts[p][wd] += 1
                    if wd == 1: monday_pn_used[p] += 1
                
                if backtrack(remaining_slots): return True
                
                del schedule[d.isoformat()][role]
                role_counts[role][p] -= 1
                total_hours[p] -= hrs
                if role == "PN":
                    wd = weekday_sun0(d)
                    pn_weekday_counts[p][wd] -= 1
                    if wd == 1: monday_pn_used[p] -= 1
            
            return False

        return backtrack(primary_slots)

    print("Attempting STRICT solve (Pass 1)...")
    reset_state()
    if not run_solver(strict_mode=True):
        print(">> Strict solve failed. Switching to RELAXED solve (Pass 2)...")
        reset_state()
        if not run_solver(strict_mode=False):
            raise RuntimeError("❌ Solver failed even in relaxed mode. Constraints are too tight.")
        else:
            print(">> Relaxed solve successful.")
    else:
        print(">> Strict solve successful.")

    
    # --- BU1 fairness (treat BU1 as a primary-like role) ---
    # We compute monthly BU1 targets and bias/limit BU1 assignment to keep it balanced,
    # especially across Mondays (avoid the same person always being BU1 on Mondays).
    bu1_opps = 0
    for _d in days:
        if _d in hols or is_sunday(_d):
            continue
        # Determine if BU1 is expected on this day (same rules as BU fill)
        if _d in push_days:
            want_bu1 = True
        elif weekday_sun0(_d) == 1:
            want_bu1 = True
        elif is_weekday(_d):
            want_bu1 = True
        else:
            want_bu1 = False
        if want_bu1:
            bu1_opps += 1

    bu1_targets = {p: 0 for p in config.roster}
    if bu1_opps > 0:
        base = bu1_opps // len(config.roster)
        rem = bu1_opps % len(config.roster)
        for p in config.roster:
            bu1_targets[p] = base
        # Remainder uses sales_order (already derived from Sales Ranking) for stable tie-breaking
        roster_set = set(config.roster)
        order = [x for x in config.sales_order if x in roster_set] + [x for x in config.roster if x not in config.sales_order]
        for i in range(rem):
            bu1_targets[order[i]] += 1

    bu1_counts = {p: 0 for p in config.roster}
    bu1_weekday_counts = {p: {i: 0 for i in range(7)} for p in config.roster}
    enforce_bu1_cap = True
    # --- end BU1 fairness ---

    # --- BU FILL ---
    def ideal_bu_roles_for_day(d: date) -> List[str]:
        if d in hols:
            return []
        if d in push_days:
            return ["BU1", "BU2", "BU3", "BU4"] if is_saturday(d) else ["BU1", "BU2", "BU3"]
        if weekday_sun0(d) == 1:
            return ["BU1", "BU2", "BU3"]  # Mon
        if is_weekday(d):
            return ["BU1", "BU2"]  # Tue-Fri
        return []

    bu_queue = []
    for d in days:
        ideal = ideal_bu_roles_for_day(d)
        
        if not ideal: continue

        assigned = set(schedule[d.isoformat()].values())
        hard = cons.hard_off.get(d, set())
        pool = sorted([p for p in config.roster if p not in assigned and p not in hard])
        
        temp_pool = pool[:]
        for rname in ideal:
            if not temp_pool: break
            
            valid_cands = [p for p in temp_pool if can_bu(d, rname, p)]
            if not valid_cands: continue

            bu_queue.append((d, rname))
            person_to_reserve = valid_cands[0]
            temp_pool.remove(person_to_reserve)

    def bu_score(d, role, p):
        s = rng.random()
        s -= total_hours[p] * 0.5
        s -= pp_hours(p, d) * 1.0
        if role == 'BU1':
            # prefer people under their BU1 target
            s += (bu1_targets[p] - bu1_counts[p]) * 50.0
            # spread BU1 across weekdays
            s -= bu1_weekday_counts[p][weekday_sun0(d)] * 10.0
            # extra spread on Mondays (avoid the same BU1 on Mondays)
            if weekday_sun0(d) == 1:
                s -= bu1_weekday_counts[p][1] * 25.0
        if p in cons.avoid_roles.get((d, role), set()): s -= 1000.0
        # Soft B2B preference for BU: discourage repeats but NEVER block filling BU.
        prev_d = d - timedelta(days=1)
        prev_roles = schedule.get(prev_d.isoformat(), {})
        if prev_d.month != month:
            prev_roles = prev_sched.get(prev_d.isoformat(), {})
        had_any_bu_yday = any(k.startswith('BU') and v == p for k, v in prev_roles.items())
        had_same_bu_slot_yday = (prev_roles.get(role) == p)
        if had_any_bu_yday:
            s -= 5.0   # small nudge away from consecutive BU days
        if had_same_bu_slot_yday:
            s -= 10.0  # safety nudge if can_bu rules are ever relaxed again
        return s

    def solve_bu(idx):
        if idx >= len(bu_queue): return True
        d, role = bu_queue[idx]
        cands = [p for p in config.roster if can_bu(d, role, p)]
        rng.shuffle(cands)
        cands.sort(key=lambda p: bu_score(d, role, p), reverse=True)
        
        for p in cands:
            schedule[d.isoformat()][role] = p
            total_hours[p] += 8
            if role == 'BU1':
                bu1_counts[p] += 1
                bu1_weekday_counts[p][weekday_sun0(d)] += 1

            if solve_bu(idx + 1):
                return True

            del schedule[d.isoformat()][role]
            total_hours[p] -= 8
            if role == 'BU1':
                bu1_counts[p] -= 1
                bu1_weekday_counts[p][weekday_sun0(d)] -= 1

        return False

    if not solve_bu(0):
        if enforce_bu1_cap:
            enforce_bu1_cap = False
            # Clear existing BU assignments and retry without the BU1 cap (still respects all hour/dayoff rules).
            for d_iso2 in list(schedule.keys()):
                for k in ['BU1','BU2','BU3','BU4']:
                    schedule[d_iso2].pop(k, None)
            for p2 in bu1_counts:
                bu1_counts[p2] = 0
            for p2 in bu1_weekday_counts:
                for wd in bu1_weekday_counts[p2]:
                    bu1_weekday_counts[p2][wd] = 0
            if not solve_bu(0):
                print("Warning: Could not fill all desired BU slots.")
        else:
            print("Warning: Could not fill all desired BU slots.")
        # --- Greedy fallback BU fill (best effort) ---
        # If full backtracking can't fill every BU slot, keep what we can.
        for d, role in bu_queue:
            d_iso = d.isoformat()
            if role in schedule[d_iso]:
                continue
            cands = [person for person in config.roster if can_bu(d, role, person)]
            if not cands:
                continue
            cands.sort(key=lambda person: bu_score(d, role, person), reverse=True)
            pick = cands[0]
            schedule[d_iso][role] = pick
            total_hours[pick] += 8
            if role == 'BU1':
                bu1_counts[pick] += 1
                bu1_weekday_counts[pick][weekday_sun0(d)] += 1
        # --- end greedy fallback ---

    # --- BU REPAIR PASS (BU2/BU3 only) ---
    bu_repair_swaps = []
    deficits = []
    for d in days:
        for role in ideal_bu_roles_for_day(d):
            if role in {'BU2', 'BU3'} and role not in schedule[d.isoformat()]:
                deficits.append((d, role))
    before_deficits = len(deficits)

    for d, role in deficits:
        d_iso = d.isoformat()
        if role in schedule[d_iso]:
            continue

        ws = week_start_sun(d)
        is_relaxed_week = week_contains_push(ws, push_days) or week_contains_holiday(ws, hols)

        # Easy win first: direct assignment if someone already passes all checks.
        direct_cands = [p for p in config.roster if can_bu(d, role, p)]
        if direct_cands:
            direct_cands.sort(key=lambda person: bu_score(d, role, person), reverse=True)
            pick = direct_cands[0]
            schedule[d_iso][role] = pick
            total_hours[pick] += 8
            bu_repair_swaps.append({
                "type": "direct",
                "deficit_day": d_iso,
                "deficit_role": role,
                "assigned": pick,
            })
            continue

        if is_relaxed_week:
            continue

        # Single-swap augmenting attempt (same-week only).
        week_days = [ws + timedelta(days=i) for i in range(7)]

        fixed = False
        for p in config.roster:
            # Candidate for deficit day must be free and not hard-off.
            if p in schedule[d_iso].values():
                continue
            if p in cons.hard_off.get(d, set()):
                continue

            # Try to free one BU2/BU3 assignment for p inside the same week.
            move_options = []
            for d2 in week_days:
                d2_iso = d2.isoformat()
                if d2_iso not in schedule:
                    continue
                for moved_role in ('BU2', 'BU3'):
                    if schedule[d2_iso].get(moved_role) == p:
                        move_options.append((d2, moved_role))

            if not move_options:
                continue

            for d2, moved_role in move_options:
                d2_iso = d2.isoformat()

                # Find receiver q for p's moved BU slot.
                q_cands = []
                for q in config.roster:
                    if q == p:
                        continue
                    if q in schedule[d2_iso].values():
                        continue
                    if q in cons.hard_off.get(d2, set()):
                        continue
                    if can_bu(d2, moved_role, q):
                        q_cands.append(q)

                if not q_cands:
                    continue

                q_cands.sort(key=lambda person: bu_score(d2, moved_role, person), reverse=True)

                for q in q_cands:
                    # Tentative move: d2 moved_role from p -> q
                    schedule[d2_iso][moved_role] = q
                    total_hours[p] -= 8
                    total_hours[q] += 8

                    # After freeing p's load, try filling deficit.
                    if can_bu(d, role, p):
                        schedule[d_iso][role] = p
                        total_hours[p] += 8
                        bu_repair_swaps.append({
                            "type": "swap",
                            "deficit_day": d_iso,
                            "deficit_role": role,
                            "filled_by": p,
                            "moved_day": d2_iso,
                            "moved_role": moved_role,
                            "moved_from": p,
                            "moved_to": q,
                        })
                        fixed = True
                        break

                    # Rollback tentative move.
                    schedule[d2_iso][moved_role] = p
                    total_hours[p] += 8
                    total_hours[q] -= 8

                if fixed:
                    break
            if fixed:
                break

    remaining_deficits = []
    for d in days:
        for role in ideal_bu_roles_for_day(d):
            if role in {'BU2', 'BU3'} and role not in schedule[d.isoformat()]:
                remaining_deficits.append((d, role))
    after_deficits = len(remaining_deficits)
    print("\n--- BU Repair Pass Summary ---")
    print(f"BU2/BU3 deficits before repair: {before_deficits}")
    print(f"BU2/BU3 deficits after repair: {after_deficits}")
    if bu_repair_swaps:
        for item in bu_repair_swaps:
            if item["type"] == "direct":
                print(f"[DIRECT] Filled {item['deficit_role']} on {item['deficit_day']} with {item['assigned']}")
            else:
                print(
                    f"[SWAP] Filled {item['deficit_role']} on {item['deficit_day']} with {item['filled_by']} "
                    f"by moving {item['moved_role']} on {item['moved_day']} from {item['moved_from']} to {item['moved_to']}"
                )
    else:
        print("No BU2/BU3 repair swaps performed.")

    def weekday_target_repair_pass(
        schedule,
        prev_sched,
        cons,
        roster,
        year,
        month,
        hols,
        push_days,
        payperiods,
        target_weekday_staff=5,
        verbose=False,
    ):
        repair_log = []
        improved_days = 0
        moves_performed = 0
        swap_attempts = 0
        max_swap_attempts = 200

        def expected_bu_roles_for_day(d: date) -> List[str]:
            return ideal_bu_roles_for_day(d)

        def desired_role_for_repair(d: date) -> Optional[str]:
            expected_roles = expected_bu_roles_for_day(d)
            d_iso = d.isoformat()
            if "BU1" in expected_roles and "BU1" not in schedule[d_iso]:
                return "BU1"
            for role in ("BU2", "BU3"):
                if role in expected_roles and role not in schedule[d_iso]:
                    return role
            return None

        def normalize_bu1_skip(d: date):
            d_iso = d.isoformat()
            expected_roles = expected_bu_roles_for_day(d)
            if "BU1" not in expected_roles:
                return
            if "BU1" in schedule[d_iso]:
                return
            if "BU2" in schedule[d_iso]:
                schedule[d_iso]["BU1"] = schedule[d_iso]["BU2"]
                del schedule[d_iso]["BU2"]

        for d in month_days(year, month):
            if d in hols or not is_weekday(d):
                continue
            available_count = len(roster) - len(cons.hard_off.get(d, set()))
            if available_count < target_weekday_staff:
                continue

            d_iso = d.isoformat()
            assigned_today = set(schedule[d_iso].values())
            pre_unique = len(assigned_today)
            if pre_unique >= target_weekday_staff:
                continue

            desired_role = desired_role_for_repair(d)
            if desired_role is None:
                continue

            available_today = [
                p for p in roster
                if p not in cons.hard_off.get(d, set()) and p not in assigned_today
            ]
            if not available_today:
                continue

            direct_success = False
            direct_cands = [p for p in available_today if can_bu(d, desired_role, p)]
            if direct_cands:
                direct_cands.sort(key=lambda person: bu_score(d, desired_role, person), reverse=True)
                pick = direct_cands[0]
                schedule[d_iso][desired_role] = pick
                total_hours[pick] += 8
                normalize_bu1_skip(d)
                post_unique = len(set(schedule[d_iso].values()))
                if post_unique > pre_unique:
                    improved_days += 1
                moves_performed += 1
                repair_log.append(
                    f"direct: assigned {pick} to {desired_role} on {d_iso} ({pre_unique}->{post_unique})"
                )
                direct_success = True

            if direct_success:
                continue

            ws = week_start_sun(d)
            week_days = [ws + timedelta(days=i) for i in range(7)]
            fixed = False
            for x in available_today:
                if fixed:
                    break
                move_options = []
                for d2 in week_days:
                    d2_iso = d2.isoformat()
                    if d2_iso not in schedule:
                        continue
                    for r2 in ("BU1", "BU2", "BU3", "BU4"):
                        if schedule[d2_iso].get(r2) == x:
                            move_options.append((d2, r2))

                for d2, r2 in move_options:
                    if swap_attempts >= max_swap_attempts:
                        break
                    d2_iso = d2.isoformat()
                    y_cands = []
                    for y in roster:
                        if y == x:
                            continue
                        if y in schedule[d2_iso].values():
                            continue
                        if y in cons.hard_off.get(d2, set()):
                            continue
                        if can_bu(d2, r2, y):
                            y_cands.append(y)

                    y_cands.sort(key=lambda person: bu_score(d2, r2, person), reverse=True)
                    for y in y_cands:
                        if swap_attempts >= max_swap_attempts:
                            break
                        swap_attempts += 1
                        schedule[d2_iso][r2] = y
                        total_hours[x] -= 8
                        total_hours[y] += 8

                        if can_bu(d, desired_role, x):
                            schedule[d_iso][desired_role] = x
                            total_hours[x] += 8
                            normalize_bu1_skip(d)
                            post_unique = len(set(schedule[d_iso].values()))
                            if post_unique > pre_unique:
                                improved_days += 1
                            moves_performed += 2
                            repair_log.append(
                                f"swap: moved {r2} on {d2_iso} from {x} -> {y} to free {x} for {desired_role} on {d_iso} ({pre_unique}->{post_unique})"
                            )
                            fixed = True
                            break

                        schedule[d2_iso][r2] = x
                        total_hours[x] += 8
                        total_hours[y] -= 8

                    if fixed:
                        break

                if swap_attempts >= max_swap_attempts:
                    break

        summary = {
            "improved_days": improved_days,
            "moves_performed": moves_performed,
            "log": repair_log,
        }
        return schedule, summary

    if config.enable_weekday_target_repair:
        schedule, repair_log = weekday_target_repair_pass(
            schedule=schedule,
            prev_sched=prev_sched,
            cons=cons,
            roster=config.roster,
            year=year,
            month=month,
            hols=hols,
            push_days=push_days,
            payperiods=payperiods,
            target_weekday_staff=config.target_weekday_staff,
            verbose=config.weekday_target_repair_verbose,
        )
        print(
            "\n--- Weekday Target Repair Summary ---\n"
            f"Weekdays improved: {repair_log['improved_days']}\n"
            f"BU moves performed: {repair_log['moves_performed']}"
        )
        if config.weekday_target_repair_verbose and repair_log["log"]:
            for action in repair_log["log"][:5]:
                print(f"[weekday-target-repair] {action}")

    # print("Warning: Could not fill all desired BU slots.")
    print("\n--- Schedule Fairness Audit ---")
    print(f"{'Name':<10} {'PN (Tgt)':<10} {'AN':<5} {'W':<5} {'TotHrs':<8}")
    for p in config.roster:
        pn = role_counts["PN"][p]
        an = role_counts["AN"][p]
        w = role_counts["W"][p]
        tgt = targets["PN"][p]
        print(f"{p:<10} {pn} ({tgt})      {an:<5} {w:<5} {total_hours[p]:<8}")

    return schedule


def validate_min_weekday_staff(
    schedule: Dict[str, Dict[str, str]],
    year: int,
    month: int,
    hols: Set[date],
    min_staff: int = 4,
) -> Tuple[bool, str]:
    """Validate minimum unique weekday staffing coverage for the built schedule."""
    for day in month_days(year, month):
        if day in hols or not is_weekday(day):
            continue

        day_key = day.isoformat()
        day_roles = schedule.get(day_key, {})
        assigned_people = set(day_roles.values())
        if len(assigned_people) < min_staff:
            return (
                False,
                f"Weekday understaffed: {day_key} has {len(assigned_people)} unique people (<{min_staff}); roles={day_roles}",
            )

    return True, "OK"


def weekday_staff_metrics(
    schedule: Dict[str, Dict[str, str]],
    year: int,
    month: int,
    hols: Set[date],
    roster: List[str],
    cons: Constraints,
    target: int = 5,
) -> Dict[str, int]:
    eligible_weekdays_checked = 0
    eligible_days_with_target_or_more = 0
    eligible_min_assigned_on_weekday: Optional[int] = None
    eligible_deficit_sum_to_target = 0

    for d in month_days(year, month):
        if d in hols or not is_weekday(d):
            continue
        available_count = len(roster) - len(cons.hard_off.get(d, set()))
        if available_count < target:
            continue

        eligible_weekdays_checked += 1
        assigned_count = len(set(schedule.get(d.isoformat(), {}).values()))
        if (
            eligible_min_assigned_on_weekday is None
            or assigned_count < eligible_min_assigned_on_weekday
        ):
            eligible_min_assigned_on_weekday = assigned_count
        if assigned_count >= target:
            eligible_days_with_target_or_more += 1
        eligible_deficit_sum_to_target += max(0, target - assigned_count)

    return {
        "eligible_weekdays_checked": eligible_weekdays_checked,
        "eligible_days_with_target_or_more": eligible_days_with_target_or_more,
        "eligible_deficit_sum_to_target": eligible_deficit_sum_to_target,
        "eligible_min_assigned_on_weekday": (
            eligible_min_assigned_on_weekday if eligible_min_assigned_on_weekday is not None else 0
        ),
    }


def expected_bu_roles_for_day(d: date, hols: Set[date], push_days: Set[date]) -> List[str]:
    if d in hols:
        return []
    if is_sunday(d):
        return []
    if d in push_days:
        if is_saturday(d):
            return ["BU1", "BU2", "BU3", "BU4"]
        return ["BU1", "BU2", "BU3"]
    if weekday_sun0(d) == 1:
        return ["BU1", "BU2", "BU3"]
    if is_weekday(d):
        return ["BU1", "BU2"]
    return []


def bu_deficit_metrics(
    schedule: Dict[str, Dict[str, str]],
    year: int,
    month: int,
    hols: Set[date],
    push_days: Set[date],
) -> Dict[str, int]:
    total_bu_missing = 0
    bu2_bu3_missing = 0
    bu3_missing = 0

    for d in month_days(year, month):
        day_roles = schedule.get(d.isoformat(), {})
        for role in expected_bu_roles_for_day(d, hols, push_days):
            assigned = day_roles.get(role, "")
            if not str(assigned).strip():
                total_bu_missing += 1
                if role in {"BU2", "BU3"}:
                    bu2_bu3_missing += 1
                if role == "BU3":
                    bu3_missing += 1

    return {
        "total_bu_missing": total_bu_missing,
        "bu2_bu3_missing": bu2_bu3_missing,
        "bu3_missing": bu3_missing,
    }


def weekly_over40_events(
    schedule: Dict[str, Dict[str, str]],
    prev_sched: Dict[str, Dict[str, str]],
    roster: List[str],
    year: int,
    month: int,
    hols: Set[date],
    push_days: Set[date],
) -> int:
    events = 0
    week_starts = sorted({week_start_sun(d) for d in month_days(year, month)})
    for ws in week_starts:
        relaxed = week_contains_push(ws, push_days) or week_contains_holiday(ws, hols)
        if relaxed:
            continue
        for person in roster:
            total_hours = 0
            for i in range(7):
                dd = ws + timedelta(days=i)
                total_hours += calculate_daily_hours(dd, person, schedule, prev_sched, year, month)
            if total_hours > 40:
                events += 1
    return events


def payperiod_over80_events(
    schedule: Dict[str, Dict[str, str]],
    prev_sched: Dict[str, Dict[str, str]],
    roster: List[str],
    payperiods: Optional[List[Tuple[date, date]]],
    year: int,
    month: int,
    push_days: Set[date],
) -> int:
    if payperiods is None:
        return 0

    events = 0
    for s, e in payperiods:
        pp_relaxed = False
        cur_wk = week_start_sun(s)
        while cur_wk <= e:
            if week_contains_push(cur_wk, push_days):
                pp_relaxed = True
                break
            cur_wk += timedelta(days=7)

        if pp_relaxed:
            continue

        for person in roster:
            hours = 0
            cur_day = s
            while cur_day <= e:
                hours += calculate_daily_hours(cur_day, person, schedule, prev_sched, year, month)
                cur_day += timedelta(days=1)
            if hours > 80:
                events += 1

    return events


def validate_target_when_available(
    schedule: Dict[str, Dict[str, str]],
    year: int,
    month: int,
    roster: List[str],
    cons: Constraints,
    hols: Set[date],
    target: int,
) -> Tuple[bool, str]:
    for d in month_days(year, month):
        if d in hols or not is_weekday(d):
            continue
        available = len(roster) - len(cons.hard_off.get(d, set()))
        if available >= target:
            assigned_unique = len(set(schedule.get(d.isoformat(), {}).values()))
            if assigned_unique < target:
                return (
                    False,
                    f"Target weekday staffing missed on {d.isoformat()}: assigned {assigned_unique} < {target} "
                    f"with {available} available; roles={schedule.get(d.isoformat(), {})}",
                )
    return True, "OK"


# ----------------------------
# Violations & Export
# ----------------------------

def audit_hours_caps(ws_vio, roster, week_starts, payperiods, hols, push_days, schedule, prev_sched, year, month):
    """Audit Weekly and Pay Period hour caps, respecting relaxations."""
    
    # 1. Weekly check
    for ws_date in week_starts:
        is_relaxed = week_contains_holiday(ws_date, hols) or week_contains_push(ws_date, push_days)
        if is_relaxed: continue
        
        for p in roster:
            tot = 0
            for i in range(7):
                dd = ws_date + timedelta(days=i)
                tot += calculate_daily_hours(dd, p, schedule, prev_sched, year, month)
            if tot > 40:
                ws_vio.append(["Weekly Cap Violation", p, f"{tot}h in week of {ws_date} (Limit 40)"])

    # 2. PP Check
    if payperiods:
        for s, e in payperiods:
            # Relax if PP contains ANY push week (start)
            pp_relaxed = False
            cur_wk = week_start_sun(s)
            while cur_wk <= e:
                if week_contains_push(cur_wk, push_days):
                    pp_relaxed = True
                    break
                cur_wk += timedelta(days=7)
            
            if pp_relaxed: continue
            
            for p in roster:
                h = 0
                cur = s
                while cur <= e:
                    h += calculate_daily_hours(cur, p, schedule, prev_sched, year, month)
                    cur += timedelta(days=1)
                if h > 80:
                    ws_vio.append(["PayPeriod Cap Violation", p, f"{h}h in PP {s}:{e} (Limit 80)"])


def write_violations_tab(wb, schedule, year, month, roster, targets, cons, prev_sched, payperiods, push_days, hols, week_starts):
    ws_vio = wb.create_sheet("Violations")
    ws_vio["A1"] = "Rule Violations / Audits"
    ws_vio["A1"].font = Font(bold=True, size=14)
    ws_vio.append(["Type", "Person", "Details"])
    for cell in ws_vio[2]: cell.font = Font(bold=True)

    # 1. Target Audit
    real_counts = {r: {p: 0 for p in roster} for r in ["PN", "AN", "W"]}
    for d_iso, rmap in schedule.items():
        for r in ["PN", "AN", "W"]:
            p = rmap.get(r)
            if p: real_counts[r][p] += 1
    for p in roster:
        for r in ["PN", "AN", "W"]:
            tgt = targets[r][p]
            act = real_counts[r][p]
            if act != tgt:
                ws_vio.append(["Target Mismatch", p, f"{r} Count {act} != Target {tgt}"])

    # 2. Daily Constraints (Unified Loop)
    sorted_dates = sorted(schedule.keys())
    
    for d_iso in sorted_dates:
        d = date.fromisoformat(d_iso)
        rmap = schedule[d_iso]
        
        roles_by_person = {} # p -> list of roles
        hard_off_names = cons.hard_off.get(d, set())
        
        for r, p in rmap.items():
            roles_by_person.setdefault(p, []).append(r)

            # B. Hard Off
            if p in hard_off_names:
                ws_vio.append(["Hard Off Violation", p, f"Worked {r} on {d} (Requested Off)"])
            
            # C. Avoid Role
            avoid_list = cons.avoid_roles.get((d, r), set())
            if p in avoid_list:
                rtype = "Avoid Violation (Hard)" if r in ROLE_PRIMARY else "Avoid Warning (Soft)"
                ws_vio.append([rtype, p, f"Assigned {r} on {d}"])
                
            # D. Back-to-Back (Same Role)
            prev_d = d - timedelta(days=1)
            prev_roles = schedule.get(prev_d.isoformat(), {})
            if prev_d.month != month:
                prev_roles = prev_sched.get(prev_d.isoformat(), {})
            
            if prev_roles.get(r) == p:
                ws_vio.append(["Back-to-Back (Same Role)", p, f"Worked {r} on {prev_d} and {d}"])

        # A. Duplicate Assignment (Aggregated)
        for p, roles in roles_by_person.items():
            if len(roles) > 1:
                role_str = ", ".join(roles)
                ws_vio.append(["Duplicate Assignment", p, f"Assigned {role_str} on {d}"])

    # 3. Hours Audit
    audit_hours_caps(ws_vio, roster, week_starts, payperiods, hols, push_days, schedule, prev_sched, year, month)

    # 4. Sunday Duplicate Audit
    sunday_counts = {p: 0 for p in roster}
    for d_iso, rmap in schedule.items():
        d = date.fromisoformat(d_iso)
        if is_sunday(d) and rmap.get("PN"):
            sunday_counts[rmap["PN"]] += 1
    for person, cnt in sunday_counts.items():
        if cnt > 1:
            ws_vio.append(["Sunday Duplicate", person, f"Assigned Sunday PN {cnt} times in {year}-{month:02d}"])


def export_to_excel(schedule, year, month, out_path, roster, hols, prev_sched,
                    payperiods, sales_order, cons, template_xlsx=None):
    """Template-first calendar export.

    Preserves sizing/borders/merges/fonts from template_xlsx; writes ONLY values + fills.
    Also writes Weekly Hours / Summary / Pay Period Hours / Violations.
    """
    import openpyxl
    import calendar as _cal
    from datetime import date as _date, timedelta as _td
    from openpyxl.styles import PatternFill, Font, Alignment

    if not template_xlsx:
        raise RuntimeError("Provide --template-xlsx (e.g., Month Template.xlsx) to preserve the proper calendar look.")

    wb = openpyxl.load_workbook(template_xlsx)
    ws = wb.active
    ws.title = _cal.month_name[month]
    ws.sheet_view.showGridLines = False

    ws["B1"] = f"{_cal.month_name[month]} {year}"

    first = _date(year, month, 1)
    last = _date(year, month, _cal.monthrange(year, month)[1])
    offset = weekday_sun0(first)
    weeks = (offset + last.day + 6) // 7
    grid_weeks = 6 if weeks > 5 else 5

    white_fill = PatternFill("solid", fgColor=WHITE_FILL_HEX)
    sunday_fill = PatternFill("solid", fgColor=SUNDAY_FILL)
    nonmonth_fill = PatternFill("solid", fgColor=NONMONTH_FILL)

    for wk in range(grid_weeks):
        R = GRID_WEEK_START_ROW + WEEK_HEIGHT * wk
        for dow in range(7):
            col_start = SUN_START_COL + DAY_WIDTH * dow
            day_num = (wk * 7 + dow) - offset + 1
            in_month = (1 <= day_num <= last.day)

            if in_month:
                d = _date(year, month, day_num)
                fill = sunday_fill if is_sunday(d) else white_fill
            else:
                d = None
                fill = nonmonth_fill

            for rr in range(R, R + WEEK_HEIGHT):
                for cc in range(col_start, col_start + DAY_WIDTH):
                    cell = ws.cell(rr, cc)
                    cell.value = None
                    cell.fill = fill

            if not in_month:
                continue

            roles = schedule.get(d.isoformat(), {})

            dn = ws.cell(R, col_start + 2, day_num)
            dn.font = FONT_DAY_NUM
            dn.alignment = Alignment(horizontal="right", vertical="top")

            ws.cell(R + 1, col_start, "PN-")
            if not is_sunday(d):
                ws.cell(R + 2, col_start, "AN-")
            if is_weekday(d):
                ws.cell(R + 3, col_start, "W-")

            ws.cell(R + 1, col_start + 1, roles.get("PN", ""))
            if not is_sunday(d):
                ws.cell(R + 2, col_start + 1, roles.get("AN", ""))
            if is_weekday(d):
                ws.cell(R + 3, col_start + 1, roles.get("W", ""))

            if not is_sunday(d):
                ws.cell(R + 4, col_start + 1, roles.get("BU1", ""))
                ws.cell(R + 5, col_start + 1, roles.get("BU2", ""))
                ws.cell(R + 6, col_start + 1, roles.get("BU3", ""))
                ws.cell(R + 7, col_start + 1, roles.get("BU4", ""))

    for nm in ["Weekly Hours", "Summary", "Pay Period Hours", "Violations"]:
        if nm in wb.sheetnames:
            del wb[nm]

    ws_weekly = wb.create_sheet("Weekly Hours")
    ws_weekly.sheet_view.showGridLines = False
    ws_weekly["A1"] = "Weekly Hours (Sun-Sat)"
    ws_weekly["A1"].font = Font(bold=True, size=14)
    ws_weekly.append(["Week Start"] + roster)

    week_starts = sorted({week_start_sun(d) for d in month_days(year, month)})
    ot_fill = PatternFill("solid", fgColor=OVERTIME_FILL)
    for ws_date in week_starts:
        vals = []
        for person in roster:
            tot = 0
            for i in range(7):
                dd = ws_date + _td(days=i)
                tot += calculate_daily_hours(dd, person, schedule, prev_sched, year, month)
            vals.append(tot)
        ws_weekly.append([ws_date.isoformat()] + vals)
        r = ws_weekly.max_row
        for i, v in enumerate(vals):
            if v > 40:
                ws_weekly.cell(row=r, column=2 + i).fill = ot_fill

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Summary"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append(["Name"] + ROLE_PRIMARY + ROLE_BUS + ["Total", "Hours"])

    for person in roster:
        counts = {r: 0 for r in ROLE_PRIMARY + ROLE_BUS}
        tot_hrs = 0
        for d_iso, rmap in schedule.items():
            for r, nm in rmap.items():
                if nm == person:
                    counts[r] += 1
                    tot_hrs += hours_for_role(_date.fromisoformat(d_iso), r, hols)
        tot_ass = sum(counts[r] for r in ROLE_PRIMARY + ROLE_BUS)
        ws2.append([person] + [counts[r] for r in ROLE_PRIMARY + ROLE_BUS] + [tot_ass, tot_hrs])

    if payperiods:
        ws3 = wb.create_sheet("Pay Period Hours")
        ws3.append(["Pay Period", "Start", "End"] + roster)
        for i, (s, e) in enumerate(payperiods, 1):
            row = [f"PP{i}", s.isoformat(), e.isoformat()]
            for person in roster:
                h = 0
                cur = s
                while cur <= e:
                    h += calculate_daily_hours(cur, person, schedule, prev_sched, year, month)
                    cur += _td(days=1)
                row.append(h)
            ws3.append(row)

    push_days = push_days_for_month(year, month)
    targets = compute_primary_targets(month_days(year, month), roster, sales_order, hols)
    write_violations_tab(wb, schedule, year, month, roster, targets, cons,
                         prev_sched, payperiods, push_days, hols, week_starts)
    wb.save(out_path)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", required=True)
    ap.add_argument("--timeoff-xlsx", required=True)
    ap.add_argument("--timeoff-sheet", required=True)
    ap.add_argument("--prev-schedule", required=True)
    ap.add_argument("--sales-sheet", default="Sales Ranking")
    ap.add_argument("--roster-json")
    ap.add_argument("--payperiods-json")
    ap.add_argument("--payperiod-anchor-start", default="2026-01-25")
    ap.add_argument("--out-json")
    ap.add_argument("--out-xlsx")
    ap.add_argument("--template-xlsx", help="Blank month template .xlsx (preserve formatting; write values/fills)")
    ap.add_argument("--seed", type=int, default=11)
    ap.add_argument("--auto-seed", action="store_true", help="Try many seeds internally until staffing validator passes")
    ap.add_argument("--seed-start", type=int, default=1)
    ap.add_argument("--seed-end", type=int, default=50, help="Default 50 (caps attempts); increase if you want deeper search.")
    ap.add_argument("--max-attempts", type=int, default=50, help="Default 50 (caps attempts); use with --seed-start to cap range.")
    ap.add_argument("--auto-seed-verbose", action="store_true", help="Print each auto-seed attempt and failure reason")
    ap.add_argument("--min-weekday-staff", type=int, default=4)
    ap.add_argument("--target-weekday-staff", type=int, default=5)
    ap.add_argument("--optimize-weekday-staff", action="store_true")
    ap.add_argument("--weekday-target-repair", action="store_true",
                    help="After BU fill, attempt BU-only micro-swaps within the same week to raise eligible weekdays to target staffing.")
    ap.add_argument("--weekday-target-repair-verbose", action="store_true")
    ap.add_argument("--enforce-target-when-available", action="store_true")
    args = ap.parse_args()

    year, month = parse_month_arg(args.month)
    roster = load_roster_from_json(args.roster_json) if args.roster_json else DEFAULT_ROSTER[:]
    timeoff = load_timeoff_from_xlsx(args.timeoff_xlsx, args.timeoff_sheet)
    cons = compile_constraints(timeoff)
    sales = load_sales_ranking_from_timeoff_xlsx(args.timeoff_xlsx, args.sales_sheet, year, month, roster)
    prev = load_schedule_json(args.prev_schedule)
    if args.payperiods_json:
        pps = load_payperiods_from_json(args.payperiods_json)
    else:
        ms = date(year, month, 1)
        me = date(year, month, calendar.monthrange(year, month)[1])
        anc = datetime.strptime(args.payperiod_anchor_start, "%Y-%m-%d").date()
        raw = generate_payperiods(anc, ms - timedelta(days=28), me + timedelta(days=28))
        pps = [p for p in raw if p[0] <= me and p[1] >= ms]
    hols = observed_holidays_for_year(year)
    push_days = push_days_for_month(year, month)

    if not args.auto_seed:
        cfg = BuildConfig(
            year,
            month,
            roster,
            sales,
            rng_seed=args.seed,
            enable_weekday_target_repair=args.weekday_target_repair,
            weekday_target_repair_verbose=args.weekday_target_repair_verbose,
            target_weekday_staff=args.target_weekday_staff,
        )
        sched = build_schedule(cfg, prev, cons, pps)
    else:
        seed_start = args.seed_start
        seed_end = args.seed_end
        seed_end_overridden = any(
            a == "--seed-end" or a.startswith("--seed-end=")
            for a in sys.argv[1:]
        )
        max_attempts_overridden = any(
            a == "--max-attempts" or a.startswith("--max-attempts=")
            for a in sys.argv[1:]
        )

        if max_attempts_overridden:
            if not seed_end_overridden:
                seed_end = seed_start + args.max_attempts - 1
            else:
                seed_end = min(seed_end, seed_start + args.max_attempts - 1)

        # Guard: prevent empty seed range (e.g., --seed-start > --seed-end)
        if seed_start > seed_end:
            raise ValueError("Resolved seed range is empty: ensure --seed-start <= --seed-end")

        for d in month_days(year, month):
            if d in hols or not is_weekday(d):
                continue
            available_count = len(roster) - len(cons.hard_off.get(d, set()))
            if available_count < args.min_weekday_staff:
                raise RuntimeError(
                    f"Impossible to satisfy min weekday staff on {d.isoformat()}: "
                    f"only {available_count} counselors available due to hard-off requests, need {args.min_weekday_staff}."
                )

        sched = None
        winning_seed = None
        winning_metrics = None
        winning_bu_metrics = None
        winning_weekly_over40 = None
        winning_payperiod_over80 = None
        best_score = None
        last_failure_reason = "No seeds attempted"

        total_required_target_days = 0
        for d in month_days(year, month):
            if d in hols or not is_weekday(d):
                continue
            available_count = len(roster) - len(cons.hard_off.get(d, set()))
            if available_count >= args.target_weekday_staff:
                total_required_target_days += 1

        for seed in range(seed_start, seed_end + 1):
            perfect_found = False
            cfg = BuildConfig(
                year,
                month,
                roster,
                sales,
                rng_seed=seed,
                enable_weekday_target_repair=args.weekday_target_repair,
                weekday_target_repair_verbose=args.weekday_target_repair_verbose,
                target_weekday_staff=args.target_weekday_staff,
            )
            if args.auto_seed_verbose:
                print(f"Trying seed {seed}")
            try:
                candidate = build_schedule(cfg, prev, cons, pps)
            except RuntimeError as e:
                last_failure_reason = f"Seed {seed} failed to solve: {e}"
                if args.auto_seed_verbose:
                    print(last_failure_reason)
                continue

            ok, reason = validate_min_weekday_staff(
                candidate,
                year,
                month,
                hols,
                min_staff=args.min_weekday_staff,
            )
            if not ok:
                last_failure_reason = f"Seed {seed} failed validator: {reason}"
                if args.auto_seed_verbose:
                    print(last_failure_reason)
                continue

            if args.enforce_target_when_available:
                ok_target, reason_target = validate_target_when_available(
                    candidate,
                    year,
                    month,
                    roster,
                    cons,
                    hols,
                    args.target_weekday_staff,
                )
                if not ok_target:
                    last_failure_reason = f"Seed {seed} failed validator: {reason_target}"
                    if args.auto_seed_verbose:
                        print(last_failure_reason)
                    continue

            metrics = weekday_staff_metrics(
                candidate,
                year,
                month,
                hols,
                roster,
                cons,
                target=args.target_weekday_staff,
            )

            if not args.optimize_weekday_staff:
                sched = candidate
                winning_seed = seed
                winning_metrics = metrics
                break

            bu_metrics = bu_deficit_metrics(candidate, year, month, hols, push_days)
            wk_over40 = weekly_over40_events(candidate, prev, roster, year, month, hols, push_days)
            pp_over80 = payperiod_over80_events(candidate, prev, roster, pps, year, month, push_days)

            score = (
                metrics["eligible_days_with_target_or_more"],
                -metrics["eligible_deficit_sum_to_target"],
                -bu_metrics["bu2_bu3_missing"],
                -bu_metrics["total_bu_missing"],
                -wk_over40,
                -pp_over80,
            )
            if best_score is None or score > best_score:
                best_score = score
                sched = candidate
                winning_seed = seed
                winning_metrics = metrics
                winning_bu_metrics = bu_metrics
                winning_weekly_over40 = wk_over40
                winning_payperiod_over80 = pp_over80

            met_required_target_days = 0
            for d in month_days(year, month):
                if d in hols or not is_weekday(d):
                    continue
                available_count = len(roster) - len(cons.hard_off.get(d, set()))
                if available_count < args.target_weekday_staff:
                    continue
                assigned_count = len(set(candidate.get(d.isoformat(), {}).values()))
                if assigned_count >= args.target_weekday_staff:
                    met_required_target_days += 1

            if met_required_target_days == total_required_target_days:
                perfect_found = True

            if perfect_found:
                break

        if sched is None:
            raise RuntimeError(
                f"Auto-seed failed for range [{seed_start}, {seed_end}]. Last failure: {last_failure_reason}"
            )

        print(f"✅ Auto-seed success: seed={winning_seed}")
        if args.optimize_weekday_staff and winning_metrics is not None:
            if winning_metrics["eligible_weekdays_checked"] == 0:
                print(
                    f"Weekday staffing: no eligible weekdays for target={args.target_weekday_staff} "
                    f"this month (availability <{args.target_weekday_staff} on all weekdays)."
                )
            else:
                print(
                    f"Weekday staffing: {winning_metrics['eligible_days_with_target_or_more']}/"
                    f"{winning_metrics['eligible_weekdays_checked']} eligible weekdays meet "
                    f"target={args.target_weekday_staff}, min eligible weekday assigned="
                    f"{winning_metrics['eligible_min_assigned_on_weekday']}"
                )
            if winning_bu_metrics is not None:
                print(
                    "BU deficits: "
                    f"BU2/BU3 missing={winning_bu_metrics['bu2_bu3_missing']}, "
                    f"total BU missing={winning_bu_metrics['total_bu_missing']}"
                )
            if winning_weekly_over40 is not None and winning_payperiod_over80 is not None:
                print(
                    "Hours over caps (non-relaxed): "
                    f"weekly>40 events={winning_weekly_over40}, "
                    f"payperiod>80 events={winning_payperiod_over80}"
                )

    oj = args.out_json or str(Path(args.prev_schedule).with_name(f"{year}_{month:02d}_schedule.json"))
    ox = args.out_xlsx or str(Path(args.prev_schedule).with_name(f"{calendar.month_name[month]}_{year}_Schedule.xlsx"))

    Path(oj).write_text(json.dumps(sched, indent=2), "utf-8")
    print(f"✅ JSON: {oj}")

    export_to_excel(
        sched, year, month, ox, roster, hols,
        prev, pps, sales, cons, template_xlsx=args.template_xlsx
    )
    print(f"✅ Excel: {ox}")


if __name__ == "__main__":
    main()
