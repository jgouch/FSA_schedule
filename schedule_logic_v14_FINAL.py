#!/usr/bin/env python3
"""
schedule_logic_v9.py
Fixes month-boundary constraint gaps identified by Codex:

✅ Cross-month "no same specific role back-to-back" (PN/AN/W and BU1/BU2/BU3/BU4)
✅ Cross-month max consecutive workdays (includes any role) using a lookback window
✅ Cross-month weekly cap hours (Sun–Sat) by seeding week-hours with overlapping prior-month days
✅ Still two-phase: Primaries first (PN/AN/W), then BU backfill (BU is scalable)

Key new concept: CARRYOVER CONTEXT
---------------------------------
When generating a month, pass in prior assignments for the days *immediately before* the month start
(and optionally any days after month end if you want whole pay periods). At minimum, we need:

- prior_days_assignments: dict[date] -> dict[role] -> person
  Include at least the last 7 days before month start so:
    * role back-to-back checks work across boundary
    * consecutive day streak works (max 5 days rule)
    * weekly cap works for week that starts in previous month

You can provide this from the already-generated previous month schedule output.

If you don't provide it, v9 behaves like v8 (month-island), but will WARN you.

No external packages required.
"""

from __future__ import annotations
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path
import calendar
import json
import random
import sys
import time
from collections import defaultdict, Counter
from typing import Dict, List, Optional, Set, Iterable, Any, Tuple


# -----------------------------
# Inputs
# -----------------------------

@dataclass(frozen=True)
class Inputs:
    year: int
    month: int
    fsas: List[str]
    sales_volume: Dict[str, float]
    seniority_order: List[str]
    time_off: Dict[str, Set[date]]
    sunday_rotation_order: List[str]
    sunday_first_assignee: str
    observed_holidays: Set[date]
    push_week_enabled: bool = True
    hours_per_day: int = 8
    sunday_hours: int = 5
    weekly_cap_hours: int = 40
    max_consecutive_days: int = 5
    monday_pn_max_per_month: int = 1

    # Pay periods (optional). If provided, enforce 80-hr cap per pay period.
    pay_periods: List[PayPeriod] = field(default_factory=list)
    pay_period_cap_hours: int = 80

    # NEW: carryover assignments for prior/adjacent dates
    # Example: include last 7-14 days before month start (and optionally overlapping pay-period days).
    carryover_assignments: Dict[date, Dict[str, str]] = field(default_factory=dict)

    # How many days before month start to consider when computing streaks, back-to-back checks, and week hours.
    lookback_days: int = 14


# -----------------------------
# Utilities
# -----------------------------

def daterange(d1: date, d2: date) -> Iterable[date]:
    cur = d1
    while cur <= d2:
        yield cur
        cur += timedelta(days=1)

def month_days(year: int, month: int) -> List[date]:
    _, last = calendar.monthrange(year, month)
    return [date(year, month, d) for d in range(1, last + 1)]

def is_saturday(d: date) -> bool:
    return d.weekday() == 5

def is_sunday(d: date) -> bool:
    return d.weekday() == 6

def week_start_sun(d: date) -> date:
    return d - timedelta(days=(d.weekday() + 1) % 7)

def last_weekdays_block(year: int, month: int) -> List[date]:
    days = month_days(year, month)
    last_day = days[-1]
    last_fri = last_day
    while last_fri.weekday() != 4:
        last_fri -= timedelta(days=1)
    last_mon = last_fri - timedelta(days=4)
    return [last_mon + timedelta(days=i) for i in range(5)]

def hours_for(d: date, role: str, inp: Inputs) -> int:
    if d in inp.observed_holidays:
        return 0
    if is_sunday(d):
        return inp.sunday_hours if role == "PN" else 0
    return inp.hours_per_day

def sort_by_hierarchy(fsas: List[str], sales: Dict[str, float], seniority: List[str]) -> List[str]:
    seniority_rank = {name: i for i, name in enumerate(seniority)}
    return sorted(fsas, key=lambda n: (-sales.get(n, 0.0), seniority_rank.get(n, 10_000)))


# -----------------------------
# Push days & cap waiver
# -----------------------------

def compute_push_days(inp: Inputs) -> Set[date]:
    if not inp.push_week_enabled:
        return set()
    days = month_days(inp.year, inp.month)
    push = set(last_weekdays_block(inp.year, inp.month))
    push.add(max(d for d in days if is_saturday(d)))
    return push

def week_contains_push(wk_start: date, push_days: Set[date]) -> bool:
    return any((wk_start + timedelta(days=i)) in push_days for i in range(7))


# -----------------------------
# Role model
# -----------------------------

PRIMARY_ROLES_ORDER = ["PN", "AN", "W"]
BU_POOL_WEEKDAY = ["BU1", "BU2", "BU3"]
BU_POOL_SAT_PUSH = ["BU1", "BU2", "BU3", "BU4"]

def required_primary_roles(d: date, inp: Inputs) -> List[str]:
    if d in inp.observed_holidays:
        return []
    if is_sunday(d):
        return ["PN"]
    if is_saturday(d):
        return ["PN", "AN"]
    return ["PN", "AN", "W"]

def desired_bu_roles(d: date, inp: Inputs, push_days: Set[date]) -> List[str]:
    if d in inp.observed_holidays:
        return []
    if is_sunday(d):
        return []

    available = [p for p in inp.fsas if d not in inp.time_off.get(p, set())]
    cap = len(available)
    prim_count = len(required_primary_roles(d, inp))

    if is_saturday(d):
        if d in push_days:
            desired = BU_POOL_SAT_PUSH[:]
            allow = max(0, min(len(desired), cap - prim_count))
            return desired[:allow]
        return []

    if d.weekday() == 0 or d in push_days:
        desired = BU_POOL_WEEKDAY[:3]
    else:
        desired = BU_POOL_WEEKDAY[:2]

    allow = max(0, min(len(desired), cap - prim_count))
    return desired[:allow]


def compute_primary_targets(inp: Inputs) -> Dict[str, Dict[str, int]]:
    days = month_days(inp.year, inp.month)
    fsas = inp.fsas
    hier = sort_by_hierarchy(fsas, inp.sales_volume, inp.seniority_order)

    totals = {r: 0 for r in PRIMARY_ROLES_ORDER}
    for d in days:
        for r in required_primary_roles(d, inp):
            if r in totals:
                totals[r] += 1

    def split(total: int) -> Dict[str, int]:
        base = total // len(fsas)
        rem = total % len(fsas)
        out = {p: base for p in fsas}
        for p in hier[:rem]:
            out[p] += 1
        return out

    return {r: split(totals[r]) for r in PRIMARY_ROLES_ORDER}


# -----------------------------
# Cross-month context
# -----------------------------


# -----------------------------
# Pay periods (bi-weekly caps)
# -----------------------------

@dataclass(frozen=True)
class PayPeriod:
    id: str
    start: date  # inclusive
    end: date    # inclusive
    payday: Optional[date] = None

def load_pay_periods(path: str) -> List[PayPeriod]:
    """
    JSON/YAML:
      pay_periods:
        - id: PP4
          start: 2026-01-25
          end: 2026-02-07
          payday: 2026-02-13
    payday is optional.
    """
    data = load_json_or_yaml(path)
    if isinstance(data, dict) and "pay_periods" in data:
        data = data["pay_periods"]
    if not isinstance(data, list):
        raise ValueError("pay periods file must be a list under 'pay_periods'.")
    out: List[PayPeriod] = []
    for row in data:
        if not isinstance(row, dict):
            continue
        pid = str(row.get("id") or row.get("pay_period") or row.get("name") or "")
        if not pid:
            raise ValueError("Each pay period must have an id (e.g., PP4).")
        start = _parse_date(str(row["start"]))
        end = _parse_date(str(row["end"]))
        payday = row.get("payday")
        pd = _parse_date(str(payday)) if payday else None
        out.append(PayPeriod(id=pid, start=start, end=end, payday=pd))
    # sort by start
    out.sort(key=lambda p: p.start)
    return out

def find_pay_period(pay_periods: List[PayPeriod], d: date) -> Optional[PayPeriod]:
    for pp in pay_periods:
        if pp.start <= d <= pp.end:
            return pp
    return None

def pay_period_contains_push(pp: PayPeriod, push_days: Set[date]) -> bool:
    # cap waiver if ANY push day falls inside this pay period window
    return any(pp.start <= pd <= pp.end for pd in push_days)

class Context:
    """
    Provides unified access to assignments on dates outside the month (carryover)
    plus in-month working schedule state.
    """
    def __init__(self, inp: Inputs, in_month_days: List[date]):
        self.inp = inp
        self.days = in_month_days
        self.start = self.days[0]
        self.carry = dict(inp.carryover_assignments or {})

        # Precompute a lookback date window
        self.lookback_start = self.start - timedelta(days=inp.lookback_days)

    def get_roles(self, sched: Dict[date, Dict[str, str]], d: date) -> Dict[str, str]:
        if d in sched:
            return sched[d]
        return self.carry.get(d, {})

    def prev_day(self, d: date) -> date:
        return d - timedelta(days=1)

    def worked_any(self, sched: Dict[date, Dict[str, str]], d: date, person: str) -> bool:
        return person in self.get_roles(sched, d).values()

    def same_role_prev_day(self, sched: Dict[date, Dict[str, str]], d: date, role: str, person: str) -> bool:
        prev = self.prev_day(d)
        if prev < self.lookback_start:
            return False
        return self.get_roles(sched, prev).get(role) == person

    def consecutive_streak_ending(self, sched: Dict[date, Dict[str, str]], d: date, person: str, assume_work_today: bool) -> int:
        """
        Counts consecutive workdays ending at d.
        If assume_work_today is False, it counts streak ending at previous day.
        """
        streak = 0
        cur = d if assume_work_today else (d - timedelta(days=1))
        while cur >= self.lookback_start:
            if self.worked_any(sched, cur, person):
                streak += 1
                cur -= timedelta(days=1)
            else:
                break
        return streak

    
    def pay_period_for(self, d: date) -> Optional[PayPeriod]:
        if not self.inp.pay_periods:
            return None
        return find_pay_period(self.inp.pay_periods, d)

    def seed_pay_period_hours(self, sched: Dict[date, Dict[str, str]], pp: PayPeriod) -> Counter:
        """
        Seed pay-period hours from carryover assignments for days in pp that are before month start.
        """
        c = Counter()
        dd = pp.start
        while dd <= pp.end:
            if dd < self.start and dd >= self.lookback_start:
                roles = self.get_roles(sched, dd)
                for role, p in roles.items():
                    c[p] += hours_for(dd, role, self.inp)
            dd += timedelta(days=1)
        return c

def seed_week_hours(self, sched: Dict[date, Dict[str, str]], week_start: date) -> Counter:
        """
        For a given week_start (Sun), seed hours from carryover assignments for days in that week that
        are BEFORE the month start. This fixes weekly cap over month boundaries.
        """
        c = Counter()
        for i in range(7):
            dd = week_start + timedelta(days=i)
            if dd < self.start and dd >= self.lookback_start:
                roles = self.get_roles(sched, dd)
                for role, p in roles.items():
                    c[p] += hours_for(dd, role, self.inp)
        return c


# -----------------------------
# Phase 1: Primary search
# -----------------------------

class PrimarySolver:
    def __init__(self, inp: Inputs, seed: int = 11):
        self.inp = inp
        self.rng = random.Random(seed)
        self.days = month_days(inp.year, inp.month)
        self.push_days = compute_push_days(inp)
        self.ctx = Context(inp, self.days)

        self._validate_inputs()

        self.sched: Dict[date, Dict[str, str]] = {d: {} for d in self.days}
        # weekly hours seeded from carryover
        self.week_hours: Dict[date, Counter] = defaultdict(Counter)
        self._seed_week_hours_from_carryover()

        # pay period hours (seeded from carryover) if pay periods provided
        self.pp_hours: Dict[str, Counter] = defaultdict(Counter)
        self._seed_pay_period_hours_from_carryover()

        self.mon_pn = Counter()
        self.primary_counts = {r: Counter() for r in PRIMARY_ROLES_ORDER}
        self.hierarchy = sort_by_hierarchy(inp.fsas, inp.sales_volume, inp.seniority_order)
        self.targets = compute_primary_targets(inp)

        self.first_dead_end: Optional[Dict[str, Any]] = None

        if not inp.carryover_assignments:
            print("⚠️  WARNING: No carryover_assignments provided. Month-boundary constraints may be violated.")
            print("   Provide at least the last 7-14 days of prior schedule to enforce cross-month rules.\n")

    def _validate_inputs(self):
        if not self.inp.fsas:
            raise ValueError("Inputs.fsas is empty.")
        if len(set(self.inp.fsas)) != len(self.inp.fsas):
            raise ValueError("Inputs.fsas contains duplicates.")
        rot = self.inp.sunday_rotation_order
        if len(set(rot)) != len(rot):
            raise ValueError("sunday_rotation_order contains duplicates.")
        if set(rot) != set(self.inp.fsas):
            raise ValueError("sunday_rotation_order must match fsas exactly.")
        if self.inp.sunday_first_assignee not in rot:
            raise ValueError("sunday_first_assignee must be in sunday_rotation_order.")

    def _seed_week_hours_from_carryover(self):
        # Seed only for weeks that intersect the month start (and any week for which carryover has days)
        for d in self.days:
            wk = week_start_sun(d)
            if wk not in self.week_hours:
                self.week_hours[wk] = self.ctx.seed_week_hours(self.sched, wk)

    def _seed_pay_period_hours_from_carryover(self):
        if not self.inp.pay_periods:
            return
        # Seed from carryover for any pay period that overlaps this month
        month_start = self.days[0]
        month_end = self.days[-1]
        for pp in self.inp.pay_periods:
            # overlap check
            if pp.end < (month_start - timedelta(days=self.inp.lookback_days)) or pp.start > month_end:
                continue
            self.pp_hours[pp.id] = self.ctx.seed_pay_period_hours(self.sched, pp)


    def _avoid_status(self, person: str, d: date, role: str) -> str:
        avoid = self.inp.role_avoid.get(person, {}).get(d, set())
        if role not in avoid:
            return "none"
        hard = self.inp.role_avoid_hard.get(person, {}).get(d, False)
        return "hard" if hard else "soft"

    def can_assign_primary(self, p: str, d: date, role: str) -> Tuple[bool, str]:
        inp = self.inp

        if d in inp.time_off.get(p, set()):
            return False, "time_off"
        if p in self.sched[d].values():
            return False, "already_today"

        # Cross-month no same role back-to-back
        if self.ctx.same_role_prev_day(self.sched, d, role, p):
            return False, "same_role_back_to_back"

        if role == "PN" and d.weekday() == 0 and self.mon_pn[p] >= inp.monday_pn_max_per_month:
            return False, "monday_pn_limit"

        # Cross-month max consecutive days: if working today would exceed cap
        if self.ctx.consecutive_streak_ending(self.sched, d, p, assume_work_today=True) > inp.max_consecutive_days:
            return False, "max_consecutive_days"

        # weekly cap seeded across month boundary
        wk = week_start_sun(d)
        waive = week_contains_push(wk, self.push_days)
        add_h = hours_for(d, role, inp)
        if (not waive) and (self.week_hours[wk][p] + add_h > inp.weekly_cap_hours):
            return False, "weekly_cap"



        # pay period cap (seeded across month boundary), waived if pay period contains any push day
        pp = self.ctx.pay_period_for(d)
        if pp is not None:
            pp_waive = pay_period_contains_push(pp, self.push_days)
            if (not pp_waive) and (self.pp_hours[pp.id][p] + add_h > inp.pay_period_cap_hours):
                return False, "pay_period_cap"
        # Saturday AN must match Sunday PN if Sunday PN already assigned
        if role == "AN" and is_saturday(d):
            nxt = d + timedelta(days=1)
            if is_sunday(nxt) and (nxt.month == inp.month) and nxt not in inp.observed_holidays:
                sunday_pn = self.sched[nxt].get("PN")
                if sunday_pn is not None and p != sunday_pn:
                    return False, "sat_an_must_match_sunday_pn"

        return True, "ok"

    def assign(self, d: date, role: str, p: str):
        self.sched[d][role] = p
        wk = week_start_sun(d)
        h = hours_for(d, role, self.inp)
        self.week_hours[wk][p] += h
        pp = self.ctx.pay_period_for(d)
        if pp is not None:
            self.pp_hours[pp.id][p] += h
        if role == "PN" and d.weekday() == 0:
            self.mon_pn[p] += 1
        self.primary_counts[role][p] += 1

    def unassign(self, d: date, role: str):
        p = self.sched[d].pop(role, None)
        if p is None:
            return
        wk = week_start_sun(d)
        h = hours_for(d, role, self.inp)
        self.week_hours[wk][p] -= h
        pp = self.ctx.pay_period_for(d)
        if pp is not None:
            self.pp_hours[pp.id][p] -= h
        if role == "PN" and d.weekday() == 0:
            self.mon_pn[p] -= 1
        self.primary_counts[role][p] -= 1

    def set_sundays_by_rotation(self, first_assignee: Optional[str] = None):
        order = self.inp.sunday_rotation_order
        n = len(order)
        start = first_assignee or self.inp.sunday_first_assignee
        idx = order.index(start)

        sundays = [d for d in self.days if is_sunday(d) and d not in self.inp.observed_holidays]
        for d in sundays:
            chosen = None
            for k in range(n):
                cand = order[(idx + k) % n]
                ok, _ = self.can_assign_primary(cand, d, "PN")
                if ok:
                    chosen = cand
                    break
            if chosen is None:
                raise RuntimeError(f"Could not assign Sunday PN on {d} (rotation candidates unavailable).")
            self.assign(d, "PN", chosen)
            idx = (order.index(chosen) + 1) % n

    def _need_score(self, role: str, p: str) -> int:
        return self.targets[role][p] - self.primary_counts[role][p]

    def candidate_list(self, d: date, role: str) -> List[str]:
        cands = [p for p in self.inp.fsas if self.can_assign_primary(p, d, role)[0]]
        if not cands:
            return []
        hier_rank = {p: i for i, p in enumerate(self.hierarchy)}
        cands.sort(key=lambda p: (self._need_score(role, p), -hier_rank.get(p, 999)), reverse=True)
        band = max(2, min(4, len(cands)))
        top = cands[:band]
        self.rng.shuffle(top)
        return top + cands[band:]

    def candidate_count(self, d: date, role: str) -> int:
        return sum(1 for p in self.inp.fsas if self.can_assign_primary(p, d, role)[0])

    def next_unfilled_primary_slot(self) -> Optional[Tuple[date, str]]:
        best = None
        best_len = 10**9
        for d in self.days:
            req = required_primary_roles(d, self.inp)
            if not req:
                continue
            for role in PRIMARY_ROLES_ORDER:
                if role not in req:
                    continue
                if role in self.sched[d]:
                    continue
                n = self.candidate_count(d, role)
                if n == 0:
                    return (d, role)
                if n < best_len:
                    best_len = n
                    best = (d, role)
        return best

    def explain_no_candidates(self, d: date, role: str) -> Dict[str, Any]:
        reasons = {}
        for p in self.inp.fsas:
            ok, why = self.can_assign_primary(p, d, role)
            reasons[p] = "OK" if ok else why
        prev = d - timedelta(days=1)
        prev_roles = self.ctx.get_roles(self.sched, prev)
        return {
            "phase": "PRIMARY",
            "date": d,
            "role": role,
            "assigned_today": dict(self.sched[d]),
            "prev_day": prev,
            "prev_day_roles": dict(prev_roles),
            "reasons": reasons
        }

    def solve(self, max_nodes: int = 2_000_000, report_every: int = 25_000, spinner: bool = True) -> Dict[date, Dict[str, str]]:
        order = self.inp.sunday_rotation_order
        start_idx = order.index(self.inp.sunday_first_assignee)
        seed_trials = 3

        last_err = None
        for s_try in range(seed_trials):
            for shift in range(len(order)):
                # reset
                self.sched = {d: {} for d in self.days}
                self.week_hours = defaultdict(Counter)
                self._seed_week_hours_from_carryover()
                self.pp_hours = defaultdict(Counter)
                self._seed_pay_period_hours_from_carryover()
                self.mon_pn = Counter()
                self.primary_counts = {r: Counter() for r in PRIMARY_ROLES_ORDER}
                self.first_dead_end = None
                self.rng = random.Random(11 + s_try * 101 + shift * 17)

                first = order[(start_idx + shift) % len(order)]
                try:
                    self.set_sundays_by_rotation(first_assignee=first)
                except RuntimeError as e:
                    last_err = e
                    continue

                try:
                    return self._backtrack(max_nodes=max_nodes, report_every=report_every, spinner=spinner)
                except RuntimeError as e:
                    last_err = e
                    continue

        raise RuntimeError(f"Primary phase failed after fallback attempts. Last error: {last_err}") from last_err

    def _backtrack(self, max_nodes: int, report_every: int, spinner: bool) -> Dict[date, Dict[str, str]]:
        t0 = time.time()
        last_report_t = t0
        nodes = 0
        spin_chars = "|/-\\"
        spin_i = 0

        def progress_line(depth: int, slot: Optional[Tuple[date, str]]):
            nonlocal spin_i
            now = time.time()
            elapsed = now - t0
            rate = nodes / elapsed if elapsed > 0 else 0.0
            s = spin_chars[spin_i % len(spin_chars)] if spinner else ""
            spin_i += 1
            if slot:
                dd, rr = slot
                msg = f"{s} [PRIMARY] nodes={nodes:,} rate={rate:,.0f}/s depth={depth} slot={dd} {rr}"
            else:
                msg = f"{s} [PRIMARY] nodes={nodes:,} rate={rate:,.0f}/s depth={depth} slot=done"
            sys.stdout.write("\r" + msg[:140].ljust(140))
            sys.stdout.flush()

        def backtrack(depth: int) -> bool:
            nonlocal nodes, last_report_t
            nodes += 1
            if nodes > max_nodes:
                return False

            slot = self.next_unfilled_primary_slot()
            if slot is None:
                progress_line(depth, None)
                return True

            dd, rr = slot
            cands = self.candidate_list(dd, rr)
            if not cands:
                if self.first_dead_end is None:
                    self.first_dead_end = self.explain_no_candidates(dd, rr)
                return False

            now = time.time()
            if nodes % report_every == 0 or (now - last_report_t) > 2.0:
                progress_line(depth, slot)
                last_report_t = now

            for p in cands:
                ok, _ = self.can_assign_primary(p, dd, rr)
                if not ok:
                    continue
                self.assign(dd, rr, p)
                if backtrack(depth + 1):
                    return True
                self.unassign(dd, rr)
            return False

        ok = backtrack(0)
        sys.stdout.write("\n")
        sys.stdout.flush()

        if not ok:
            if self.first_dead_end:
                de = self.first_dead_end
                print("❌ PRIMARY phase failed. First dead-end:")
                print(f"  Date: {de['date']}  Role: {de['role']}")
                print(f"  Assigned today: {de['assigned_today']}")
                print(f"  Prev day ({de['prev_day']}): {de['prev_day_roles']}")
                print("  Block reasons:")
                for p, why in de["reasons"].items():
                    print(f"    - {p}: {why}")
            raise RuntimeError("Failed to assign all primary roles.")
        return self.sched


# -----------------------------
# Phase 2: BU backfill (scalable) with cross-month checks
# -----------------------------

class BUBackfill:
    def __init__(self, inp: Inputs, primary_sched: Dict[date, Dict[str, str]], seed: int = 11):
        self.inp = inp
        self.rng = random.Random(seed)
        self.days = month_days(inp.year, inp.month)
        self.push_days = compute_push_days(inp)
        self.ctx = Context(inp, self.days)

        self.sched: Dict[date, Dict[str, str]] = {d: dict(primary_sched.get(d, {})) for d in self.days}

        # Seed week hours from carryover assignments plus primaries already assigned
        self.week_hours: Dict[date, Counter] = defaultdict(Counter)
        # Seed pay period hours from carryover + primaries already assigned (if pay periods provided)
        self.pp_hours: Dict[str, Counter] = defaultdict(Counter)
        for d in self.days:
            wk = week_start_sun(d)
            if wk not in self.week_hours:
                self.week_hours[wk] = self.ctx.seed_week_hours(self.sched, wk)
            pp = self.ctx.pay_period_for(d)
            if pp is not None and pp.id not in self.pp_hours:
                self.pp_hours[pp.id] = self.ctx.seed_pay_period_hours(self.sched, pp)
            for role, p in self.sched[d].items():
                h = hours_for(d, role, inp)
                self.week_hours[wk][p] += h
                if pp is not None:
                    self.pp_hours[pp.id][p] += h

        self.shortages: List[Dict[str, Any]] = []

    def _avoid_status(self, person: str, d: date, role: str) -> str:
        avoid = self.inp.role_avoid.get(person, {}).get(d, set())
        if role not in avoid:
            return "none"
        hard = self.inp.role_avoid_hard.get(person, {}).get(d, False)
        return "hard" if hard else "soft"

    def can_assign_bu(self, p: str, d: date, bu_role: str) -> Tuple[bool, str]:
        inp = self.inp

        if d in inp.time_off.get(p, set()):
            return False, "time_off"
        if self._avoid_status(p, d, bu_role) == "hard":
            return False, "role_avoid_hard"
        if p in self.sched[d].values():
            return False, "already_today"

        # Cross-month no same specific BU role back-to-back
        if self.ctx.same_role_prev_day(self.sched, d, bu_role, p):
            return False, "same_bu_role_back_to_back"

        # Cross-month max consecutive days INCLUDING BU
        if self.ctx.consecutive_streak_ending(self.sched, d, p, assume_work_today=True) > inp.max_consecutive_days:
            return False, "max_consecutive_days"

        wk = week_start_sun(d)
        waive = week_contains_push(wk, self.push_days)
        add_h = hours_for(d, bu_role, inp)
        if (not waive) and (self.week_hours[wk][p] + add_h > inp.weekly_cap_hours):
            return False, "weekly_cap"

        pp = self.ctx.pay_period_for(d)
        if pp is not None:
            pp_waive = pay_period_contains_push(pp, self.push_days)
            if (not pp_waive) and (self.pp_hours[pp.id][p] + add_h > inp.pay_period_cap_hours):
                return False, "pay_period_cap"

        return True, "ok"

    def assign_bu(self, d: date, bu_role: str, p: str):
        self.sched[d][bu_role] = p
        wk = week_start_sun(d)
        h = hours_for(d, bu_role, self.inp)
        self.week_hours[wk][p] += h
        pp = self.ctx.pay_period_for(d)
        if pp is not None:
            self.pp_hours[pp.id][p] += h

    def candidates(self, d: date, bu_role: str) -> List[str]:
        cands = [p for p in self.inp.fsas if self.can_assign_bu(p, d, bu_role)[0]]
        if not cands:
            return []
        wk = week_start_sun(d)
        cands.sort(key=lambda x: (self.week_hours[wk][x], x))
        band = max(2, min(4, len(cands)))
        top = cands[:band]
        self.rng.shuffle(top)
        return top + cands[band:]

    def fill_day(self, d: date):
        desired = desired_bu_roles(d, self.inp, self.push_days)
        if not desired:
            return

        for bu_role in desired:
            if bu_role in self.sched[d]:
                continue
            cands = self.candidates(d, bu_role)
            if not cands:
                reasons = {p: ("OK" if self.can_assign_bu(p, d, bu_role)[0] else self.can_assign_bu(p, d, bu_role)[1])
                           for p in self.inp.fsas}
                self.shortages.append({
                    "date": d,
                    "missing_from": bu_role,
                    "missing_roles": [bu_role] + [r for r in desired if r not in self.sched[d] and r != bu_role],
                    "assigned_today": dict(self.sched[d]),
                    "prev_day": d - timedelta(days=1),
                    "prev_day_roles": dict(self.ctx.get_roles(self.sched, d - timedelta(days=1))),
                    "reasons_for_first_missing": reasons
                })
                break

            self.assign_bu(d, bu_role, cands[0])

    def fill_month(self, report_every_days: int = 2):
        t0 = time.time()
        for i, d in enumerate(self.days):
            self.fill_day(d)
            if (i % report_every_days) == 0:
                elapsed = time.time() - t0
                sys.stdout.write(f"\r[BU] day {i+1}/{len(self.days)} elapsed {elapsed:0.1f}s")
                sys.stdout.flush()
        sys.stdout.write("\n")
        sys.stdout.flush()
        return self.sched


# -----------------------------
# High-level API
# -----------------------------

def build_schedule(inp: Inputs, seed: int = 11) -> Tuple[Dict[date, Dict[str, str]], List[Dict[str, Any]]]:
    prim = PrimarySolver(inp, seed=seed)
    primary_sched = prim.solve(max_nodes=2_000_000, report_every=25_000, spinner=True)

    bu = BUBackfill(inp, primary_sched, seed=seed)
    full = bu.fill_month(report_every_days=2)

    return full, bu.shortages


def print_shortages(shortages: List[Dict[str, Any]]):
    if not shortages:
        print("\n✅ No BU shortages (all desired BU roles were staffed).")
        return
    print(f"\n⚠️  BU SHORTAGES: {len(shortages)} day(s) ran short on BU coverage.\n")
    for s in shortages:
        d = s["date"]
        miss = s["missing_roles"]
        print(f"- {d}: missing {miss}  assigned={s['assigned_today']}")
        reasons = s["reasons_for_first_missing"]
        why_counts = Counter(reasons.values())
        why_summary = ", ".join(f"{k}:{v}" for k, v in why_counts.items())
        print(f"  blockers summary: {why_summary}")


def print_schedule(sched: Dict[date, Dict[str, str]], inp: Inputs):
    days = month_days(inp.year, inp.month)
    push_days = compute_push_days(inp)
    for d in days:
        prim = required_primary_roles(d, inp)
        bus = desired_bu_roles(d, inp, push_days)
        req = prim + bus
        if not req:
            print(f"{d} (HOLIDAY)")
            continue
        a = sched.get(d, {})
        parts = []
        for r in ["PN", "AN", "W", "BU1", "BU2", "BU3", "BU4"]:
            if r in req:
                parts.append(f"{r}:{a.get(r,'-')}")
        print(f"{d}  " + "  ".join(parts))


# -----------------------------
# Example run (Feb 2026)
# -----------------------------

# -----------------------------
# File I/O helpers (JSON + minimal YAML)
# -----------------------------

def _parse_date(s: str) -> date:
    y, m, d = s.strip().split("-")
    return date(int(y), int(m), int(d))

def _load_text(path: str) -> str:
    return Path(path).expanduser().read_text(encoding="utf-8")

def _looks_like_json(txt: str) -> bool:
    t = txt.lstrip()
    return t.startswith("{") or t.startswith("[")

def load_json_or_yaml(path: str) -> Any:
    """
    Loads either JSON or a minimal YAML subset (enough for our config files).
    Supported YAML features:
      - key: value (strings, numbers)
      - lists using "- item"
      - inline lists: [a, b, c]
      - nested dicts via indentation (2 spaces recommended)
    If you want full YAML, you can install PyYAML and swap this loader later.
    """
    txt = _load_text(path)
    if _looks_like_json(txt):
        return json.loads(txt)

    # Minimal YAML parser (subset)
    lines = [ln.rstrip("\n") for ln in txt.splitlines()]
    # strip comments and empty lines
    cleaned = []
    for ln in lines:
        ln2 = ln.split("#", 1)[0].rstrip()
        if ln2.strip():
            cleaned.append(ln2)

    def parse_inline_list(val: str) -> List[str]:
        inner = val.strip()[1:-1].strip()
        if not inner:
            return []
        parts = [p.strip().strip('"').strip("'") for p in inner.split(",")]
        return [p for p in parts if p]

    # Stack of (indent, container, last_key)
    root: Any = None
    stack: List[Tuple[int, Any, Optional[str]]] = []

    def current():
        return stack[-1][1] if stack else None

    def set_root(obj):
        nonlocal root
        root = obj

    def push(indent: int, obj: Any, last_key: Optional[str] = None):
        stack.append((indent, obj, last_key))

    def pop_to(indent: int):
        while stack and stack[-1][0] >= indent:
            stack.pop()

    for ln in cleaned:
        indent = len(ln) - len(ln.lstrip(" "))
        s = ln.lstrip(" ")

        # list item
        if s.startswith("- "):
            item = s[2:].strip()
            cont = current()
            if cont is None:
                set_root([])
                push(-1, root, None)
                cont = current()
            if not isinstance(cont, list):
                # create list under last dict key
                parent_indent, parent, last_key = stack[-1]
                if isinstance(parent, dict) and last_key:
                    new_list = []
                    parent[last_key] = new_list
                    pop_to(parent_indent + 1)
                    push(parent_indent + 1, new_list, None)
                    cont = current()
                else:
                    raise ValueError("YAML structure error near: " + ln)

            # item could be "key: value" dict inline
            if ":" in item and not item.strip().startswith('"') and not item.strip().startswith("'"):
                k, v = item.split(":", 1)
                k = k.strip()
                v = v.strip()
                obj = {}
                cont.append(obj)
                # prepare to possibly nest more keys into this dict
                pop_to(indent + 1)
                push(indent + 1, obj, k)
                if v:
                    if v.startswith("[") and v.endswith("]"):
                        obj[k] = parse_inline_list(v)
                    else:
                        obj[k] = v.strip('"').strip("'")
                else:
                    obj[k] = None
            else:
                # scalar
                val = item.strip().strip('"').strip("'")
                cont.append(val)
            continue

        # key: value
        if ":" in s:
            k, v = s.split(":", 1)
            k = k.strip()
            v = v.strip()

            # decide container
            if not stack:
                set_root({})
                push(-1, root, None)

            pop_to(indent + 1)
            cont = current()
            if cont is None or not isinstance(cont, dict):
                raise ValueError("YAML structure error near: " + ln)

            # set last_key for potential list attachment
            stack[-1] = (stack[-1][0], cont, k)

            if v == "":
                cont[k] = None
                continue

            # inline list
            if v.startswith("[") and v.endswith("]"):
                cont[k] = parse_inline_list(v)
            else:
                # try number
                vv = v.strip('"').strip("'")
                if re.fullmatch(r"-?\d+(\.\d+)?", vv):
                    cont[k] = float(vv) if "." in vv else int(vv)
                else:
                    cont[k] = vv
            continue

        raise ValueError("Unsupported YAML line: " + ln)

    return root

def load_time_off(path: str) -> Dict[str, Set[date]]:
    """
    Accepts JSON/YAML structures:
    time_off:
      Will:
        - 2026-02-07..2026-02-12
        - 2026-02-20
      CJ: [2026-02-05, 2026-02-06]
    or JSON:
    {"Will":["2026-02-07..2026-02-12","2026-02-20"], "CJ":["2026-02-05","2026-02-06"]}
    """
    data = load_json_or_yaml(path)
    # allow top-level {"time_off": {...}} or direct dict
    if isinstance(data, dict) and "time_off" in data and isinstance(data["time_off"], dict):
        data = data["time_off"]

    out: Dict[str, Set[date]] = defaultdict(set)
    if not isinstance(data, dict):
        raise ValueError("time_off file must be a dict keyed by FSA name.")
    for name, items in data.items():
        if items is None:
            continue
        if isinstance(items, str):
            items = [items]
        if not isinstance(items, list):
            raise ValueError(f"time_off[{name}] must be a list.")
        for it in items:
            it = str(it).strip()
            if ".." in it:
                a, b = it.split("..", 1)
                d1 = _parse_date(a)
                d2 = _parse_date(b)
                for d in daterange(d1, d2):
                    out[name].add(d)
            else:
                out[name].add(_parse_date(it))
    return out


def load_requests_xlsx(path: str, sheet: str = "TimeOff") -> Tuple[Dict[str, Set[date]], Dict[str, Dict[date, Set[str]]], Dict[str, Dict[date, bool]]]:
    """
    Excel requests sheet columns (row 1):
      Name | Date | Start | End | Hard | AvoidRoles

    - Full day off: AvoidRoles blank; Date or Start/End filled (treated as HARD).
    - Role avoidance: AvoidRoles filled (e.g. "PN, AN"). Hard controls hard-vs-soft:
        blank/Y/YES/TRUE => hard, N/NO/FALSE => soft
    """
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl is required to read Excel files. Install with: pip install openpyxl") from e

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found. Found: {wb.sheetnames}")
    ws = wb[sheet]

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        headers[str(v).strip().lower()] = c

    def col(name: str) -> int:
        return headers.get(name.lower(), -1)

    name_c = col("name")
    date_c = col("date")
    start_c = col("start")
    end_c = col("end")
    hard_c = col("hard")
    avoid_c = col("avoidroles")

    if name_c == -1:
        raise ValueError("Requests sheet must include a 'Name' column in row 1.")

    time_off: Dict[str, Set[date]] = defaultdict(set)
    role_avoid: Dict[str, Dict[date, Set[str]]] = defaultdict(dict)
    role_avoid_hard: Dict[str, Dict[date, bool]] = defaultdict(dict)

    def to_date(val) -> Optional[date]:
        if val is None or val == "":
            return None
        if isinstance(val, date):
            return val
        try:
            from datetime import datetime
            if isinstance(val, datetime):
                return val.date()
        except Exception:
            pass
        s = str(val).strip()
        if "/" in s and len(s.split("/")) == 3:
            mm, dd, yy = s.split("/")
            return date(int(yy), int(mm), int(dd))
        return _parse_date(s)

    def hard_bool(val) -> bool:
        if val is None or str(val).strip() == "":
            return True
        s = str(val).strip().lower()
        if s.startswith("n") or s.startswith("f") or s == "0":
            return False
        return True

    def parse_roles(val) -> Set[str]:
        if val is None or str(val).strip() == "":
            return set()
        import re as _re
        s = str(val).upper()
        parts = _re.split(r"[,\;\s]+", s)
        parts = [p.strip() for p in parts if p.strip()]
        norm = set()
        for p in parts:
            if p in {"PN","AN","W","BU1","BU2","BU3","BU4"}:
                norm.add(p)
            elif p == "BU":
                norm.update({"BU1","BU2","BU3","BU4"})
        return norm

    for r in range(2, ws.max_row + 1):
        nm = ws.cell(row=r, column=name_c).value
        if nm is None or str(nm).strip() == "":
            continue
        name = str(nm).strip()

        d_single = to_date(ws.cell(row=r, column=date_c).value) if date_c != -1 else None
        d_start  = to_date(ws.cell(row=r, column=start_c).value) if start_c != -1 else None
        d_end    = to_date(ws.cell(row=r, column=end_c).value) if end_c != -1 else None

        hard = hard_bool(ws.cell(row=r, column=hard_c).value) if hard_c != -1 else True
        avoid = parse_roles(ws.cell(row=r, column=avoid_c).value) if avoid_c != -1 else set()

        if d_single:
            dates = [d_single]
        elif d_start and d_end:
            dates = list(daterange(d_start, d_end))
        elif d_start:
            dates = [d_start]
        else:
            continue

        if avoid:
            for dd in dates:
                role_avoid[name][dd] = set(role_avoid[name].get(dd, set())) | set(avoid)
                role_avoid_hard[name][dd] = bool(role_avoid_hard[name].get(dd, False) or hard)
        else:
            for dd in dates:
                time_off[name].add(dd)

    return time_off, role_avoid, role_avoid_hard

def load_time_off_xlsx(path: str, sheet: str = "TimeOff") -> Dict[str, Set[date]]:
    t, _, _ = load_requests_xlsx(path, sheet=sheet)
    return t


def load_sales_ranking(path: str, month_first_day: date) -> List[str]:
    """
    sales ranking file supports JSON or minimal YAML.

    Recommended YAML:
      periods:
        - start: 2026-01-01
          end: 2026-03-31
          ranking: [Mark, Will, Dawn, CJ, Kyle, Greg]
        - start: 2026-04-01
          end: 2026-06-30
          ranking:
            - Mark
            - Dawn
            - Will
            - Kyle
            - CJ
            - Greg

    If file contains only:
      ranking: [...]
    we'll use it directly.
    """
    data = load_json_or_yaml(path)

    if isinstance(data, dict) and "ranking" in data and isinstance(data["ranking"], list):
        return [str(x) for x in data["ranking"]]

    periods = None
    if isinstance(data, dict) and "periods" in data and isinstance(data["periods"], list):
        periods = data["periods"]
    elif isinstance(data, list):
        # treat as list of periods
        periods = data

    if not periods:
        raise ValueError("sales file must contain 'ranking' or 'periods'.")

    for p in periods:
        if not isinstance(p, dict):
            continue
        start = _parse_date(str(p.get("start")))
        end = _parse_date(str(p.get("end")))
        if start <= month_first_day <= end:
            ranking = p.get("ranking")
            if not isinstance(ranking, list):
                raise ValueError("period.ranking must be a list.")
            return [str(x) for x in ranking]

    raise ValueError(f"No sales ranking period covers {month_first_day}.")

def ranking_to_sales_volume(ranking_high_to_low: List[str]) -> Dict[str, float]:
    """
    Convert a ranking list into synthetic sales volumes so existing solver logic can sort.
    Higher value = higher rank.
    """
    n = len(ranking_high_to_low)
    return {name: float(n - i) for i, name in enumerate(ranking_high_to_low)}

def load_carryover(path: str) -> Dict[date, Dict[str, str]]:
    """
    JSON/YAML mapping date -> role dict.
    Example JSON:
      {"2026-01-31":{"PN":"Kyle","AN":"Mark","W":"CJ","BU1":"Greg"}}
    """
    data = load_json_or_yaml(path)
    if not isinstance(data, dict):
        raise ValueError("carryover file must be a dict keyed by date strings.")
    out: Dict[date, Dict[str, str]] = {}
    for ds, roles in data.items():
        dd = _parse_date(str(ds))
        if not isinstance(roles, dict):
            continue
        out[dd] = {str(r): str(p) for r, p in roles.items()}
    return out


def load_schedule_json(path: str) -> Dict[date, Dict[str, str]]:
    """
    Reads schedule.json written by this tool:
      {"YYYY-MM-DD":{"PN":"Name","AN":"Name",...}, ...}
    """
    data = json.loads(_load_text(path))
    if not isinstance(data, dict):
        raise ValueError("schedule.json must be a dict keyed by YYYY-MM-DD.")
    out: Dict[date, Dict[str, str]] = {}
    for ds, roles in data.items():
        dd = _parse_date(str(ds))
        if not isinstance(roles, dict):
            continue
        out[dd] = {str(r): str(p) for r, p in roles.items()}
    return out

def make_carryover_from_prev_schedule(prev_schedule: Dict[date, Dict[str, str]], month_first_day: date, days: int) -> Dict[date, Dict[str, str]]:
    """
    Builds carryover_assignments by taking the last N days immediately before month_first_day
    from a previously generated schedule map.
    """
    start = month_first_day - timedelta(days=days)
    end = month_first_day - timedelta(days=1)
    out: Dict[date, Dict[str, str]] = {}
    for d in daterange(start, end):
        if d in prev_schedule:
            out[d] = dict(prev_schedule[d])
    return out
def save_schedule_json(path: str, sched: Dict[date, Dict[str, str]]):
    data = {d.isoformat(): roles for d, roles in sorted(sched.items(), key=lambda x: x[0])}
    Path(path).write_text(json.dumps(data, indent=2), encoding="utf-8")

def save_shortages_json(path: str, shortages: List[Dict[str, Any]]):
    def conv(obj):
        if isinstance(obj, date):
            return obj.isoformat()
        return obj
    data = []
    for s in shortages:
        ss = {k: conv(v) for k, v in s.items()}
        # nested dicts may include dates already as date objects
        if "assigned_today" in ss and isinstance(ss["assigned_today"], dict):
            ss["assigned_today"] = {k: v for k, v in ss["assigned_today"].items()}
        data.append(ss)
    Path(path).write_text(json.dumps(data, indent=2), encoding="utf-8")
def compute_pay_period_hours(sched: Dict[date, Dict[str, str]], inp: Inputs) -> Dict[str, Dict[str, int]]:
    """
    Returns {PPID: {person: hours}} for the pay periods that overlap this month.
    Includes only hours that fall within each PP window, regardless of month.
    Note: If you want complete PP totals, provide carryover days that include the full PP.
    """
    if not inp.pay_periods:
        return {}
    # Build a unified assignment map from carryover + in-month schedule for the lookback window
    # We'll count only within PP windows.
    # Gather assignments: carryover plus month days
    assign_map: Dict[date, Dict[str, str]] = {}
    for d, roles in (inp.carryover_assignments or {}).items():
        assign_map[d] = dict(roles)
    for d, roles in sched.items():
        assign_map[d] = dict(roles)

    out: Dict[str, Dict[str, int]] = {}
    month_start = date(inp.year, inp.month, 1)
    month_end = month_days(inp.year, inp.month)[-1]
    for pp in inp.pay_periods:
        # only report pay periods that overlap this month
        if pp.end < month_start or pp.start > month_end:
            continue
        hours = Counter()
        dd = pp.start
        while dd <= pp.end:
            roles = assign_map.get(dd, {})
            for role, person in roles.items():
                hours[person] += hours_for(dd, role, inp)
            dd += timedelta(days=1)
        out[pp.id] = dict(hours)
    return out

def save_pay_period_hours_json(path: str, pp_hours: Dict[str, Dict[str, int]]):
    Path(path).write_text(json.dumps(pp_hours, indent=2), encoding="utf-8")


# -----------------------------
# CLI runner
# -----------------------------
if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser(description="Harpeth Hills FSA scheduler (logic only).")
    ap.add_argument("--month", required=True, help="Target month in YYYY-MM (e.g., 2026-02)")
    ap.add_argument("--fsas", default="Mark,Dawn,Will,Kyle,CJ,Greg", help="Comma-separated FSA short names.")
    ap.add_argument("--seniority", default="Mark,Dawn,Will,Kyle,CJ,Greg", help="Comma-separated seniority order.")
    ap.add_argument("--sales", default="", help="Path to sales ranking file (JSON/YAML). If omitted, uses ./sales_ranking.yaml if present; otherwise falls back to seniority order.")
    ap.add_argument("--timeoff", default="", help="Path to time off file (JSON/YAML). If omitted, uses ./time_off.yaml if present; otherwise assumes none.")
    ap.add_argument("--timeoff-xlsx", default="", help="Optional Excel file to read time off requests from (sheet TimeOff). Overrides --timeoff if provided.")
    ap.add_argument("--timeoff-sheet", default="", help="Sheet name in the Excel file. Default: TimeOff, else <MonthName> Requests if present.")
    ap.add_argument("--pay-periods", default="", help="Optional path to pay periods file (JSON/YAML) to enforce 80-hr cap per pay period (waived if pay period contains push days). If omitted, uses ./pay_periods.yaml if present.")
    ap.add_argument("--carryover", default="", help="Optional path to carryover assignments (JSON/YAML).")
    ap.add_argument("--prev-schedule", default="", help="Optional path to previous month schedule.json to auto-build carryover.")
    ap.add_argument("--carryover-days", type=int, default=14, help="How many days before month start to include in auto-carryover.")
    ap.add_argument("--out-next-carryover", default="next_carryover.json", help="Output carryover.json for the next month.")
    ap.add_argument("--sunday-rotation", default="Greg,Mark,Dawn,Will,CJ,Kyle", help="Comma-separated Sunday rotation order.")
    ap.add_argument("--sunday-first", default="Greg", help="First assignee for Sunday PN rotation (carryover-based typically).")
    ap.add_argument("--seed", type=int, default=11)
    ap.add_argument("--lookback-days", type=int, default=14)
    ap.add_argument("--out-schedule", default="schedule.json")
    ap.add_argument("--out-shortages", default="shortages.json")
    ap.add_argument("--out-payperiod-hours", default="pay_period_hours.json", help="Output pay-period hours summary (JSON).")
    args = ap.parse_args()

    y, m = args.month.split("-")
    year = int(y); month = int(m)
    first_day = date(year, month, 1)

    fsas = [x.strip() for x in args.fsas.split(",") if x.strip()]
    seniority = [x.strip() for x in args.seniority.split(",") if x.strip()]
    sunday_rotation = [x.strip() for x in args.sunday_rotation.split(",") if x.strip()]


    # Time off loading precedence:
    #   1) --timeoff-xlsx (shared Excel, OneDrive synced locally)
    #   2) --timeoff (yaml/json)
    #   3) ./time_off.yaml if present
    #   4) none
    role_avoid = defaultdict(dict)
    role_avoid_hard = defaultdict(dict)

    if args.timeoff_xlsx.strip():
        sheet = args.timeoff_sheet.strip()
        if not sheet:
            import calendar as _cal
            candidate = f"{_cal.month_name[month]} Requests"
            try:
                import openpyxl
                _wb = openpyxl.load_workbook(args.timeoff_xlsx.strip(), read_only=True, data_only=True)
                if "TimeOff" in _wb.sheetnames:
                    sheet = "TimeOff"
                elif candidate in _wb.sheetnames:
                    sheet = candidate
                else:
                    sheet = _wb.sheetnames[0]
                _wb.close()
            except Exception:
                sheet = "TimeOff"
        time_off, role_avoid, role_avoid_hard = load_requests_xlsx(args.timeoff_xlsx.strip(), sheet=sheet)
        print(f"ℹ️  Loaded requests from Excel: {args.timeoff_xlsx.strip()} (sheet='{sheet}')")
    else:
        timeoff_path = args.timeoff.strip()
        if not timeoff_path and Path("time_off.yaml").exists():
            timeoff_path = "time_off.yaml"
        time_off = load_time_off(timeoff_path) if timeoff_path else {p: set() for p in fsas}

    # Sales ranking loading (updates quarterly):
    #   1) --sales
    #   2) ./sales_ranking.yaml if present
    #   3) fallback to seniority order (least preferred)
    sales_path = args.sales.strip()
    if not sales_path and Path("sales_ranking.yaml").exists():
        sales_path = "sales_ranking.yaml"
    if sales_path:
        ranking = load_sales_ranking(sales_path, first_day)
    else:
        ranking = seniority[:]  # fallback
        print("⚠️  No sales ranking file provided/found. Falling back to seniority order for hierarchy decisions.")
    sales_volume = ranking_to_sales_volume(ranking)


    pay_periods = []
    pp_path = args.pay_periods.strip()
    if not pp_path and Path("pay_periods.yaml").exists():
        pp_path = "pay_periods.yaml"
    if pp_path:
        pay_periods = load_pay_periods(pp_path)

    carry = {}
    if args.carryover.strip():
        carry = load_carryover(args.carryover)
    elif args.prev_schedule.strip():
        prev = load_schedule_json(args.prev_schedule)
        carry = make_carryover_from_prev_schedule(prev, first_day, args.carryover_days)
        if not carry:
            print("⚠️  prev-schedule provided, but no matching days were found for the carryover window.")


    inp = Inputs(
        year=year,
        month=month,
        fsas=fsas,
        sales_volume=sales_volume,
        seniority_order=seniority,
        time_off=time_off,
        role_avoid=role_avoid,
        role_avoid_hard=role_avoid_hard,
        pay_periods=pay_periods,
        sunday_rotation_order=sunday_rotation,
        sunday_first_assignee=args.sunday_first,
        observed_holidays=set(),
        push_week_enabled=True,
        carryover_assignments=carry,
        lookback_days=max(args.lookback_days, args.carryover_days),
    )

    sched, shortages = build_schedule(inp, seed=args.seed)
    save_schedule_json(args.out_schedule, sched)
    save_shortages_json(args.out_shortages, shortages)

    print(f"\n✅ Wrote schedule -> {args.out_schedule}")

    if inp.pay_periods:
        pp_hours = compute_pay_period_hours(sched, inp)
        save_pay_period_hours_json(args.out_payperiod_hours, pp_hours)
        print(f"✅ Wrote pay period hours -> {args.out_payperiod_hours}")
    print(f"✅ Wrote BU shortages -> {args.out_shortages}")

    # Write carryover file for the next month automatically (so month-to-month runs are hands-off)
    next_month_first = (first_day.replace(day=28) + timedelta(days=4)).replace(day=1)  # first day of next month
    next_carry = make_carryover_from_prev_schedule(load_schedule_json(args.out_schedule), next_month_first, args.carryover_days)
    Path(args.out_next_carryover).write_text(json.dumps({d.isoformat(): roles for d, roles in sorted(next_carry.items())}, indent=2), encoding="utf-8")
    print(f"✅ Wrote next-month carryover -> {args.out_next_carryover}")
    print_shortages(shortages)