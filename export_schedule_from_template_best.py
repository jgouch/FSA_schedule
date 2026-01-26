#!/usr/bin/env python3
"""
export_schedule_from_template_best.py

Best-combined exporter for your filled calendar-template Excel -> schedule.json
(for use as --prev-schedule in schedule_logic).

Combines:
- v1 simplicity (minimal assumptions, easy CLI)
- v2 robustness (header parsing, day-number detection, safer scanning)

Key features:
- Month header parsing supports: "JANUARY 2026", "Jan 2026", "Sept. 2026", "February-2026", etc.
- Day cell detection supports int, float (1.0), or string ("1")
- Scans each day-square downward until it hits the next day-number cell in the same date column
  (or a safe max window)
- Primary roles detected by labels starting with PN / AN / W (case-insensitive)
- BU roles are unlabeled: extra names after primaries become BU1..BU4 top-to-bottom
- Optional: include empty days (store {} for days with no roles, e.g., holidays)

Usage:
  python3 export_schedule_from_template_best.py \
    --in "/path/January_2026_Filled_Template copy.xlsx" \
    --out "/path/jan_2026_schedule.json"

Optional:
  --sheet "Sheet1"
  --include-empty-days
  --date-col-offset 0    (rare; default 0 means the day number is in the date column itself)
  --label-col-offset 1   (default: labels are 1 col to the right of the date)
  --name-col-offset 2    (default: names are 2 cols to the right of the date)
"""

import argparse
import json
import re
import calendar
from datetime import date
from pathlib import Path
from typing import Dict, Tuple, List, Optional

import openpyxl

MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

PRIMARY_LABELS = ("PN", "AN", "W")
BU_ROLES = ("BU1", "BU2", "BU3", "BU4")


def parse_month_year(header_val) -> Tuple[int, int]:
    if header_val is None:
        raise ValueError("Month/year header cell is blank (expected something like 'January 2026').")

    s = str(header_val).strip()
    # Common patterns with optional punctuation and separators
    m = re.search(r"([A-Za-z]+)\.?\s*[-/ ]\s*(\d{4})", s)
    if not m:
        m = re.search(r"([A-Za-z]+)\.?\s+(\d{4})", s)
    if not m:
        raise ValueError(f"Could not parse month/year from header '{s}'. Expected like 'January 2026'.")

    mon_token = m.group(1).strip().lower()
    year = int(m.group(2))

    if mon_token not in MONTHS:
        mon3 = mon_token[:3]
        if mon3 in MONTHS:
            month = MONTHS[mon3]
        else:
            raise ValueError(
                f"Unrecognized month token '{mon_token}' in header '{s}'. "
                f"Use a standard month name like 'January 2026'."
            )
    else:
        month = MONTHS[mon_token]

    return year, month


def as_day_number(v) -> Optional[int]:
    """Return 1..31 if v looks like a day number, else None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v if 1 <= v <= 31 else None
    if isinstance(v, float):
        if abs(v - round(v)) < 1e-9:
            iv = int(round(v))
            return iv if 1 <= iv <= 31 else None
        return None
    s = str(v).strip()
    if s.isdigit():
        iv = int(s)
        return iv if 1 <= iv <= 31 else None
    return None


def role_from_label(label_s: str) -> Optional[str]:
    ls = label_s.strip().upper()
    if ls.startswith("PN"):
        return "PN"
    if ls.startswith("AN"):
        return "AN"
    if ls.startswith("W"):
        return "W"
    return None


def export_template(
    path_in: str,
    sheet_name: str = "",
    include_empty_days: bool = False,
    date_col_offset: int = 0,
    label_col_offset: int = 1,
    name_col_offset: int = 2,
) -> Tuple[int, int, Dict[str, Dict[str, str]]]:

    wb = openpyxl.load_workbook(path_in, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    # Month/year label is typically in B1, fallback to A1.
    header = ws.cell(1, 2).value
    if header is None or str(header).strip() == "":
        header = ws.cell(1, 1).value

    year, month = parse_month_year(header)

    # Find all day-number anchor cells (1..31)
    date_cells: List[Tuple[int, int, int]] = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            dn = as_day_number(cell.value)
            if dn is not None:
                date_cells.append((cell.row, cell.column, dn))

    if not date_cells:
        raise ValueError(
            "No day-number cells (1..31) were found. "
            "Check that the calendar template contains day numbers as values."
        )

    schedule: Dict[str, Dict[str, str]] = {}

    # We'll build from each anchor. If duplicates exist (unlikely), last one wins.
    for r, c, daynum in date_cells:
        # Respect offsets (mostly date_col_offset=0)
        date_col = c + date_col_offset
        label_col = c + label_col_offset
        name_col = c + name_col_offset

        try:
            d_iso = date(year, month, daynum).isoformat()
        except ValueError:
            # e.g. "31" appearing in a Feb template area; ignore
            continue

        # Scan downward until we hit another day number in the same date_col, or max window
        max_scan = 24
        stop_row = min(ws.max_row, r + max_scan)
        for rr in range(r + 1, r + max_scan + 1):
            if rr > ws.max_row:
                break
            dn2 = as_day_number(ws.cell(rr, date_col).value)
            if dn2 is not None:
                stop_row = rr - 1
                break

        roles: Dict[str, str] = {}

        # Primary pass
        last_primary_row = r
        for rr in range(r + 1, stop_row + 1):
            label = ws.cell(rr, label_col).value
            if label is None:
                continue
            label_s = str(label).strip()
            if not label_s:
                continue

            role = role_from_label(label_s)
            if role is None:
                continue

            name = ws.cell(rr, name_col).value
            if name is not None and str(name).strip() != "":
                roles[role] = str(name).strip()
            last_primary_row = rr

        # BU pass: unlabeled names after primaries
        bu_idx = 1
        blank_streak = 0
        for rr in range(last_primary_row + 1, stop_row + 1):
            label = ws.cell(rr, label_col).value
            label_s = str(label).strip().upper() if label is not None else ""

            # If this row is actually a primary label row, skip it
            if label_s.startswith(PRIMARY_LABELS):
                continue

            name = ws.cell(rr, name_col).value
            name_s = str(name).strip() if name is not None else ""

            if not name_s:
                blank_streak += 1
                if blank_streak >= 2:
                    break
                continue

            blank_streak = 0

            # guard against capturing stray role-label text
            if name_s.strip().upper() in set(PRIMARY_LABELS) | set(BU_ROLES) | {"BU"}:
                continue

            roles[f"BU{bu_idx}"] = name_s
            bu_idx += 1
            if bu_idx > 4:
                break

        if roles or include_empty_days:
            schedule[d_iso] = roles

    if include_empty_days:
        _, last_day = calendar.monthrange(year, month)
        for dd in range(1, last_day + 1):
            iso = date(year, month, dd).isoformat()
            schedule.setdefault(iso, {})

    return year, month, schedule


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="Input filled template .xlsx")
    ap.add_argument("--out", dest="out", default="", help="Output schedule.json path (default: YYYY-MM_schedule.json)")
    ap.add_argument("--sheet", dest="sheet", default="", help="Optional sheet name (default: first sheet)")
    ap.add_argument("--include-empty-days", action="store_true", help="Include all days of month with {} when no roles found.")
    ap.add_argument("--date-col-offset", type=int, default=0, help="Column offset from day-number cell to the date column (default 0).")
    ap.add_argument("--label-col-offset", type=int, default=1, help="Column offset from day-number cell to label column (default 1).")
    ap.add_argument("--name-col-offset", type=int, default=2, help="Column offset from day-number cell to name column (default 2).")
    args = ap.parse_args()

    year, month, sched = export_template(
        args.inp,
        sheet_name=args.sheet.strip(),
        include_empty_days=args.include_empty_days,
        date_col_offset=args.date_col_offset,
        label_col_offset=args.label_col_offset,
        name_col_offset=args.name_col_offset,
    )

    out_path = args.out.strip()
    if not out_path:
        out_path = f"{year:04d}-{month:02d}_schedule.json"

    Path(out_path).write_text(json.dumps(sched, indent=2), encoding="utf-8")
    print(f"✅ Exported {len(sched)} day entries for {year:04d}-{month:02d} → {out_path}")


if __name__ == "__main__":
    main()
