#!/usr/bin/env python3
"""
export_schedule_from_template_best_v3.py

More robust exporter for your filled calendar-template Excel -> schedule.json
(for use as --prev-schedule in schedule_logic).

Fixes the main failure modes:
1) Header parsing supports strings, Excel dates/datetimes, and common "03/2026" formats.
2) Day cell detection is validated to reduce false positives:
   - a candidate day cell must have a plausible "day-square" structure nearby.
3) Duplicate day anchors are detected; we keep the first and warn.
4) Invalid day-of-month is skipped with a warning.
5) Scans each day-square downward until the next day-number cell in the same date column
   (or a configurable max window).
6) If a primary label exists but name is blank, we still advance the primary boundary so BU
   scanning doesn't start too early.

Template assumptions:
- date number in cell (r,c)
- role labels in column (c+1)
- names in column (c+2)
- BU roles are unlabeled names after primaries (BU1..BU4)

Usage:
  python3 export_schedule_from_template_best_v3.py --in IN.xlsx --out jan_2026_schedule.json
"""

import argparse
import json
import re
import calendar
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Tuple, List, Optional

import openpyxl

MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

PRIMARY_LABELS = ("PN", "AN", "W")
BU_ROLES = ("BU1", "BU2", "BU3", "BU4")


def warn(msg: str, strict: bool = False):
    if strict:
        raise ValueError(msg)
    print(f"⚠️  {msg}", file=sys.stderr)


def parse_month_year(header_val) -> Tuple[int, int]:
    if header_val is None or str(header_val).strip() == "":
        raise ValueError("Month/year header cell is blank (expected like 'January 2026' or '03/2026').")

    if isinstance(header_val, (date, datetime)):
        dt = header_val.date() if isinstance(header_val, datetime) else header_val
        return dt.year, dt.month

    s = str(header_val).strip()

    m = re.search(r"^\s*(\d{1,2})\s*/\s*(\d{4})\s*$", s)
    if m:
        mm = int(m.group(1))
        yy = int(m.group(2))
        if 1 <= mm <= 12:
            return yy, mm
        raise ValueError(f"Header looks like MM/YYYY but month invalid: '{s}'")

    m = re.search(r"([A-Za-z]+)\.?\s*[-/ ]\s*(\d{4})", s) or re.search(r"([A-Za-z]+)\.?\s+(\d{4})", s)
    if not m:
        raise ValueError(f"Could not parse month/year from header '{s}'. Expected like 'January 2026' or '03/2026'.")

    mon_token = m.group(1).strip().lower()
    year = int(m.group(2))

    if mon_token not in MONTHS:
        mon3 = mon_token[:3]
        if mon3 in MONTHS:
            month = MONTHS[mon3]
        else:
            raise ValueError(f"Unrecognized month token '{mon_token}' in header '{s}'.")
    else:
        month = MONTHS[mon_token]

    return year, month


def as_day_number(v) -> Optional[int]:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v if 1 <= v <= 31 else None
    if isinstance(v, float):
        if abs(v - round(v)) < 1e-9:
            iv = int(round(v))
            return iv if 1 <= iv <= 31 else None
        return None
    s = str(v).strip()
    if s.isdigit():
        iv = int(s)
        return iv if 1 <= iv <= 31 else None
    return None


def role_from_label(label_s: str) -> Optional[str]:
    ls = label_s.strip().upper()
    if ls.startswith("PN"):
        return "PN"
    if ls.startswith("AN"):
        return "AN"
    if ls.startswith("W"):
        return "W"
    return None


def looks_like_day_square(ws, r: int, date_col: int, label_col: int, name_col: int, max_probe: int = 14) -> bool:
    found_label = False
    name_hits = 0

    for rr in range(r + 1, min(ws.max_row, r + max_probe) + 1):
        lab = ws.cell(rr, label_col).value
        if lab is not None and str(lab).strip():
            if str(lab).strip().upper().startswith(PRIMARY_LABELS):
                found_label = True
                break
    if found_label:
        return True

    for rr in range(r + 1, min(ws.max_row, r + max_probe) + 1):
        nm = ws.cell(rr, name_col).value
        if nm is not None and str(nm).strip():
            name_hits += 1
            if name_hits >= 2:
                return True
    return False


def export_template(
    path_in: str,
    sheet_name: str = "",
    include_empty_days: bool = False,
    date_col_offset: int = 0,
    label_col_offset: int = 1,
    name_col_offset: int = 2,
    max_scan: int = 40,
    strict: bool = False,
) -> Tuple[int, int, Dict[str, Dict[str, str]]]:

    wb = openpyxl.load_workbook(path_in, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header = ws.cell(1, 2).value
    if header is None or str(header).strip() == "":
        header = ws.cell(1, 1).value

    year, month = parse_month_year(header)

    date_cells: List[Tuple[int, int, int]] = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            dn = as_day_number(cell.value)
            if dn is None:
                continue
            r, c = cell.row, cell.column
            date_col = c + date_col_offset
            label_col = c + label_col_offset
            name_col = c + name_col_offset
            if not looks_like_day_square(ws, r, date_col, label_col, name_col):
                continue
            date_cells.append((r, c, dn))

    if not date_cells:
        raise ValueError("No day-number cells found that matched the day-square structure.")

    schedule: Dict[str, Dict[str, str]] = {}
    seen_iso = set()

    for r, c, daynum in date_cells:
        date_col = c + date_col_offset
        label_col = c + label_col_offset
        name_col = c + name_col_offset

        try:
            d_iso = date(year, month, daynum).isoformat()
        except ValueError:
            warn(f"Ignoring invalid day number {daynum} for {year:04d}-{month:02d} at cell ({r},{c}).", strict=strict)
            continue

        if d_iso in seen_iso:
            warn(f"Duplicate anchor for {d_iso} at cell ({r},{c}). Keeping first occurrence.", strict=strict)
            continue
        seen_iso.add(d_iso)

        stop_row = min(ws.max_row, r + max_scan)
        for rr in range(r + 1, min(ws.max_row, r + max_scan) + 1):
            dn2 = as_day_number(ws.cell(rr, date_col).value)
            if dn2 is not None:
                stop_row = rr - 1
                break

        roles: Dict[str, str] = {}

        last_primary_row = r
        for rr in range(r + 1, stop_row + 1):
            label = ws.cell(rr, label_col).value
            if label is None:
                continue
            label_s = str(label).strip()
            if not label_s:
                continue
            role = role_from_label(label_s)
            if role is None:
                continue
            last_primary_row = rr  # advance even if name is blank
            name = ws.cell(rr, name_col).value
            if name is not None and str(name).strip():
                roles[role] = str(name).strip()

        bu_idx = 1
        blank_streak = 0
        for rr in range(last_primary_row + 1, stop_row + 1):
            label = ws.cell(rr, label_col).value
            label_s = str(label).strip().upper() if label is not None else ""
            if label_s.startswith(PRIMARY_LABELS):
                continue

            name = ws.cell(rr, name_col).value
            name_s = str(name).strip() if name is not None else ""

            if not name_s:
                blank_streak += 1
                if blank_streak >= 2:
                    break
                continue

            blank_streak = 0
            if name_s.upper() in set(PRIMARY_LABELS) | set(BU_ROLES) | {"BU"}:
                continue

            roles[f"BU{bu_idx}"] = name_s
            bu_idx += 1
            if bu_idx > 4:
                break

        if roles or include_empty_days:
            schedule[d_iso] = roles

    if include_empty_days:
        _, last_day = calendar.monthrange(year, month)
        for dd in range(1, last_day + 1):
            iso = date(year, month, dd).isoformat()
            schedule.setdefault(iso, {})

    return year, month, schedule


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="Input filled template .xlsx")
    ap.add_argument("--out", dest="out", default="", help="Output schedule.json path (default: YYYY-MM_schedule.json)")
    ap.add_argument("--sheet", dest="sheet", default="", help="Optional sheet name (default: first sheet)")
    ap.add_argument("--include-empty-days", action="store_true", help="Include all days of month with {} when no roles found.")
    ap.add_argument("--date-col-offset", type=int, default=0, help="Column offset from day-number cell to the date column (default 0).")
    ap.add_argument("--label-col-offset", type=int, default=1, help="Column offset from day-number cell to label column (default 1).")
    ap.add_argument("--name-col-offset", type=int, default=2, help="Column offset from day-number cell to name column (default 2).")
    ap.add_argument("--max-scan", type=int, default=40, help="Max rows to scan downward per day-square (default 40).")
    ap.add_argument("--strict", action="store_true", help="Turn warnings into errors.")
    args = ap.parse_args()

    year, month, sched = export_template(
        args.inp,
        sheet_name=args.sheet.strip(),
        include_empty_days=args.include_empty_days,
        date_col_offset=args.date_col_offset,
        label_col_offset=args.label_col_offset,
        name_col_offset=args.name_col_offset,
        max_scan=args.max_scan,
        strict=args.strict,
    )

    out_path = args.out.strip() or f"{year:04d}-{month:02d}_schedule.json"
    Path(out_path).write_text(json.dumps(sched, indent=2), encoding="utf-8")
    print(f"✅ Exported {len(sched)} day entries for {year:04d}-{month:02d} → {out_path}")


if __name__ == "__main__":
    main()
