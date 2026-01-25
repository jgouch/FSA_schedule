#!/usr/bin/env python3
"""
export_schedule_from_template_v2.py

Export an existing filled calendar-template workbook (Excel) into schedule.json format
for use as --prev-schedule in the scheduler.

Why v2:
- Friendlier month header parsing (supports "JANUARY 2026", "Jan 2026", "Sept. 2026", etc.)
- Day cell detection accepts ints, floats like 1.0, or strings like "1"
- Less brittle vertical scanning: scans until it hits the next day-number cell in the same column
  (or a safe max window), and uses "blank-streak" stopping for BU names.

Template assumptions (same as your current calendar layout):
- Each day "square" has:
    date number in cell (r,c)
    role labels in column (c+1)
    names in column (c+2)
- Primaries are detected by label text starting with PN / AN / W (case-insensitive)
- Backup roles are unlabeled: extra names in the name column after primaries become BU1..BU4
  top-to-bottom.

Usage:
  python3 export_schedule_from_template_v2.py \
    --in "/path/January_2026_Filled_Template copy.xlsx" \
    --out "/path/jan_2026_schedule.json"
"""

import argparse
import json
import re
from datetime import date
from pathlib import Path
from typing import Dict, Tuple, List, Optional

import openpyxl


MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


def parse_month_year(header_val) -> Tuple[int, int]:
    """
    Parse month/year from common header strings like:
      "JANUARY 2026", "Jan 2026", "Sept. 2026", "February-2026"
    """
    if header_val is None:
        raise ValueError("Month/year header cell is blank (expected something like 'January 2026').")

    s = str(header_val).strip()
    # Extract month token + year
    m = re.search(r"([A-Za-z]+)\.?\s*[-/ ]\s*(\d{4})", s)
    if not m:
        m = re.search(r"([A-Za-z]+)\.?\s+(\d{4})", s)
    if not m:
        raise ValueError(f"Could not parse month/year from header '{s}'. Expected like 'January 2026'.")

    mon_token = m.group(1).strip().lower()
    year = int(m.group(2))

    if mon_token not in MONTHS:
        # Try first 3 letters fallback
        mon3 = mon_token[:3]
        if mon3 in MONTHS:
            month = MONTHS[mon3]
        else:
            raise ValueError(f"Unrecognized month token '{mon_token}' in header '{s}'. "
                             f"Use a standard month name like 'January 2026'.")
    else:
        month = MONTHS[mon_token]

    return year, month


def as_day_number(v) -> Optional[int]:
    """Return 1..31 if v looks like a day number, else None."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v if 1 <= v <= 31 else None
    if isinstance(v, float):
        if abs(v - round(v)) < 1e-9:
            iv = int(round(v))
            return iv if 1 <= iv <= 31 else None
        return None
    s = str(v).strip()
    if s.isdigit():
        iv = int(s)
        return iv if 1 <= iv <= 31 else None
    return None


def export_template(path_in: str, sheet_name: str = "") -> Tuple[int, int, Dict[str, Dict[str, str]]]:
    wb = openpyxl.load_workbook(path_in, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    # Month/year label is typically in B1, but we’ll fallback to A1 if needed.
    header = ws.cell(1, 2).value
    if header is None or str(header).strip() == "":
        header = ws.cell(1, 1).value

    year, month = parse_month_year(header)

    # Find all day-number cells (1..31)
    date_cells: List[Tuple[int, int, int]] = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            dn = as_day_number(cell.value)
            if dn is not None:
                date_cells.append((cell.row, cell.column, dn))

    if not date_cells:
        raise ValueError("No day-number cells (1..31) were found. "
                         "Check that the calendar template contains day numbers as values.")

    schedule: Dict[str, Dict[str, str]] = {}

    for r, c, daynum in date_cells:
        d_iso = date(year, month, daynum).isoformat()
        label_col = c + 1
        name_col = c + 2

        roles: Dict[str, str] = {}

        # Scan downward until we hit another day-number cell in the same column (next week row), or max window.
        max_scan = 22  # safe, generous
        stop_row = r + max_scan
        for rr in range(r + 1, r + max_scan + 1):
            if rr > ws.max_row:
                break
            # stop if we hit another day number in the same column c
            if rr != r:
                dn2 = as_day_number(ws.cell(rr, c).value)
                if dn2 is not None:
                    stop_row = rr - 1
                    break

        # Primary roles: any rows in [r+1..stop_row] where label cell starts with PN/AN/W
        last_primary_row = r
        for rr in range(r + 1, stop_row + 1):
            label = ws.cell(rr, label_col).value
            name = ws.cell(rr, name_col).value
            if label is None:
                continue
            label_s = str(label).strip().upper()
            if not label_s:
                continue
            if not (label_s.startswith("PN") or label_s.startswith("AN") or label_s.startswith("W")):
                continue
            # name may be blank on holidays or intentionally empty; only store if present
            if name is not None and str(name).strip() != "":
                roles["PN" if label_s.startswith("PN") else ("AN" if label_s.startswith("AN") else "W")] = str(name).strip()
            last_primary_row = rr

        # BU roles: unlabeled names after primaries in the name column
        bu_idx = 1
        blank_streak = 0
        for rr in range(last_primary_row + 1, stop_row + 1):
            name = ws.cell(rr, name_col).value
            label = ws.cell(rr, label_col).value
            name_s = str(name).strip() if name is not None else ""
            label_s = str(label).strip().upper() if label is not None else ""

            # If label column has any primary label, don't treat as BU row.
            if label_s.startswith(("PN", "AN", "W")):
                continue

            if not name_s:
                blank_streak += 1
                if blank_streak >= 2:
                    break
                continue

            blank_streak = 0

            # Guard: don't accidentally capture header-y text
            if name_s.upper() in {"PN", "AN", "W", "BU", "BU1", "BU2", "BU3", "BU4"}:
                continue

            roles[f"BU{bu_idx}"] = name_s
            bu_idx += 1
            if bu_idx > 4:
                break

        if roles:
            schedule[d_iso] = roles

    return year, month, schedule


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="Input filled template .xlsx")
    ap.add_argument("--out", dest="out", required=True, help="Output schedule.json path")
    ap.add_argument("--sheet", dest="sheet", default="", help="Optional sheet name (default: first sheet)")
    args = ap.parse_args()

    year, month, sched = export_template(args.inp, sheet_name=args.sheet.strip())
    Path(args.out).write_text(json.dumps(sched, indent=2), encoding="utf-8")
    print(f"✅ Exported {len(sched)} days for {year:04d}-{month:02d} → {args.out}")


if __name__ == "__main__":
    main()
