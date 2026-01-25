#!/usr/bin/env python3
import argparse
import json
import re
from datetime import date
from pathlib import Path
from typing import Tuple
import openpyxl

def parse_month_year(s: str) -> Tuple[int, int]:
    s = str(s).strip()
    m = re.search(r"([A-Za-z.]+)\s+(\d{4})", s)
    if not m:
        raise ValueError(f"Cannot parse month/year from '{s}'")
    month_name = m.group(1).lower().rstrip(".")
    year = int(m.group(2))
    months = {
        "january": 1, "jan": 1,
        "february": 2, "feb": 2,
        "march": 3, "mar": 3,
        "april": 4, "apr": 4,
        "may": 5,
        "june": 6, "jun": 6,
        "july": 7, "jul": 7,
        "august": 8, "aug": 8,
        "september": 9, "sep": 9, "sept": 9,
        "october": 10, "oct": 10,
        "november": 11, "nov": 11,
        "december": 12, "dec": 12,
    }
    if month_name not in months:
        raise ValueError(f"Unrecognized month name '{month_name}' in '{s}'")
    return year, months[month_name]


def parse_day(value) -> int | None:
    if isinstance(value, int):
        return value if 1 <= value <= 31 else None
    if isinstance(value, float) and value.is_integer():
        day = int(value)
        return day if 1 <= day <= 31 else None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.isdigit():
            day = int(stripped)
            return day if 1 <= day <= 31 else None
    return None

def export_template(path_in: str):
    wb = openpyxl.load_workbook(path_in, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header = ws.cell(1, 2).value
    year, month = parse_month_year(header)

    date_cells = []
    for row in ws.iter_rows():
        for cell in row:
            daynum = parse_day(cell.value)
            if daynum is not None:
                date_cells.append((cell.row, cell.column, daynum))

    schedule = {}
    for r, c, daynum in date_cells:
        d = date(year, month, daynum).isoformat()
        label_col = c + 1
        name_col = c + 2
        roles = {}

        last_primary_row = r
        empty_primary = 0
        for rr in range(r + 1, r + 20):
            label = ws.cell(rr, label_col).value
            name = ws.cell(rr, name_col).value
            if not label and not name:
                empty_primary += 1
                if empty_primary >= 3:
                    break
                continue
            empty_primary = 0
            if not label or not name:
                continue
            label = str(label).strip().upper()
            name = str(name).strip()
            if label.startswith("PN"):
                roles["PN"] = name
                last_primary_row = rr
            elif label.startswith("AN"):
                roles["AN"] = name
                last_primary_row = rr
            elif label.startswith("W"):
                roles["W"] = name
                last_primary_row = rr

        bu_idx = 1
        empty_bu = 0
        for rr in range(last_primary_row + 1, r + 30):
            name = ws.cell(rr, name_col).value
            if not name:
                empty_bu += 1
                if empty_bu >= 3:
                    break
                continue
            empty_bu = 0
            roles[f"BU{bu_idx}"] = str(name).strip()
            bu_idx += 1

        if roles:
            schedule[d] = roles

    return year, month, schedule

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True)
    ap.add_argument("--out", dest="out", required=True)
    args = ap.parse_args()

    year, month, sched = export_template(args.inp)
    Path(args.out).write_text(json.dumps(sched, indent=2), encoding="utf-8")
    print(f"✅ Exported {len(sched)} days → {args.out}")

if __name__ == "__main__":
    main()
