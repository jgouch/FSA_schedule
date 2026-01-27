#!/usr/bin/env python3
"""
build_month_schedule_v7.py

End-to-end month builder:
- Reads TimeOff.xlsx (monthly request sheet + Sales Ranking tab)
- Reads prior month schedule JSON for cross-month constraints
- Generates schedule for target month (primaries then BU)
- Exports:
    - schedule JSON (machine truth)
    - Excel calendar workbook (no template required) matching your geometry + colors

LOCKED CALENDAR GEOMETRY (per John):
- Grid weeks: 5 (rows 3..42), each week is 8 rows tall
- Each day block: 3 columns x 8 rows
- Day number: top-right cell of block (col_end, week_start_row)
- Role mapping (R = week_start_row, cols = start/mid/end):
    PN label: (start, R+1), PN name: (mid, R+1)
    AN label: (start, R+2), AN name: (mid, R+2)
    W  label: (start, R+3), W  name: (mid, R+3)
    BU1 name: (mid, R+4)
    BU2 name: (mid, R+5)
    BU3 name: (mid, R+6)
    BU4 name (push-week Saturday only): (mid, R+7)

Color rules:
- Sundays (real dates): #FFF2CC
- Non-month blocks:     #F2F2F2
- Other in-month days:  no fill

Gridlines must be OFF.
"""

from __future__ import annotations

import argparse
import calendar
import json
import random
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ----------------------------
# Constants / settings
# ----------------------------

SUNDAY_FILL = "FFF2CC"
NONMONTH_FILL = "F2F2F2"

FONT_MAIN = Font(name="Calibri", size=11)
FONT_HEADER = Font(name="Calibri", size=18, bold=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

GRID_WEEK_START_ROW = 3
WEEK_HEIGHT = 8
DAY_WIDTH = 3
SUN_START_COL = 2  # B
MAX_WEEKS = 6
ROLE_PRIMARY = ["PN", "AN", "W"]
ROLE_BUS = ["BU1", "BU2", "BU3", "BU4"]

# Park Lawn observed holidays (no roles on observed date)
OBSERVED_HOLIDAYS_BY_YEAR = {
    2026: {
        date(2026, 1, 1),   # New Year's Day
        date(2026, 5, 25),  # Memorial Day
        date(2026, 7, 3),   # Independence Day observed
        date(2026, 9, 7),   # Labor Day
        date(2026, 11, 26), # Thanksgiving
        date(2026, 12, 25), # Christmas
    }
}

# Default roster (display names used in Excel). You can expand later.
DEFAULT_ROSTER = ["Greg", "Mark", "Dawn", "Will", "CJ", "Kyle"]

# Seniority tie-break (only used if sales ranking missing)
SENIORITY_ORDER = ["Mark", "Dawn", "Will", "Kyle", "CJ", "Greg"]


# ----------------------------
# Helpers
# ----------------------------

def weekday_sun0(d: date) -> int:
    """Return weekday where Sunday=0..Saturday=6."""
    return (d.weekday() + 1) % 7

def is_sunday(d: date) -> bool:
    return weekday_sun0(d) == 0

def is_saturday(d: date) -> bool:
    return weekday_sun0(d) == 6

def is_weekday(d: date) -> bool:
    return d.weekday() < 5

def daterange(a: date, b: date) -> List[date]:
    out = []
    cur = a
    while cur <= b:
        out.append(cur)
        cur += timedelta(days=1)
    return out

def month_days(year: int, month: int) -> List[date]:
    n = calendar.monthrange(year, month)[1]
    return [date(year, month, d) for d in range(1, n+1)]

def quarter_tag_for_month(year: int, month: int) -> str:
    q = (month - 1) // 3 + 1
    return f"{year}Q{q}"

def clamp_int(x, default=0) -> int:
    try:
        return int(x)
    except Exception:
        return default

def norm_name(x) -> str:
    return re.sub(r"\s+", " ", str(x).strip())

def parse_month_arg(s: str) -> Tuple[int, int]:
    m = re.match(r"^\s*(\d{4})-(\d{2})\s*$", s)
    if not m:
        raise ValueError("Month must be YYYY-MM, e.g. 2026-02")
    return int(m.group(1)), int(m.group(2))


def compute_observed_holidays_us(year: int) -> Set[date]:
    """Compute a small set of observed US holidays used in scheduling."""
    out: Set[date] = set()

    # New Year's Day
    out.add(date(year, 1, 1))

    # Memorial Day: last Monday in May
    d = date(year, 5, 31)
    while d.weekday() != 0:  # Monday
        d -= timedelta(days=1)
    out.add(d)

    # Independence Day observed: if July 4 is Sat -> Fri 3; if Sun -> Mon 5; else Jul 4
    july4 = date(year, 7, 4)
    if july4.weekday() == 5:
        out.add(date(year, 7, 3))
    elif july4.weekday() == 6:
        out.add(date(year, 7, 5))
    else:
        out.add(july4)

    # Labor Day: first Monday in September
    d = date(year, 9, 1)
    while d.weekday() != 0:
        d += timedelta(days=1)
    out.add(d)

    # Thanksgiving: fourth Thursday in November
    d = date(year, 11, 1)
    while d.weekday() != 3:  # Thursday
        d += timedelta(days=1)
    out.add(d + timedelta(days=21))

    # Christmas
    out.add(date(year, 12, 25))

    return out

def observed_holidays_for_year(year: int) -> Set[date]:
    return OBSERVED_HOLIDAYS_BY_YEAR.get(year, compute_observed_holidays_us(year))

def parse_excel_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date format: {value!r}")


# ----------------------------
# Inputs from Excel (TimeOff.xlsx)
# ----------------------------

@dataclass
class TimeOffRule:
    name: str
    d: date
    hard: bool = True
    avoid_roles: Set[str] = None  # e.g., {"PN","AN"}

def load_timeoff_from_xlsx(xlsx_path: str, sheet_name: str) -> List[TimeOffRule]:
    """    Load time-off / avoidance rules from a sheet in TimeOff.xlsx.

    Supported sheet formats (headers in row 1, case-insensitive):

      A) Single-day rows:
         Date | Name | Hard | AvoidRoles

      B) Date ranges:
         Start | End | Name | Hard | AvoidRoles

      C) Mixed:
         Date | Start | End | Name | Hard | AvoidRoles
         - If Date is present, it wins.
         - Else Start/End is expanded inclusive.

    Notes:
    - Date/Start/End can be Excel dates or strings parsed by parse_excel_date().
    - Hard defaults to TRUE when blank; accepts TRUE/FALSE, 1/0, Y/N, YES/NO.
    - AvoidRoles is optional, comma-separated (e.g., "PN,AN").
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing timeoff sheet '{sheet_name}' in {xlsx_path}")
    ws = wb[sheet_name]

    # Map headers
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        headers[str(v).strip().lower()] = c

    def col_optional(name: str) -> Optional[int]:
        return headers.get(name.lower(), None)

    def col_required(name: str) -> int:
        k = name.lower()
        if k not in headers:
            raise ValueError(
                f"TimeOff sheet missing column '{name}' (headers row 1). Found: {list(headers.keys())}"
            )
        return headers[k]

    c_name = col_required("name")
    c_date = col_optional("date")
    c_start = col_optional("start")
    c_end = col_optional("end")
    c_hard = col_optional("hard")
    c_avoid = col_optional("avoidroles")

    def parse_hard(v) -> bool:
        if v is None or str(v).strip() == "":
            return True
        s = str(v).strip().lower()
        return s in ("true", "1", "yes", "y")

    rules: List[TimeOffRule] = []

    for r in range(2, ws.max_row + 1):
        nv = ws.cell(r, c_name).value
        if nv is None or str(nv).strip() == "":
            continue
        name = norm_name(nv)

        # Determine date(s) represented by this row
        dates: List[date] = []
        dv = ws.cell(r, c_date).value if c_date else None
        sv = ws.cell(r, c_start).value if c_start else None
        ev = ws.cell(r, c_end).value if c_end else None

        if dv not in (None, ""):
            dates = [parse_excel_date(dv)]
        elif sv not in (None, "") and ev not in (None, ""):
            start_d = parse_excel_date(sv)
            end_d = parse_excel_date(ev)
            if end_d < start_d:
                raise ValueError(f"End before Start at row {r}: {start_d} .. {end_d}")
            cur = start_d
            while cur <= end_d:
                dates.append(cur)
                cur += timedelta(days=1)
        else:
            # No usable date info
            continue

        hard = parse_hard(ws.cell(r, c_hard).value) if c_hard else True

        avoid_roles: Set[str] = set()
        if c_avoid:
            av = ws.cell(r, c_avoid).value
            if av:
                parts = [p.strip().upper() for p in str(av).split(",") if p.strip()]
                avoid_roles = set(parts)

        for dd in dates:
            rules.append(TimeOffRule(name=name, d=dd, hard=hard, avoid_roles=avoid_roles))

    return rules


def load_roster_from_sales_ranking(
    xlsx_path: str,
    sheet_name: str,
    year: int,
    month: int,
) -> List[str]:
    """    Derive the active roster dynamically from the Sales Ranking sheet.

    Uses the row for the target quarter (YYYYQ#) if present, otherwise the latest available.
    Returns the list of names in ranking order (highest -> lowest) for that row.

    Falls back to DEFAULT_ROSTER if the sheet has no usable rows.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing sheet '{sheet_name}' in {xlsx_path}")
    ws = wb[sheet_name]

    target_period = quarter_tag_for_month(year, month)
    rows: Dict[str, List[str]] = {}

    for r in range(2, ws.max_row + 1):
        period = ws.cell(r, 1).value
        if not period:
            continue
        period = str(period).strip()
        ranks: List[str] = []
        for c in range(2, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None or str(v).strip() == "":
                continue
            ranks.append(norm_name(v))
        if period and ranks:
            rows[period] = ranks

    def key(p: str):
        m = re.match(r"^\s*(\d{4})\s*Q([1-4])\s*$", p)
        if not m:
            return (-1, -1)
        return (int(m.group(1)), int(m.group(2)))

    if target_period in rows:
        return rows[target_period]

    candidates = sorted(rows.keys(), key=key)
    if candidates:
        return rows[candidates[-1]]

    return DEFAULT_ROSTER[:]


def load_roster_from_json(path: str) -> List[str]:
    """Load roster list from a JSON file containing an array of names."""
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(data, list) or not data:
        raise ValueError("Roster JSON must be a non-empty JSON array of names")
    return [norm_name(x) for x in data if str(x).strip()]


def load_payperiods_from_json(path: str) -> List[Tuple[date, date]]:
    """    Load pay periods from JSON.

    Expected format: a JSON array of objects like:
      {"start":"YYYY-MM-DD","end":"YYYY-MM-DD"}

    Periods are treated as inclusive [start, end].
    """
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(data, list) or not data:
        raise ValueError("Payperiods JSON must be a non-empty JSON array")
    out: List[Tuple[date, date]] = []
    for obj in data:
        if not isinstance(obj, dict) or "start" not in obj or "end" not in obj:
            raise ValueError("Each pay period must be an object with 'start' and 'end'")
        s = datetime.strptime(str(obj["start"]).strip(), "%Y-%m-%d").date()
        e = datetime.strptime(str(obj["end"]).strip(), "%Y-%m-%d").date()
        if e < s:
            raise ValueError(f"Pay period end before start: {s} .. {e}")
        out.append((s, e))
    out.sort(key=lambda t: t[0])
    return out

def load_sales_ranking_from_timeoff_xlsx(
    xlsx_path: str,
    sheet_name: str,
    year: int,
    month: int,
    roster_names: List[str],
) -> List[str]:
    """
    Sheet format:
      Period | Rank 1 | Rank 2 | Rank 3 | Rank 4 | Rank 5 | Rank 6

    Example Period: 2025Q4
    Returns list highest->lowest.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing sheet '{sheet_name}' in {xlsx_path}")
    ws = wb[sheet_name]

    target_period = quarter_tag_for_month(year, month)

    rows: Dict[str, List[str]] = {}
    for r in range(2, ws.max_row + 1):
        period = ws.cell(r, 1).value
        if not period:
            continue
        period = str(period).strip()
        ranks = []
        for c in range(2, 8):  # B..G
            v = ws.cell(r, c).value
            if v is None or str(v).strip() == "":
                continue
            ranks.append(norm_name(v))
        if period and ranks:
            rows[period] = ranks

    def key(p: str):
        m = re.match(r"^\s*(\d{4})\s*Q([1-4])\s*$", p)
        if not m:
            return (-1, -1)
        return (int(m.group(1)), int(m.group(2)))

    if target_period in rows:
        order = rows[target_period]
    else:
        # fall back to latest available
        candidates = sorted(rows.keys(), key=key)
        order = rows[candidates[-1]] if candidates else []

    roster_set = set(roster_names)
    order_set = set(order)

    if not order:
        # fallback to seniority, then append any remaining roster members
        fallback = [n for n in SENIORITY_ORDER if n in roster_set]
        fallback += [n for n in roster_names if n not in fallback]
        return fallback

    missing = roster_set - order_set
    extra = order_set - roster_set
    if missing or extra:
        raise ValueError(
            f"Sales Ranking invalid for {target_period}. "
            f"Missing: {sorted(missing)} Extra/Unknown: {sorted(extra)} "
            f"Roster: {roster_names} Got: {order}"
        )
    return order


# ----------------------------
# Previous schedule JSON (carryover)
# ----------------------------

def load_schedule_json(path: str) -> Dict[str, Dict[str, str]]:
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Schedule JSON must be a dict of ISO-date -> role dict")
    return data

def role_of(prev: Dict[str, Dict[str, str]], d: date, name: str) -> Optional[str]:
    roles = prev.get(d.isoformat(), {})
    for r, n in roles.items():
        if n == name:
            return r
    return None

def worked_prev(prev: Dict[str, Dict[str, str]], d: date, name: str) -> bool:
    roles = prev.get(d.isoformat(), {})
    return any(n == name for n in roles.values())


# ----------------------------
# Scheduling rules
# ----------------------------

def roles_required_for_day(d: date,
                           push_week_days: Set[date],
                           observed_holidays: Set[date]) -> List[str]:
    """
    Per your rules:
    - No roles on observed holidays.
    - Sunday: PN only (5 hrs)
    - Mon-Thu: PN, AN, W + BU1/BU2 normally; Monday wants BU3 if possible.
    - Friday: PN, AN, W + BU1/BU2 (and BU3 if push week / all hands)
    - Saturday: PN, AN + BU1/BU2 normally; push-week Saturday has BU3 + BU4 (no W)
    """
    if d in observed_holidays:
        return []

    if is_sunday(d):
        return ["PN"]

    # base primaries
    req = ["PN", "AN"]
    if is_weekday(d):
        req.append("W")

    # BU rules
    if d in push_week_days:
        # all hands on deck M-F and Saturday in push week
        if is_saturday(d):
            # no W; add BU3 + BU4
            req += ["BU1", "BU2", "BU3", "BU4"]
        else:
            # weekdays push: we want 6 people total -> 3 BU roles
            req += ["BU1", "BU2", "BU3"]
    else:
        # normal weeks
        if weekday_sun0(d) == 1:  # Monday
            req += ["BU1", "BU2", "BU3"]  # full team if possible
        elif is_weekday(d):
            req += ["BU1", "BU2"]
        else:
            # Saturday (non-push): no W
            req += ["BU1", "BU2"]

    # Ensure order: primaries first then BUs
    return req

def hours_for_role(d: date, role: str, observed_holidays: Set[date]) -> int:
    if d in observed_holidays:
        return 0
    if is_sunday(d) and role == "PN":
        return 5
    return 8

def week_start_sun(d: date) -> date:
    return d - timedelta(days=weekday_sun0(d))

def push_days_for_month(year: int, month: int) -> Set[date]:
    days = set()
    # last Mon-Fri
    last_dom = calendar.monthrange(year, month)[1]
    last = date(year, month, last_dom)
    cur = last
    while cur.weekday() != 4:
        cur -= timedelta(days=1)
    last_friday = cur
    last_monday = last_friday - timedelta(days=4)
    for i in range(5):
        days.add(last_monday + timedelta(days=i))
    # add Saturday after Friday if within month
    sat = last_friday + timedelta(days=1)
    if sat.month == month:
        days.add(sat)
    return days


# ----------------------------
# Solver
# ----------------------------

@dataclass
class BuildConfig:
    year: int
    month: int
    roster: List[str]
    sales_order: List[str]   # highest -> lowest
    max_consecutive_days: int = 5
    lookback_days: int = 14
    rng_seed: int = 11

@dataclass
class Constraints:
    hard_off: Dict[date, Set[str]]
    avoid_roles: Dict[Tuple[date, str], Set[str]]  # (date, role)->names who avoid that role

def compile_constraints(timeoff: List[TimeOffRule]) -> Constraints:
    hard_off: Dict[date, Set[str]] = {}
    avoid_roles: Dict[Tuple[date, str], Set[str]] = {}
    for r in timeoff:
        if r.hard:
            hard_off.setdefault(r.d, set()).add(r.name)
        if r.avoid_roles:
            for role in r.avoid_roles:
                avoid_roles.setdefault((r.d, role.upper()), set()).add(r.name)
    return Constraints(hard_off=hard_off, avoid_roles=avoid_roles)

def compute_primary_targets(
    days: List[date],
    push_days: Set[date],
    roster: List[str],
    sales_order: List[str],
    observed_holidays: Set[date],
) -> Dict[str, Dict[str, int]]:
    """    Targets for PN/AN/W across the month (roughly even).

    Returns targets[role][name] = target_count.

    Remainders (extras after even split) are distributed by sales_order priority
    (highest -> lowest), but only among names present in roster.
    """
    opp = {"PN": 0, "AN": 0, "W": 0}
    for d in days:
        req = roles_required_for_day(d, push_days, observed_holidays)
        for role in ("PN", "AN", "W"):
            if role in req:
                opp[role] += 1

    n = len(roster)
    if n == 0:
        raise ValueError("Roster is empty")

    priority = [x for x in sales_order if x in roster] + [x for x in roster if x not in sales_order]

    targets: Dict[str, Dict[str, int]] = {r: {} for r in opp}
    for role, total in opp.items():
        base = total // n
        rem = total % n
        for name in roster:
            targets[role][name] = base
        for i in range(rem):
            targets[role][priority[i % len(priority)]] += 1

    return targets

def pick_week_index_for_day(year: int, month: int, day: int) -> Tuple[int, int, int, int]:
    """
    Returns (R, col_start, col_mid, col_end) for a date in this month using calendar math
    matching the fixed Excel grid (not scanning).
    """
    first = date(year, month, 1)
    offset = weekday_sun0(first)
    idx = offset + (day - 1)
    week_index = idx // 7
    dow = idx % 7
    if week_index < 0 or week_index >= MAX_WEEKS:
        raise ValueError("Date falls outside 5-week grid. Check template assumptions.")
    R = GRID_WEEK_START_ROW + WEEK_HEIGHT * week_index
    col_start = SUN_START_COL + DAY_WIDTH * dow
    col_mid = col_start + 1
    col_end = col_start + 2
    return R, col_start, col_mid, col_end

def build_schedule(config: BuildConfig,
                   prev_sched: Dict[str, Dict[str, str]],
                   cons: Constraints,
                   payperiods: Optional[List[Tuple[date, date]]] = None) -> Dict[str, Dict[str, str]]:
    
    rng = random.Random(config.rng_seed)
    year, month = config.year, config.month
    days = month_days(year, month)
    
    # Pre-calculate holidays for the target year once
    observed_holidays = observed_holidays_for_year(year)

    push_days = push_days_for_month(year, month)

    def week_contains_push(week_start: date) -> bool:
        # If any day in this Sunday-start week is in push_days, weekly caps are relaxed.
        return any((week_start + timedelta(days=i)) in push_days for i in range(7))

    def hours_for_any_role_on_day(dd: date, roles_map: Dict[str, Dict[str, str]], name: str) -> int:
        # defensive: return max role hours for the person on a date
        day_hours = 0
        roles = roles_map.get(dd.isoformat(), {})
        # If dd is in same year, use local observed_holidays, else calc
        hols = observed_holidays if dd.year == year else observed_holidays_for_year(dd.year)
        
        for rr, nn in roles.items():
            if nn == name:
                day_hours = max(day_hours, hours_for_role(dd, rr, hols))
        return day_hours

    def week_hours(name: str, week_start: date) -> int:
        total = 0
        first_day = date(year, month, 1)
        last_day = date(year, month, calendar.monthrange(year, month)[1])
        for i in range(7):
            dd = week_start + timedelta(days=i)
            if dd.year == year and dd.month == month:
                total += hours_for_any_role_on_day(dd, schedule, name)
            elif dd < first_day:
                # carryover from previous schedule only
                roles = prev_sched.get(dd.isoformat(), {})
                day_hours = 0
                hols = observed_holidays if dd.year == year else observed_holidays_for_year(dd.year)
                for rr, nn in roles.items():
                    if nn == name:
                        day_hours = max(day_hours, hours_for_role(dd, rr, hols))
                total += day_hours
            else:
                # ignore future days beyond this month
                continue
        return total

    def week_days_worked(name: str, week_start: date) -> int:
        # Count distinct days in the week where the person works any role.
        cnt = 0
        first_day = date(year, month, 1)
        last_day = date(year, month, calendar.monthrange(year, month)[1])
        for i in range(7):
            dd = week_start + timedelta(days=i)
            if dd > last_day:
                continue
            if dd.year == year and dd.month == month:
                if hours_for_any_role_on_day(dd, schedule, name) > 0:
                    cnt += 1
            elif dd < first_day:
                roles = prev_sched.get(dd.isoformat(), {})
                if any(nn == name for nn in roles.values()):
                    cnt += 1
        return cnt

    def would_exceed_week_cap(d: date, role: str, name: str) -> bool:
        ws = week_start_sun(d)
        if week_contains_push(ws):
            return False  # fully relaxed during any week that includes push-week days
        cur = week_hours(name, ws)
        add = hours_for_role(d, role, observed_holidays)
        return (cur + add) > 40

    def would_exceed_week_days_cap(d: date, name: str) -> bool:
        ws = week_start_sun(d)
        if week_contains_push(ws):
            return False
        # one role per person per day is enforced elsewhere, so +1 is safe
        return (week_days_worked(name, ws) + 1) > 5

    def find_pay_period(d: date) -> Optional[Tuple[date, date]]:
        if not payperiods:
            return None
        for s, e in payperiods:
            if s <= d <= e:
                return (s, e)
        return None

    def pay_period_hours(name: str, pp: Tuple[date, date]) -> int:
        s, e = pp
        total = 0
        first_day = date(year, month, 1)
        # Sum across the pay period range; include carryover from prev schedule if pp starts before month
        dd = s
        while dd <= e:
            if dd.year == year and dd.month == month:
                total += hours_for_any_role_on_day(dd, schedule, name)
            elif dd < first_day:
                roles = prev_sched.get(dd.isoformat(), {})
                day_hours = 0
                hols = observed_holidays if dd.year == year else observed_holidays_for_year(dd.year)
                for rr, nn in roles.items():
                    if nn == name:
                        day_hours = max(day_hours, hours_for_role(dd, rr, hols))
                total += day_hours
            dd += timedelta(days=1)
        return total

    def would_exceed_payperiod_cap(d: date, role: str, name: str) -> bool:
        pp = find_pay_period(d)
        if not pp:
            return False
        cur = pay_period_hours(name, pp)
        add = hours_for_role(d, role, observed_holidays)
        return (cur + add) > 80

    # Primary targets: PN/AN should be as even as possible across the month.
    # We compute targets based on total opportunities; extras go to top of sales_order.
    # We'll use these as soft scoring during search, not hard limits.
    targets = compute_primary_targets(days, push_days, config.roster, config.sales_order, observed_holidays)

    # Helper: last role day (carryover)
    def last_role(name: str, d: date) -> Optional[str]:
        # look back within month first, then prev_sched
        for k in range(1, config.lookback_days + 1):
            p = d - timedelta(days=k)
            if p.month == month:
                roles = schedule.get(p.isoformat(), {})
                for rr, nn in roles.items():
                    if nn == name:
                        return rr
            else:
                rr = role_of(prev_sched, p, name)
                if rr:
                    return rr
        return None

    def consecutive_streak(name: str, d: date) -> int:
        # count consecutive days worked ending yesterday, across boundary
        streak = 0
        for k in range(1, config.lookback_days + 1):
            p = d - timedelta(days=k)
            worked = False
            if p.month == month:
                worked = any(n == name for n in schedule.get(p.isoformat(), {}).values())
            else:
                worked = worked_prev(prev_sched, p, name)
            if worked:
                streak += 1
            else:
                break
        return streak

    # Sunday rotation: 1 Sunday in 6, and if PN on Sunday -> AN on Saturday before
    # We treat Sunday PN as a hard pre-assignment based on previous month last Sunday assignee if present.
    roster = config.roster[:]
    sales_order = config.sales_order[:]

    def get_prev_sunday_assignee() -> Optional[str]:
        # find last Sunday in prev_sched and who had PN
        prev_dates = sorted(prev_sched.keys())
        for ds in reversed(prev_dates):
            dd = datetime.strptime(ds, "%Y-%m-%d").date()
            if is_sunday(dd):
                roles = prev_sched.get(ds, {})
                pn = roles.get("PN")
                if pn in roster:
                    return pn
        return None

    sunday_assignees = {}
    prev_sun = get_prev_sunday_assignee()
    sundays = [d for d in days if is_sunday(d) and d not in observed_holidays]
    if sundays:
        start_idx = 0
        if prev_sun and prev_sun in roster:
            # next in rotation after previous assignee
            start_idx = (roster.index(prev_sun) + 1) % len(roster)
        for i, sd in enumerate(sundays):
            who = roster[(start_idx + i) % len(roster)]
            sunday_assignees[sd] = who

    # Prepare schedule dict
    schedule: Dict[str, Dict[str, str]] = {d.isoformat(): {} for d in days}

    # Pre-assign Sunday PN and Saturday-before AN
    # Enforce pairing (Sunday PN == Saturday-before AN) when Saturday exists in-month.
    # Respect hard off, avoid_roles, no same-role back-to-back, and max_consecutive_days across month boundary.
    def can_preassign(d: date, role: str, name: str) -> bool:
        if d in observed_holidays:
            return False
        if name in cons.hard_off.get(d, set()):
            return False
        if name in cons.avoid_roles.get((d, role), set()):
            return False
        if name in schedule[d.isoformat()].values():
            return False
        prev_role = last_role(name, d)
        if prev_role == role:
            return False
        if consecutive_streak(name, d) >= config.max_consecutive_days:
            return False
        if would_exceed_week_cap(d, role, name):
            return False
        if would_exceed_week_days_cap(d, name):
            return False
        if would_exceed_payperiod_cap(d, role, name):
            return False
        return True


    for sd, who0 in sunday_assignees.items():
        sat = sd - timedelta(days=1)
        need_sat_pair = (sat.month == month and sat not in observed_holidays)

        chosen = None
        for step in range(len(roster)):
            cand = roster[(roster.index(who0) + step) % len(roster)]
            if not can_preassign(sd, "PN", cand):
                continue
            if need_sat_pair and not can_preassign(sat, "AN", cand):
                continue
            chosen = cand
            break

        if chosen is None:
            raise RuntimeError(f"No valid Sunday PN assignee found for {sd.isoformat()} under current constraints")

        schedule[sd.isoformat()]["PN"] = chosen
        if need_sat_pair:
            schedule[sat.isoformat()]["AN"] = chosen
    # Primary role assignment order (fill all PN/AN/W for month)
    primary_slots: List[Tuple[date, str]] = []
    for d in days:
        req = roles_required_for_day(d, push_days, observed_holidays)
        for role in ("PN", "AN", "W"):
            if role in req:
                # if already pre-assigned, skip
                if role in schedule[d.isoformat()]:
                    continue
                primary_slots.append((d, role))

    # Scoring for candidates (soft constraints):
    # - avoid repeating same role back-to-back (hard)
    # - avoid same weekday concentration for PN (soft; hard cap: max 1 Monday PN per person per month)
    # - push toward target counts per role
    # - avoid scheduling on hard-off
    pn_weekday_counts = {n: {i: 0 for i in range(7)} for n in roster}
    monday_pn_used = {n: 0 for n in roster}

    # Preload from already pre-assigned Sunday PN and any prefilled AN
    for d in days:
        roles = schedule[d.isoformat()]
        if "PN" in roles:
            pn_weekday_counts[roles["PN"]][weekday_sun0(d)] += 1
            if weekday_sun0(d) == 1:
                monday_pn_used[roles["PN"]] += 1

    role_counts = {r: {n: 0 for n in roster} for r in ("PN", "AN", "W")}
    for d in days:
        roles = schedule[d.isoformat()]
        for r in ("PN", "AN", "W"):
            if r in roles:
                role_counts[r][roles[r]] += 1
    def can_assign_primary(d: date, role: str, name: str) -> bool:
        if d in observed_holidays:
            return False
        if name in cons.hard_off.get(d, set()):
            return False
        if name in cons.avoid_roles.get((d, role), set()):
            return False
        if name in schedule[d.isoformat()].values():
            return False
        prev_role = last_role(name, d)
        if prev_role == role:
            return False
        if role == "PN" and weekday_sun0(d) == 1 and monday_pn_used[name] >= 1:
            return False
        if consecutive_streak(name, d) >= config.max_consecutive_days:
            return False
        if would_exceed_week_cap(d, role, name):
            return False
        if would_exceed_week_days_cap(d, name):
            return False
        if would_exceed_payperiod_cap(d, role, name):
            return False
        return True


    def score_primary(d: date, role: str, name: str) -> float:
        s = 0.0
        # prefer under-target
        t = targets.get(role, {}).get(name, 0)
        c = role_counts[role][name]
        s += (t - c) * 5.0

        # soften weekday concentration for PN
        if role == "PN":
            wd = weekday_sun0(d)
            s -= pn_weekday_counts[name][wd] * 3.0
            if wd == 1:
                s -= 50.0  # discourage more Monday PN strongly (hard already blocks >1)

        # prefer sales-order fairness only for extras (handled by targets), but slight bias:
        if name in sales_order:
            s += (len(sales_order) - sales_order.index(name)) * 0.05

        # randomness for variety
        s += rng.random() * 0.25
        return s

    # Backtracking for primaries (MRV-ish by choosing next slot with fewest candidates)
    start_time = datetime.now()
    nodes = 0
    last_fail_slot: Tuple[date, str] = (days[0], "")

    def candidates_for(slot: Tuple[date, str]) -> List[str]:
        d, role = slot
        cands = [n for n in roster if can_assign_primary(d, role, n)]
        cands.sort(key=lambda n: score_primary(d, role, n), reverse=True)
        return cands

    def select_next_slot(slots_left: List[Tuple[date, str]]) -> Tuple[int, Tuple[date, str], List[str]]:
        best_i = -1
        best_slot = None
        best_cands = None
        best_len = 10**9
        for i, sl in enumerate(slots_left):
            cands = candidates_for(sl)
            L = len(cands)
            if L < best_len:
                best_len = L
                best_i = i
                best_slot = sl
                best_cands = cands
                if L <= 1:
                    break
        return best_i, best_slot, best_cands

    def assign_primary(d: date, role: str, name: str):
        schedule[d.isoformat()][role] = name
        role_counts[role][name] += 1
        if role == "PN":
            wd = weekday_sun0(d)
            pn_weekday_counts[name][wd] += 1
            if wd == 1:
                monday_pn_used[name] += 1

    def unassign_primary(d: date, role: str, name: str):
        del schedule[d.isoformat()][role]
        role_counts[role][name] -= 1
        if role == "PN":
            wd = weekday_sun0(d)
            pn_weekday_counts[name][wd] -= 1
            if wd == 1:
                monday_pn_used[name] -= 1

    def solve_primary(slots_left: List[Tuple[date, str]]) -> bool:
        nonlocal nodes, last_fail_slot
        if not slots_left:
            return True
        idx, slot, cands = select_next_slot(slots_left)
        if not cands:
            return False
        d, role = slot
        # try best candidates first
        for name in cands:
            nodes += 1
            assign_primary(d, role, name)
            nxt = slots_left[:idx] + slots_left[idx+1:]
            if solve_primary(nxt):
                return True
            unassign_primary(d, role, name)
        return False

    ok = solve_primary(primary_slots)
    if not ok:
        raise RuntimeError(f"❌ Failed to assign primary roles. Stuck at {last_fail_slot[0].isoformat()} role={last_fail_slot[1]} with 0 candidates under current constraints.")

    # ---- BU backfill phase (fully relaxed caps per your latest instruction)
    # We still enforce:
    # - one person per day
    # - no same specific BU role back-to-back for same person (BU1->BU1 forbidden, BU1->BU2 ok)
    # - hard time off
    # - avoid_roles on BU* if specified (optional)
    bu_slots: List[Tuple[date, str]] = []
    for d in days:
        req = roles_required_for_day(d, push_days, observed_holidays)
        for role in ("BU1","BU2","BU3","BU4"):
            if role in req:
                if role in schedule[d.isoformat()]:
                    continue
                bu_slots.append((d, role))

    # track BU role last assigned to person yesterday (across boundary)
    def can_assign_bu(d: date, role: str, name: str) -> bool:
        if d in observed_holidays:
            return False
        if name in cons.hard_off.get(d, set()):
            return False
        if name in cons.avoid_roles.get((d, role), set()):
            return False
        if name in schedule[d.isoformat()].values():
            return False
        prev_role = last_role(name, d)
        if prev_role == role:
            return False
        if consecutive_streak(name, d) >= config.max_consecutive_days:
            return False
        if would_exceed_week_cap(d, role, name):
            return False
        if would_exceed_week_days_cap(d, name):
            return False
        if would_exceed_payperiod_cap(d, role, name):
            return False
        return True


    def score_bu(d: date, role: str, name: str) -> float:
        # prefer people with fewer total assignments that week to spread workload,
        # but since fully relaxed, this is just a soft preference.
        s = rng.random()
        # slight preference to keep Monday full-team if possible is handled by required roles already
        return s

    nodes2 = 0
    last_fail_slot_bu: Tuple[date, str] = (days[0], "")
    start2 = datetime.now()

    def candidates_for_bu(slot: Tuple[date, str]) -> List[str]:
        d, role = slot
        cands = [n for n in roster if can_assign_bu(d, role, n)]
        cands.sort(key=lambda n: score_bu(d, role, n), reverse=True)
        return cands

    def select_next_slot_bu(slots_left):
        best_i, best_slot, best_cands, best_len = -1, None, None, 10**9
        for i, sl in enumerate(slots_left):
            cands = candidates_for_bu(sl)
            L = len(cands)
            if L < best_len:
                best_len = L
                best_i, best_slot, best_cands = i, sl, cands
                if L <= 1:
                    break
        return best_i, best_slot, best_cands

    def solve_bu(slots_left: List[Tuple[date, str]]) -> bool:
        nonlocal nodes2, last_fail_slot_bu
        if not slots_left:
            return True
        idx, slot, cands = select_next_slot_bu(slots_left)
        if not cands:
            return False
        d, role = slot
        for name in cands:
            nodes2 += 1
            schedule[d.isoformat()][role] = name
            nxt = slots_left[:idx] + slots_left[idx+1:]
            if solve_bu(nxt):
                return True
            del schedule[d.isoformat()][role]
        return False

    ok2 = solve_bu(bu_slots)
    if not ok2:
        raise RuntimeError(f"❌ Failed to assign BU roles. Stuck at {last_fail_slot_bu[0].isoformat()} role={last_fail_slot_bu[1]} with 0 candidates under current constraints.")

    return schedule


# ----------------------------
# Excel export (no template)
# ----------------------------

def export_schedule_to_excel(schedule: Dict[str, Dict[str, str]],
    year: int,
    month: int,
    out_xlsx: str,
    roster: List[str],
    observed_holidays: Set[date],
    prev_sched: Optional[Dict[str, Dict[str, str]]] = None,
    payperiods: Optional[List[Tuple[date, date]]] = None,
    title: Optional[str] = None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = calendar.month_name[month]

    # Hide gridlines (the "lines" you were seeing)
    ws.sheet_view.showGridLines = False

    # Set a reasonable column width for B..V (21 columns)
    for col in range(2, 23):  # B..V
        ws.column_dimensions[get_column_letter(col)].width = 10

    # Set row heights for 3..42
    for r in range(3, GRID_WEEK_START_ROW + WEEK_HEIGHT * MAX_WEEKS):
        ws.row_dimensions[r].height = 18

    # Header merged B1:V1
    header = title or f"{calendar.month_name[month]} {year}"
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=22)
    c = ws.cell(1, 2, header)
    c.font = FONT_HEADER
    c.alignment = ALIGN_CENTER

    # Build grid: for each of 35 blocks (5*7), fill day number if in month else gray
    first = date(year, month, 1)
    offset = weekday_sun0(first)
    days_in_month = calendar.monthrange(year, month)[1]

    # Write each day into its block, and set background fills
    for week_index in range(MAX_WEEKS):
        R = GRID_WEEK_START_ROW + WEEK_HEIGHT * week_index
        for dow in range(7):
            col_start = SUN_START_COL + DAY_WIDTH * dow
            col_mid = col_start + 1
            col_end = col_start + 2

            cell_idx = week_index * 7 + dow
            daynum = cell_idx - offset + 1
            in_month = 1 <= daynum <= days_in_month

            # Determine fill for this block
            fill_hex = None
            if not in_month:
                fill_hex = NONMONTH_FILL
            else:
                d = date(year, month, daynum)
                if is_sunday(d):
                    fill_hex = SUNDAY_FILL

            fill = PatternFill("solid", fgColor=fill_hex) if fill_hex else None

            # Apply fill to all 24 cells of block
            for rr in range(R, R + WEEK_HEIGHT):
                for cc in range(col_start, col_start + DAY_WIDTH):
                    cell = ws.cell(rr, cc)
                    cell.font = FONT_MAIN
                    cell.alignment = ALIGN_CENTER
                    if fill:
                        cell.fill = fill

            if not in_month:
                continue

            d = date(year, month, daynum)
            # Day number in top-right
            ws.cell(R, col_end, daynum).font = FONT_MAIN
            ws.cell(R, col_end).alignment = Alignment(horizontal="right", vertical="top")

            # Labels
            ws.cell(R+1, col_start, "PN-").font = FONT_MAIN
            ws.cell(R+2, col_start, "AN-").font = FONT_MAIN
            if is_weekday(d):
                ws.cell(R+3, col_start, "W-").font = FONT_MAIN

            # Names from schedule
            roles = schedule.get(d.isoformat(), {})

            def put_name(row_off: int, role_key: str):
                name = roles.get(role_key, "")
                ws.cell(R + row_off, col_mid, name).font = FONT_MAIN

            # primaries
            put_name(1, "PN")
            if not is_sunday(d):
                put_name(2, "AN")
            if is_weekday(d):
                put_name(3, "W")

            # BU rows
            put_name(4, "BU1")
            put_name(5, "BU2")
            put_name(6, "BU3")
            put_name(7, "BU4")

            # Holiday: if observed holiday, you may want to merge and write text;
            # for February 2026 none, but leaving here for future:
            if d in observed_holidays:
                # Clear names/labels and write holiday text across the block
                for off in range(1, 8):
                    ws.cell(R+off, col_start, "").value = ""
                    ws.cell(R+off, col_mid, "").value = ""
                ws.merge_cells(start_row=R, start_column=col_start, end_row=R+7, end_column=col_end)
                ws.cell(R, col_start, "HOLIDAY").alignment = ALIGN_CENTER
                ws.cell(R, col_start).font = Font(name="Calibri", size=14, bold=True)

    # Summary tab
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2["A1"] = f"Summary: {calendar.month_name[month]} {year}"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True)

    roles_to_count = ["PN","AN","W","BU1","BU2","BU3","BU4"]
    # headers
    ws2["A3"] = "Name"
    for j, r in enumerate(roles_to_count, start=2):
        ws2.cell(3, j, r)
    ws2.cell(3, len(roles_to_count)+2, "Total Assignments")
    ws2.cell(3, len(roles_to_count)+3, "Total Hours")

    counts = {n: {r:0 for r in roles_to_count} for n in roster}
    hours = {n:0 for n in roster}

    for ds, roles in schedule.items():
        d = datetime.strptime(ds, "%Y-%m-%d").date()
        for r, n in roles.items():
            if n not in counts:
                continue
            if r in counts[n]:
                counts[n][r] += 1
            hours[n] += hours_for_role(d, r, observed_holidays)

    for i, n in enumerate(roster, start=4):
        ws2.cell(i, 1, n)
        tot = 0
        for j, r in enumerate(roles_to_count, start=2):
            v = counts[n][r]
            ws2.cell(i, j, v)
            tot += v
        ws2.cell(i, len(roles_to_count)+2, tot)
        ws2.cell(i, len(roles_to_count)+3, hours[n])

    # simple widths
    ws2.column_dimensions["A"].width = 14
    for col in range(2, len(roles_to_count)+4):
        ws2.column_dimensions[get_column_letter(col)].width = 14



    # Weekly Hours tab (Sun–Sat)
    ws3 = wb.create_sheet("Weekly Hours")
    ws3.sheet_view.showGridLines = False
    ws3["A1"] = f"Weekly Hours (Sun–Sat): {calendar.month_name[month]} {year}"
    ws3["A1"].font = Font(name="Calibri", size=14, bold=True)

    # Headers
    ws3["A3"] = "Week Start (Sun)"
    ws3["A3"].font = Font(name="Calibri", size=11, bold=True)
    for j, name in enumerate(roster, start=2):
        cell = ws3.cell(3, j, name)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = ALIGN_CENTER

    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    # Determine which Sunday-start weeks intersect this month
    week_starts = sorted({week_start_sun(d) for d in month_days(year, month)})

    def hours_on_day(dd: date, person: str) -> int:
        if dd in observed_holidays: # Use correct holiday set
            return 0
        if dd.month == month and dd.year == year:
            roles = schedule.get(dd.isoformat(), {})
        elif dd < first_day and prev_sched:
            roles = prev_sched.get(dd.isoformat(), {})
        else:
            roles = {}
        day_hours = 0
        for rr, nn in roles.items():
            if nn == person:
                # pass observed_holidays because if we are here, we are checking this year
                day_hours = max(day_hours, hours_for_role(dd, rr, observed_holidays))
        return day_hours

    red_fill = PatternFill("solid", fgColor="FFC7CE")

    row = 4
    for wsun in week_starts:
        ws3.cell(row, 1, wsun.isoformat())
        ws3.cell(row, 1).alignment = ALIGN_CENTER
        for j, person in enumerate(roster, start=2):
            tot = 0
            for i in range(7):
                dd = wsun + timedelta(days=i)
                # include carryover before month, ignore after month
                if dd > last_day:
                    continue
                tot += hours_on_day(dd, person)
            c = ws3.cell(row, j, tot)
            c.alignment = ALIGN_CENTER
            if tot > 40:
                c.fill = red_fill
        row += 1

    # Column widths
    ws3.column_dimensions["A"].width = 16
    for col in range(2, 2 + len(roster)):
        ws3.column_dimensions[get_column_letter(col)].width = 12
    

    # Pay Period Hours tab (optional, requires payperiods)
    if payperiods:
        ws4 = wb.create_sheet("Pay Period Hours")
        ws4.sheet_view.showGridLines = False
        ws4["A1"] = f"Pay Period Hours (Cap 80): {calendar.month_name[month]} {year}"
        ws4["A1"].font = Font(name="Calibri", size=14, bold=True)

        ws4["A3"] = "Pay Period"
        ws4["B3"] = "Start"
        ws4["C3"] = "End"
        for j, name in enumerate(roster, start=4):
            ws4.cell(3, j, name).font = Font(name="Calibri", size=11, bold=True)
            ws4.cell(3, j).alignment = ALIGN_CENTER

        red_fill = PatternFill("solid", fgColor="FFC7CE")
        first_day = date(year, month, 1)

        def pp_hours(person: str, pp: Tuple[date, date]) -> int:
            s, e = pp
            tot = 0
            dd = s
            while dd <= e:
                if dd.month == month and dd.year == year:
                    roles = schedule.get(dd.isoformat(), {})
                elif dd < first_day and prev_sched:
                    roles = prev_sched.get(dd.isoformat(), {})
                else:
                    roles = {}
                day_hours = 0
                for rr, nn in roles.items():
                    if nn == person:
                        # Use passed holidays if safe, otherwise calc
                        hols = observed_holidays if dd.year == year else observed_holidays_for_year(dd.year)
                        day_hours = max(day_hours, hours_for_role(dd, rr, hols))
                tot += day_hours
                dd += timedelta(days=1)
            return tot

        row = 4
        for idx_pp, (s, e) in enumerate(payperiods, start=1):
            ws4.cell(row, 1, f"PP{idx_pp}")
            ws4.cell(row, 2, s.isoformat())
            ws4.cell(row, 3, e.isoformat())
            for j, person in enumerate(roster, start=4):
                tot = pp_hours(person, (s, e))
                c = ws4.cell(row, j, tot)
                c.alignment = ALIGN_CENTER
                if tot > 80:
                    c.fill = red_fill
            row += 1

        for col in range(1, 4 + len(roster)):
            ws4.column_dimensions[get_column_letter(col)].width = 14
    
    wb.save(out_xlsx)


# ----------------------------
# Main
# ----------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", required=True, help="YYYY-MM, e.g. 2026-02")
    ap.add_argument("--timeoff-xlsx", required=True)
    ap.add_argument("--timeoff-sheet", required=True)
    ap.add_argument("--prev-schedule", required=True, help="Previous month schedule JSON (exported from Excel)")
    ap.add_argument("--sales-sheet", default="Sales Ranking")
    ap.add_argument("--out-json", default=None)
    ap.add_argument("--out-xlsx", default=None)
    ap.add_argument("--seed", type=int, default=11)
    ap.add_argument("--roster-json", default=None, help="Optional JSON file with roster list (overrides DEFAULT_ROSTER / Sales Ranking)")
    ap.add_argument("--payperiods-json", default=None, help="Optional JSON file with pay period boundaries (for 80h cap + audit)")
    args = ap.parse_args()

    year, month = parse_month_arg(args.month)

    roster = DEFAULT_ROSTER[:]  # display names
    # constraints
    timeoff = load_timeoff_from_xlsx(args.timeoff_xlsx, args.timeoff_sheet)
    cons = compile_constraints(timeoff)

    # sales ranking from TimeOff.xlsx (Option B)
    sales_order = load_sales_ranking_from_timeoff_xlsx(
        xlsx_path=args.timeoff_xlsx,
        sheet_name=args.sales_sheet,
        year=year,
        month=month,
        roster_names=roster,
    )

    prev = load_schedule_json(args.prev_schedule)

    cfg = BuildConfig(year=year, month=month, roster=roster, sales_order=sales_order, rng_seed=args.seed)

    payperiods = load_payperiods_from_json(args.payperiods_json) if args.payperiods_json else None
    sched = build_schedule(cfg, prev, cons, payperiods)

    out_json = args.out_json or str(Path(args.prev_schedule).with_name(f"{year}_{month:02d}_schedule.json"))
    out_xlsx = args.out_xlsx or str(Path(args.prev_schedule).with_name(f"{calendar.month_name[month]}_{year}_Schedule.xlsx"))

    Path(out_json).write_text(json.dumps(sched, indent=2), encoding="utf-8")
    print(f"✅ Wrote schedule JSON → {out_json}")

    export_schedule_to_excel(
        sched,
        year,
        month,
        out_xlsx,
        roster,
        observed_holidays_for_year(year),
        prev,
        payperiods,
    )
    print(f"✅ Wrote schedule Excel → {out_xlsx}")

if __name__ == "__main__":
    main()
