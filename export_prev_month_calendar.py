#!/usr/bin/env python3
"""
export_prev_month_calendar.py

Deterministic exporter for your Harpeth Hills calendar template.

Grid rules (LOCKED FROM YOUR SPEC):
- Calendar grid spans rows 3..42 (5 week rows * 8 rows per week).
- Each day is 3 columns wide x 8 rows tall (24 cells).
- Day number is in the top-right cell of the day block.
- For a day block with week_start_row = R and columns start/mid/end:
    Day number: (col_end, R)
    PN name:    (col_mid, R+1)
    AN name:    (col_mid, R+2)
    W  name:    (col_mid, R+3)
    BU1 name:   (col_mid, R+4)
    BU2 name:   (col_mid, R+5)
    BU3 name:   (col_mid, R+6)
- Labels exist but are NOT required for export (we ignore them).

Columns by weekday (Sun=0..Sat=6):
- Sunday    B-D  (start=B)
- Monday    E-G
- Tuesday   H-J
- Wednesday K-M
- Thursday  N-P
- Friday    Q-S
- Saturday  T-V

Holiday text may appear (even merged). We treat any non-roster text as "no assignment".
"""

from __future__ import annotations

import argparse
import calendar
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl


# ----------------------------
# Month/year parsing (B1 merged is fine)
# ----------------------------

MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

def parse_month_year(v) -> Tuple[int, int]:
    """
    Accepts:
      - "January 2026", "Jan 2026"
      - "03/2026"
      - Excel date/datetime objects
    Returns (year, month).
    """
    if v is None or (isinstance(v, str) and not v.strip()):
        raise ValueError("B1 header is empty; cannot determine month/year.")

    if isinstance(v, datetime):
        return v.year, v.month
    if isinstance(v, date):
        return v.year, v.month

    s = str(v).strip()

    m = re.match(r"^\s*(\d{1,2})\s*/\s*(\d{4})\s*$", s)
    if m:
        mm = int(m.group(1))
        yy = int(m.group(2))
        if 1 <= mm <= 12:
            return yy, mm
        raise ValueError(f"Invalid month in header: {s!r}")

    m = re.search(r"([A-Za-z]+)\s+(\d{4})", s)
    if m:
        mon = m.group(1).lower()
        yy = int(m.group(2))
        if mon in MONTHS:
            return yy, MONTHS[mon]
        key = mon[:4] if mon.startswith("sept") else mon[:3]
        if key in MONTHS:
            return yy, MONTHS[key]
        raise ValueError(f"Unrecognized month name in header: {m.group(1)!r}")

    raise ValueError(f"Could not parse month/year from header value: {v!r}")


# ----------------------------
# Roster + aliases
# ----------------------------

@dataclass(frozen=True)
class RosterPerson:
    person_id: str
    display: str
    aliases: Tuple[str, ...]

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip()).lower()

def load_roster(path: str) -> Tuple[Dict[str, RosterPerson], Dict[str, str]]:
    """
    Roster JSON format:

    {
      "mark_bryant": {"display":"Mark", "aliases":["Mark","Mark B","MB"]},
      "will_sutherland": {"display":"Will", "aliases":["Will","William","WS"]}
    }

    Returns:
      - people_by_id
      - alias_to_id (normalized alias -> person_id)
    """
    raw = json.loads(Path(path).read_text(encoding="utf-8"))
    people: Dict[str, RosterPerson] = {}
    alias_to_id: Dict[str, str] = {}

    for pid, data in raw.items():
        display = str(data["display"]).strip()
        aliases = tuple(str(a).strip() for a in data.get("aliases", []))
        if not aliases:
            aliases = (display,)
        rp = RosterPerson(pid, display, aliases)
        people[pid] = rp
        for a in aliases:
            alias_to_id[_norm(a)] = pid

    return people, alias_to_id

def normalize_to_display(cell_value, people: Dict[str, RosterPerson], alias_to_id: Dict[str, str]) -> Optional[str]:
    """
    Returns display name if cell_value matches a roster alias, else None.
    This safely ignores holiday greetings and other text.
    """
    if cell_value is None:
        return None
    if isinstance(cell_value, (int, float, bool)):
        return None
    s = str(cell_value).strip()
    if not s:
        return None
    pid = alias_to_id.get(_norm(s))
    if not pid:
        return None
    return people[pid].display


# ----------------------------
# Geometry: row/col math
# ----------------------------

GRID_WEEK_START_ROW = 3          # Week 1 starts at row 3
WEEK_HEIGHT = 8                  # 8 rows per week block
DAY_WIDTH = 3                    # 3 columns per day block
SUN_START_COL = 2                # Column B = 2 in openpyxl
MAX_WEEKS = 5                    # grid supports 5 weeks (rows 3..42)

def weekday_sun0(d: date) -> int:
    """
    Python: Monday=0..Sunday=6
    Convert to Sunday=0..Saturday=6
    """
    return (d.weekday() + 1) % 7

def cell_for_day(year: int, month: int, day: int) -> Tuple[int, int, int, int]:
    """
    Returns (week_start_row R, col_start, col_mid, col_end) for the given date.
    """
    d = date(year, month, day)
    first = date(year, month, 1)
    offset = weekday_sun0(first)            # where day 1 lands in week row (Sun0)
    idx = offset + (day - 1)                # index into the 35-cell grid
    week_index = idx // 7                   # 0..4
    dow = idx % 7                           # 0..6 (Sun0)

    if week_index < 0 or week_index >= MAX_WEEKS:
        raise ValueError(f"Date {d} falls outside 5-week grid (unexpected).")

    R = GRID_WEEK_START_ROW + WEEK_HEIGHT * week_index

    col_start = SUN_START_COL + DAY_WIDTH * dow
    col_mid = col_start + 1
    col_end = col_start + 2
    return R, col_start, col_mid, col_end


# ----------------------------
# Export
# ----------------------------

def export_month_from_calendar_xlsx(
    xlsx_path: str,
    roster_path: str,
    out_json_path: str,
    sheet: Optional[str] = None,
    include_empty_days: bool = True,
    validate_day_numbers: bool = True,
):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]

    year, month = parse_month_year(ws["B1"].value)
    days_in_month = calendar.monthrange(year, month)[1]

    people, alias_to_id = load_roster(roster_path)

    schedule: Dict[str, Dict[str, str]] = {}

    for day in range(1, days_in_month + 1):
        R, col_start, col_mid, col_end = cell_for_day(year, month, day)

        # Optional validation: day number is in top-right (col_end, R)
        if validate_day_numbers:
            v = ws.cell(R, col_end).value
            # Allow blank/merged holiday text, but if it's numeric and wrong, warn
            if isinstance(v, (int, float)) and int(v) != day:
                print(f"⚠️  Day number mismatch at {ws.cell(R,col_end).coordinate}: found {v!r}, expected {day}")

        roles: Dict[str, str] = {}

        # Read names from fixed cells; ignore non-roster text
        pn = normalize_to_display(ws.cell(R + 1, col_mid).value, people, alias_to_id)
        an = normalize_to_display(ws.cell(R + 2, col_mid).value, people, alias_to_id)
        w  = normalize_to_display(ws.cell(R + 3, col_mid).value, people, alias_to_id)

        bu1 = normalize_to_display(ws.cell(R + 4, col_mid).value, people, alias_to_id)
        bu2 = normalize_to_display(ws.cell(R + 5, col_mid).value, people, alias_to_id)
        bu3 = normalize_to_display(ws.cell(R + 6, col_mid).value, people, alias_to_id)

        if pn: roles["PN"] = pn
        if an: roles["AN"] = an
        if w:  roles["W"]  = w
        if bu1: roles["BU1"] = bu1
        if bu2: roles["BU2"] = bu2
        if bu3: roles["BU3"] = bu3

        if roles or include_empty_days:
            schedule[date(year, month, day).isoformat()] = roles

    Path(out_json_path).write_text(json.dumps(schedule, indent=2), encoding="utf-8")
    print(f"✅ Exported {len(schedule)} days for {year}-{month:02d} → {out_json_path}")


# ----------------------------
# CLI
# ----------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="Filled calendar XLSX (previous month)")
    ap.add_argument("--roster", required=True, help="Roster JSON with aliases")
    ap.add_argument("--out", required=True, help="Output JSON path")
    ap.add_argument("--sheet", default=None, help="Worksheet name (default: first)")
    ap.add_argument("--no-include-empty-days", action="store_true", help="Do not include empty days")
    ap.add_argument("--no-validate-daynums", action="store_true", help="Do not validate top-right day numbers")
    args = ap.parse_args()

    export_month_from_calendar_xlsx(
        xlsx_path=args.inp,
        roster_path=args.roster,
        out_json_path=args.out,
        sheet=args.sheet,
        include_empty_days=(not args.no_include_empty_days),
        validate_day_numbers=(not args.no_validate_daynums),
    )

if __name__ == "__main__":
    main()
