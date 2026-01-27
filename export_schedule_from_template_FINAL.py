#!/usr/bin/env python3
"""
export_schedule_from_template_FINAL.py

Robust exporter to convert a filled monthly calendar template into JSON:
{
  "YYYY-MM-DD": {"PN":"Name", "AN":"Name", "W":"Name", "BU1":"Name", ...},
  ...
}

Hardening improvements vs earlier versions:
- Header parsing accepts string or Excel date/datetime; configurable header cell.
- Day detection restricted to calendar grid region (default B4:V40) to avoid random 1–31.
- Invalid days of month are skipped with warning (or error with --strict).
- Duplicate day anchors warn (or error with --strict).
- Role scanning is bounded to a fixed block below the day cell and stops at next day cell.
- Sheet selection supported.
"""

from __future__ import annotations

import argparse
import calendar
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Tuple, Optional, List

import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


# ----------------------------
# Parsing helpers
# ----------------------------

MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

def parse_month_year(value) -> Tuple[int, int]:
    """
    Accepts:
      - "January 2026", "Jan 2026"
      - "03/2026", "3/2026"
      - datetime/date objects (Excel often stores header as date)
    Returns (year, month).
    """
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError("Header cell is empty; cannot determine month/year.")

    # If header is an Excel date/datetime, use it directly
    if isinstance(value, datetime):
        return value.year, value.month
    if isinstance(value, date):
        return value.year, value.month

    s = str(value).strip()

    # Try "MM/YYYY"
    m = re.match(r"^\s*(\d{1,2})\s*/\s*(\d{4})\s*$", s)
    if m:
        month = int(m.group(1))
        year = int(m.group(2))
        if 1 <= month <= 12:
            return year, month
        raise ValueError(f"Invalid month in header: {s}")

    # Try "Month YYYY"
    m = re.search(r"([A-Za-z]+)\s+(\d{4})", s)
    if m:
        mon_raw = m.group(1).lower()
        year = int(m.group(2))
        mon_key = mon_raw[:4] if mon_raw.startswith("sept") else mon_raw[:3]
        # support full names too
        if mon_raw in MONTHS:
            month = MONTHS[mon_raw]
        elif mon_key in MONTHS:
            month = MONTHS[mon_key]
        else:
            raise ValueError(f"Unrecognized month name in header: {m.group(1)}")
        return year, month

    raise ValueError(f"Could not parse month/year from header value: {value!r}")


def a1_to_rowcol(a1: str) -> Tuple[int, int]:
    col_letters, row = coordinate_from_string(a1.upper())
    return int(row), column_index_from_string(col_letters)


def clamp_month_day(year: int, month: int, daynum: int) -> Optional[date]:
    """Return date if daynum valid for year/month, else None."""
    max_day = calendar.monthrange(year, month)[1]
    if 1 <= daynum <= max_day:
        return date(year, month, daynum)
    return None


def cell_as_int(v) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        t = v.strip()
        if t.isdigit():
            return int(t)
    return None


# ----------------------------
# Exporter
# ----------------------------

@dataclass
class ExportConfig:
    header_cell: str = "B1"
    grid_min: str = "B4"
    grid_max: str = "V40"
    # relative offsets from the day number cell:
    date_col_offset: int = 0
    label_col_offset: int = 0
    name_col_offset: int = 1
    max_scan: int = 40
    include_empty_days: bool = False
    strict: bool = False


PRIMARY_LABELS = {"PN", "AN", "W", "PH"}  # W only in your world, but keep PH safe if appears
BU_PREFIX = "BU"

def looks_like_day_block(ws, r: int, c: int, max_scan: int) -> bool:
    """
    Heuristic: below the day cell, we should see:
    - at least one primary label like "PN" or "AN" or "W"
      OR
    - multiple name-like entries
    This avoids picking up random integers in the grid.
    """
    labels_found = 0
    names_found = 0
    for rr in range(r + 1, min(r + 1 + max_scan, ws.max_row + 1)):
        label = ws.cell(rr, c).value
        name = ws.cell(rr, c + 1).value
        if isinstance(label, str):
            t = label.strip().upper().replace(":", "")
            # match "PN -" or "PN"
            t = t.split()[0]
            if t in PRIMARY_LABELS or t.startswith(BU_PREFIX):
                labels_found += 1
        if name not in (None, ""):
            names_found += 1
        if labels_found >= 1:
            return True
    return names_found >= 2


def find_day_cells(ws, year: int, month: int, cfg: ExportConfig) -> List[Tuple[date, int, int]]:
    """
    Scans only within the calendar grid region and returns list of (date, row, col) anchors.
    """
    r1, c1 = a1_to_rowcol(cfg.grid_min)
    r2, c2 = a1_to_rowcol(cfg.grid_max)
    found: Dict[int, Tuple[int, int]] = {}

    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            v = ws.cell(r, c).value
            daynum = cell_as_int(v)
            if not daynum:
                continue
            d = clamp_month_day(year, month, daynum)
            if d is None:
                msg = f"⚠️  Found invalid day '{daynum}' for {year}-{month:02d} at {ws.cell(r,c).coordinate}"
                if cfg.strict:
                    raise ValueError(msg)
                else:
                    print(msg)
                    continue

            # ensure it is actually a day block
            if not looks_like_day_block(ws, r, c, cfg.max_scan):
                continue

            if daynum in found:
                msg = f"⚠️  Duplicate day '{daynum}' anchor at {ws.cell(r,c).coordinate} (already have {ws.cell(*found[daynum]).coordinate}). Keeping first."
                if cfg.strict:
                    raise ValueError(msg)
                else:
                    print(msg)
                    continue

            found[daynum] = (r, c)

    anchors: List[Tuple[date, int, int]] = []
    for daynum, (r, c) in sorted(found.items(), key=lambda x: x[0]):
        anchors.append((date(year, month, daynum), r, c))
    return anchors


def parse_roles_for_day(ws, anchor_r: int, anchor_c: int, cfg: ExportConfig, next_anchor_row: Optional[int]) -> Dict[str, str]:
    """
    Parse roles in the block beneath the day cell.
    We scan downward until:
      - we reach max_scan rows, OR
      - we reach the next day anchor row (same column region), OR
      - we hit a sustained empty stretch
    Primary roles: label in (PN, AN, W, PH) with a name on same row in adjacent col.
    BU roles: name-only rows after primaries become BU1..BUk, bounded to the day block.
    """
    roles: Dict[str, str] = {}
    start = anchor_r + 1
    end_limit = min(ws.max_row, anchor_r + cfg.max_scan)
    if next_anchor_row is not None:
        end_limit = min(end_limit, next_anchor_row - 1)

    # First pass: read primary labeled rows
    last_primary_row = None
    for r in range(start, end_limit + 1):
        label = ws.cell(r, anchor_c + cfg.label_col_offset).value
        name = ws.cell(r, anchor_c + cfg.name_col_offset).value

        if label is None and name is None:
            continue

        if isinstance(label, str):
            t = label.strip().upper().replace(":", "")
            # common formats: "PN -", "PN", "PN-"
            t = t.split()[0].replace("-", "")
            if t in PRIMARY_LABELS:
                if name not in (None, ""):
                    roles[t] = str(name).strip()
                last_primary_row = r
                continue

        # If it looks like primary label row but missing name, still set last_primary_row
        if isinstance(label, str):
            t2 = label.strip().upper()
            if any(t2.startswith(x) for x in ["PN", "AN", "W", "PH"]):
                last_primary_row = r

    # Second pass: BU rows — start right after last primary row if found,
    # otherwise start after anchor (but still bounded)
    bu_start = (last_primary_row + 1) if last_primary_row else start

    bu_idx = 1
    empty_streak = 0
    for r in range(bu_start, end_limit + 1):
        label = ws.cell(r, anchor_c + cfg.label_col_offset).value
        name = ws.cell(r, anchor_c + cfg.name_col_offset).value

        # Stop conditions:
        # - If we hit a new labeled primary block, stop BU scan
        if isinstance(label, str):
            t = label.strip().upper().replace(":", "")
            t = t.split()[0].replace("-", "")
            if t in PRIMARY_LABELS:
                break

        if name in (None, ""):
            empty_streak += 1
            if empty_streak >= 4:
                break
            continue

        empty_streak = 0
        # If label column has obvious non-empty note text, skip counting it as BU
        if label not in (None, ""):
            # If it explicitly says BU, accept; else treat as a note and stop
            if isinstance(label, str) and label.strip().upper().startswith("BU"):
                pass
            else:
                # likely notes creeping in
                break

        roles[f"BU{bu_idx}"] = str(name).strip()
        bu_idx += 1

    return roles


def export_schedule(in_path: str, out_path: str, sheet: Optional[str], cfg: ExportConfig):
    wb = openpyxl.load_workbook(in_path, data_only=True)

    ws = wb[sheet] if sheet else wb.worksheets[0]

    # header parsing
    hr, hc = a1_to_rowcol(cfg.header_cell)
    header_val = ws.cell(hr, hc).value
    try:
        year, month = parse_month_year(header_val)
    except Exception as e:
        raise ValueError(f"Header parse failed at {cfg.header_cell} value={header_val!r}. Error: {e}") from e

    anchors = find_day_cells(ws, year, month, cfg)
    if not anchors:
        raise ValueError("No day cells found in the configured grid region. Check --grid-min/--grid-max.")

    # build mapping for next anchor row to bound scanning
    anchor_rows_by_day = [(d, r, c) for (d, r, c) in anchors]
    schedule: Dict[str, Dict[str, str]] = {}

    for idx, (d, r, c) in enumerate(anchor_rows_by_day):
        next_r = anchor_rows_by_day[idx + 1][1] if idx + 1 < len(anchor_rows_by_day) else None
        roles = parse_roles_for_day(ws, r, c, cfg, next_r)

        if roles or cfg.include_empty_days:
            schedule[d.isoformat()] = roles

    Path(out_path).write_text(json.dumps(schedule, indent=2), encoding="utf-8")
    print(f"✅ Exported {len(schedule)} day entries for {year}-{month:02d} → {out_path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="Input XLSX template")
    ap.add_argument("--out", dest="out", required=True, help="Output JSON path")
    ap.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet)")

    ap.add_argument("--header-cell", default="B1", help="A1 address of month header (default B1)")
    ap.add_argument("--grid-min", default="B4", help="A1 min cell of calendar grid (default B4)")
    ap.add_argument("--grid-max", default="V40", help="A1 max cell of calendar grid (default V40)")
    ap.add_argument("--max-scan", type=int, default=40, help="Max rows scanned under each day cell (default 40)")
    ap.add_argument("--include-empty-days", action="store_true", help="Include days with no roles as {}")
    ap.add_argument("--strict", action="store_true", help="Treat warnings as errors")

    args = ap.parse_args()

    cfg = ExportConfig(
        header_cell=args.header_cell,
        grid_min=args.grid_min,
        grid_max=args.grid_max,
        max_scan=args.max_scan,
        include_empty_days=args.include_empty_days,
        strict=args.strict,
    )

    export_schedule(args.inp, args.out, args.sheet, cfg)


if __name__ == "__main__":
    main()
